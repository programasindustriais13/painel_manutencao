import os
import sys
import django

# Set up Django environment
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "maintenance_project.settings")
django.setup()

from maintenance.models import Sector, Machine

def get_sector_by_mapping(tag, name):
    """
    Infece o setor correspondente com base no TAG e no Nome do Equipamento,
    caso a planilha não tenha a coluna específica de Setor.
    """
    tag_lower = tag.lower().strip()
    name_lower = name.lower().strip()
    
    # 1. Estamparia (Prensas)
    if (any(x in tag_lower for x in ['pb', 'pr', 'lbp', 'pt']) or 
            'prensa' in name_lower or 'blader' in name_lower or 'bladder' in name_lower):
        return 'Estamparia'
        
    # 2. Pintura
    if 'pintura' in name_lower or 'desmoldante' in name_lower:
        return 'Pintura'
        
    # 3. Montagem Mecânica (Construtoras, Montagem)
    if 'cst' in tag_lower or 'construtora' in name_lower or 'montagem' in name_lower:
        return 'Montagem Mecânica'
        
    # 4. Usinagem (Tornos, CNCs, Usinagem)
    if 'cnc' in tag_lower or any(x in name_lower for x in ['torno', 'usinagem', 'centro de usinagem', 'extrusora']):
        return 'Usinagem'
        
    # 5. Logística (Empilhadeiras, Pontes)
    if 'ets' in tag_lower or 'ponte' in name_lower or any(x in name_lower for x in ['empilhadeira', 'logistica', 'logística', 'talha']):
        return 'Logística'
        
    # 6. Utilidades / Manutenção Geral (Caldeiras, Compressores, Geradores, Reservatórios, etc.)
    return 'Utilidades'

def map_criticidade(val):
    """
    Mapeia a criticidade da planilha para as opções do modelo: 'BAIXA', 'MEDIA', 'ALTA'.
    """
    if not val:
        return 'BAIXA'
    val_lower = str(val).lower().strip()
    if 'alta' in val_lower or 'grave' in val_lower or 'urgente' in val_lower:
        return 'ALTA'
    elif 'media' in val_lower or 'média' in val_lower or 'med' in val_lower:
        return 'MEDIA'
    else:
        return 'BAIXA'

def populate():
    import openpyxl
    
    excel_path = 'TAG DAS MÁQUINAS.xlsx'
    if not os.path.exists(excel_path):
        print(f"Erro: O arquivo '{excel_path}' não foi encontrado na raiz do projeto.")
        sys.exit(1)
        
    print(f"Lendo planilha: '{excel_path}'...")
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"Erro ao abrir arquivo Excel: {e}")
        sys.exit(1)
        
    if 'Cronograma 2026' not in wb.sheetnames:
        print("Erro: A aba 'Cronograma 2026' não foi encontrada na planilha.")
        sys.exit(1)
        
    sheet = wb['Cronograma 2026']
    
    # Procurar a linha de cabeçalho (esperado na linha 11)
    header_row_index = None
    headers = []
    
    for r_idx, r in enumerate(sheet.iter_rows(max_row=30), 1):
        row_vals = [cell.value for cell in r]
        # Verifica se esta linha contém os cabeçalhos cruciais
        if any(val is not None and str(val).strip().lower() in ['tag', 'nome do equipamento'] for val in row_vals):
            headers = [str(x).strip() if x is not None else '' for x in row_vals]
            header_row_index = r_idx
            break
            
    if not header_row_index:
        print("Aviso: Cabeçalhos ('TAG', 'Nome do Equipamento') não detectados nas primeiras 30 linhas. Usando mapeamento fixo padrão.")
        # Mapeamento padrão de colunas (0: EQU, 1: TAG, 2: Nome, 3: Criticidade)
        tag_col = 1
        name_col = 2
        crit_col = 3
        sector_col = None
        start_row = 13
    else:
        print(f"Cabeçalho encontrado na linha {header_row_index}: {headers[:10]}...")
        tag_col = next((i for i, h in enumerate(headers) if 'tag' in h.lower()), 1)
        name_col = next((i for i, h in enumerate(headers) if 'equipamento' in h.lower() or 'nome' in h.lower()), 2)
        crit_col = next((i for i, h in enumerate(headers) if 'criticidade' in h.lower() or 'prioridade' in h.lower() or 'crit' in h.lower()), 3)
        sector_col = next((i for i, h in enumerate(headers) if 'setor' in h.lower() or 'área' in h.lower() or 'linha' in h.lower()), None)
        start_row = header_row_index + 1
        
    print(f"Mapeamento de Colunas - TAG: col {tag_col}, Nome: col {name_col}, Criticidade: col {crit_col}, Setor: {f'col {sector_col}' if sector_col is not None else 'Inferido dinamicamente'}")
    
    created_sectors = 0
    created_machines = 0
    skipped_machines = 0
    
    # Iterar sobre as linhas de dados
    for r_idx, r in enumerate(sheet.iter_rows(min_row=start_row), start_row):
        row_vals = [cell.value for cell in r]
        if not row_vals or len(row_vals) <= max(tag_col, name_col):
            continue
            
        # O id de ordenação/EQU está na coluna 0. Ignoramos linhas completamente vazias
        equ_val = row_vals[0]
        if equ_val is None:
            continue
            
        tag = str(row_vals[tag_col]).strip() if row_vals[tag_col] is not None else ''
        name = str(row_vals[name_col]).strip() if row_vals[name_col] is not None else ''
        criticidade_raw = row_vals[crit_col] if len(row_vals) > crit_col else 'BAIXA'
        
        if not tag or not name or tag.lower() == 'tag':
            continue
            
        # Determinar o setor
        if sector_col is not None and len(row_vals) > sector_col and row_vals[sector_col] is not None:
            sector_name = str(row_vals[sector_col]).strip()
        else:
            sector_name = get_sector_by_mapping(tag, name)
            
        # Mapeamento de prioridade
        criticidade = map_criticidade(criticidade_raw)
        
        # 1. Tratar/Criar o Setor
        sector_obj, created_sec = Sector.objects.get_or_create(nome=sector_name)
        if created_sec:
            created_sectors += 1
            print(f"[Setor] Criado novo setor: '{sector_name}'")
            
        # 2. Tratar/Criar a Máquina
        # Usamos o nome limpo e também verificamos se já existe
        machine_name = f"{name} ({tag})" if tag else name
        
        # Também verificamos se já existe por nome exato (ou incluindo a tag para evitar colisões de nomes idênticos em setores diferentes)
        # Vamos manter o nome simples, mas se já existir, não criamos
        if Machine.objects.filter(nome=name).exists() or Machine.objects.filter(nome=machine_name).exists():
            skipped_machines += 1
            continue
            
        # Criar a máquina associada ao setor
        Machine.objects.create(
            nome=name,
            setor=sector_obj,
            criticidade=criticidade
        )
        created_machines += 1
        
    print("\n--- Relatório de Importação ---")
    print(f"Novos Setores criados: {created_sectors}")
    print(f"Novas Máquinas criadas: {created_machines}")
    print(f"Máquinas ignoradas (já existentes): {skipped_machines}")
    print("---------------------------------")

if __name__ == "__main__":
    populate()
