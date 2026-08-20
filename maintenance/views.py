from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse, FileResponse, Http404, HttpResponseForbidden, JsonResponse
from functools import wraps
import json
import io
import datetime
import os
import mimetypes
from django.urls import reverse
from django.db import models
from django.db.models import Prefetch, Q, Case, When, Value, IntegerField
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


from .models import (
    Sector, Machine, Technician, Allocation, 
    HistoricoPausa, HistoricoEscala, WhatsAppGroup, AllocationProgressUpdate,
    OrdemServico, OrdemServicoPeca
)
from .forms import (
    SectorForm, MachineForm, TechnicianForm,
    StartServiceForm, PauseServiceForm, FinishServiceForm,
    OrdemServicoCreateForm
)



# ─────────────────────────────────────────────────────────────────────────────
# Helpers de permissão de grupo
# ─────────────────────────────────────────────────────────────────────────────

def _user_is_operador(user):
    """Retorna True apenas para Operadores/Administradores (grupo 'Operadores', superuser, staff, ou grupo legado 'Operador')."""
    if not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    return user.groups.filter(name__in=['Operadores', 'Operador']).exists()


def _user_is_lider_ou_operador(user):
    """Retorna True para Operadores e Técnicos Líderes (grupos 'Operadores', 'Tecnicos_Lideres', 'Operador', superuser, staff)."""
    if _user_is_operador(user):
        return True
    return user.groups.filter(name='Tecnicos_Lideres').exists()


def _get_technician_proprio(user):
    """Retorna o Técnico vinculado ao usuário, ou None se não houver."""
    try:
        return user.technician_profile
    except Exception:
        return None


def _user_can_access_maintenance(user):
    """Retorna True se for superuser, staff, pertencer aos grupos de manutenção ou possuir technician_profile."""
    if not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    if user.groups.filter(name__in=['Operadores', 'Operador', 'Tecnicos_Lideres', 'Tecnicos']).exists():
        return True
    if _get_technician_proprio(user):
        return True
    return False


def _user_can_access_production(user):
    """Retorna True se for superuser, staff, pertencer ao grupo 'Liderança de Produção', 'Operadores', 'Operador' ou grupo PCP."""
    if not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    if user.groups.filter(name__in=['Liderança de Produção', 'Operadores', 'Operador', 'PCP']).exists():
        return True
    if user.has_perm('production.add_productiontarget') or user.has_perm('production.view_productiontarget'):
        return True
    return False


def _user_has_dual_access(user):
    """Retorna True se o usuário tem permissão para AMBOS os módulos (Manutenção e Produção)."""
    return _user_can_access_maintenance(user) and _user_can_access_production(user)


def _user_has_maintenance_access(user):
    """Alias para compatibilidade retroativa com verificações existentes."""
    return _user_can_access_maintenance(user)


def _user_can_create_os(user):
    """
    Retorna True se o usuário pode abrir novas Ordens de Serviço:
    Operadores, Técnicos Líderes, Liderança de Produção, PCP, Superuser ou Staff.
    """
    if not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    if user.groups.filter(name__in=[
        'Operadores', 'Operador', 'Tecnicos_Lideres', 
        'Liderança de Produção', 'Lideres_Producao', 'Lideranca_Producao', 'Producao', 'PCP'
    ]).exists():
        return True
    tech = _get_technician_proprio(user)
    if tech and tech.perfil in ['OPERADOR', 'TECNICO_LIDER']:
        return True
    return False


def os_creation_required(view_func):
    """Decorator que protege a tela de abertura de novas Ordens de Serviço."""
    @login_required
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if _user_can_create_os(request.user):
            return view_func(request, *args, **kwargs)
        messages.error(request, "Acesso restrito. Você não possui permissão para abrir novas Ordens de Serviço.")
        return redirect('technician_management')
    return wrapper



# Decorator para views que exigem Operador/Admin COMPLETO (cadastros, etc.).
# Redireciona usuários sem acesso para /management/ com alerta.
def operador_required(view_func):
    @login_required
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if _user_is_operador(request.user):
            return view_func(request, *args, **kwargs)
        if request.user.groups.filter(name="Liderança de Produção").exists() and not _user_can_access_maintenance(request.user):
            messages.error(request, "Acesso restrito. Esta seção não está disponível para o seu perfil.")
            return redirect("production:dashboard")
        messages.error(
            request,
            "Acesso restrito a Operadores/Administradores. Esta seção não está disponível para o seu perfil."
        )
        return redirect('technician_management')
    return wrapper


# Decorator para /dashboard/ e exportação Excel.
# Permite: Operadores e Tecnicos_Lideres. Bloqueia outros (redireciona para /management/).
def lider_ou_operador_required(view_func):
    @login_required
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if _user_is_lider_ou_operador(request.user):
            return view_func(request, *args, **kwargs)
        if request.user.groups.filter(name="Liderança de Produção").exists() and not _user_can_access_maintenance(request.user):
            messages.error(request, "Acesso restrito. Esta página requer perfil de Técnico Líder ou superior.")
            return redirect("production:dashboard")
        messages.error(
            request,
            "Acesso restrito. Esta página requer perfil de Técnico Líder ou superior."
        )
        return redirect('technician_management')
    return wrapper


# Decorator para /management/ — acessível por todos os perfis com login vinculado ou do grupo adequado.
def tecnico_or_operador_required(view_func):
    @login_required
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        user = request.user
        if _user_can_access_maintenance(user):
            return view_func(request, *args, **kwargs)
        if user.groups.filter(name="Liderança de Produção").exists():
            messages.error(request, "Acesso restrito. Esta área é exclusiva para a Manutenção.")
            return redirect("production:dashboard")
        messages.error(request, "Acesso negado. Faça login com credenciais válidas.")
        return redirect('login')
    return wrapper


@login_required
def home_redirect(request):
    user = request.user
    # 1. Usuário de TV ('tv' ou grupo 'Visualizador')
    if user.username == 'tv' or user.groups.filter(name='Visualizador').exists():
        return redirect('tv_dashboard')

    # 2. Usuário com Acesso Duplo (Manutenção + Produção) -> Portal de Escolha
    if _user_has_dual_access(user):
        return redirect('portal_select')

    # 3. Usuário com Acesso Apenas à Produção
    if _user_can_access_production(user) and not _user_can_access_maintenance(user):
        return redirect('production:dashboard')

    # 4. Usuário com Acesso Apenas à Manutenção
    if _user_can_access_maintenance(user):
        # Técnico Líder ou Técnico comum -> tela de gerenciamento de técnicos (/management/)
        if (
            user.groups.filter(name__in=['Tecnicos_Lideres', 'Tecnicos']).exists()
            or _get_technician_proprio(user)
        ):
            return redirect('technician_management')
        # Operadores puros de manutenção -> dashboard
        return redirect('dashboard')

    # 5. Fallback para usuários sem permissões válidas
    messages.error(request, "Acesso restrito. Seu usuário não possui permissão para acessar os módulos.")
    return redirect('login')


@login_required
def portal_select(request):
    """
    Tela de Seleção de Módulos (Hub / Portal de Entrada).
    Se o usuário não possuir acesso duplo, redireciona-o automaticamente
    para o único módulo ao qual tem direito.
    """
    user = request.user
    if not _user_has_dual_access(user):
        if _user_can_access_production(user):
            return redirect('production:dashboard')
        if _user_can_access_maintenance(user):
            if (
                user.groups.filter(name__in=['Tecnicos_Lideres', 'Tecnicos']).exists()
                or _get_technician_proprio(user)
            ):
                return redirect('technician_management')
            return redirect('dashboard')
        if user.username == 'tv' or user.groups.filter(name='Visualizador').exists():
            return redirect('tv_dashboard')
        messages.error(request, "Acesso restrito. Seu usuário não possui módulos atribuídos.")
        return redirect('login')

    # Destino do botão de manutenção no portal
    if _user_is_operador(user) or _user_is_lider_ou_operador(user):
        maintenance_url = 'dashboard'
    else:
        maintenance_url = 'technician_management'

    context = {
        'maintenance_url': maintenance_url,
    }
    return render(request, 'maintenance/portal_select.html', context)


# ----------------------------------------------------
# 1. TELA A: PAINEL INFORMATIVO DE FÁBRICA (MODO TV)
# ----------------------------------------------------
@login_required
def tv_dashboard(request):
    user = request.user
    # Permitir apenas se for Operador/Admin ou se estiver no grupo Visualizador ou username tv
    is_tv_viewer = user.groups.filter(name='Visualizador').exists() or user.username == 'tv'
    is_operador = user.is_superuser or user.is_staff or user.groups.filter(name__in=['Operadores', 'Operador']).exists()
    
    if not (is_tv_viewer or is_operador):
        if user.groups.filter(name="Liderança de Produção").exists() and not _user_has_maintenance_access(user):
            messages.error(request, "Acesso negado. A TV de exibição é restrita para o seu perfil.")
            return redirect("production:dashboard")
        messages.error(request, "Acesso negado. A TV de exibição é restrita para o seu perfil.")
        return redirect('technician_management')


    active_allocations = Allocation.objects.filter(
        data_fim__isnull=True,
        status='EM_ATENDIMENTO'
    ).select_related('maquina', 'maquina__setor').prefetch_related('pausas')

    technicians = Technician.objects.filter(is_active=True).prefetch_related(
        Prefetch('allocations', queryset=active_allocations)
    ).order_by('nome')

    context = {
        'technicians': technicians,
        'now': timezone.now(),
    }
    return render(request, 'maintenance/tv_dashboard.html', context)


# ----------------------------------------------------
# 2. TELA C: GERENCIAMENTO DE TÉCNICOS EM TEMPO REAL
# Acessível por Operadores (acesso total) e Técnicos vinculados (somente próprio card).
# ----------------------------------------------------
@tecnico_or_operador_required
def technician_management(request):
    technicians = Technician.objects.filter(is_active=True).order_by('nome')
    pending_os_list = OrdemServico.objects.filter(status='PENDENTE').select_related('maquina', 'setor').order_by('-criticidade', 'data_abertura')

    # Instantiate blank forms to render in the modals
    start_form = StartServiceForm()
    pause_form = PauseServiceForm()
    finish_form = FinishServiceForm()

    # Determina contexto de permissão do usuário logado
    is_operador = _user_is_operador(request.user)          # True apenas para OPERADOR puro
    can_manage  = _user_is_lider_ou_operador(request.user) # True para OPERADOR e TECNICO_LIDER
    tecnico_proprio = _get_technician_proprio(request.user)
    technician_proprio_id = tecnico_proprio.id if tecnico_proprio else None

    context = {
        'technicians': technicians,
        'pending_os_list': pending_os_list,
        'start_form': start_form,
        'pause_form': pause_form,
        'finish_form': finish_form,
        'user_is_operador': is_operador,   # usado apenas para o botão "Cadastros"
        'user_can_manage': can_manage,     # usado para ações dos cards e widget de escala
        'technician_proprio_id': technician_proprio_id,
    }
    return render(request, 'maintenance/technician_management.html', context)



# Action: Start Service
@login_required
def start_service(request, technician_id):
    if request.method == 'POST':
        technician = get_object_or_404(Technician, id=technician_id)

        # Verificação de permissão: apenas TECNICO comum tem restrição ao próprio card.
        # TECNICO_LIDER e OPERADOR podem agir em qualquer card.
        if not _user_is_lider_ou_operador(request.user):
            tecnico_proprio = _get_technician_proprio(request.user)
            if not tecnico_proprio or tecnico_proprio.id != technician.id:
                messages.error(request, "Acesso negado. Você só pode iniciar serviços no seu próprio card.")
                return redirect('technician_management')

        # Bloquear se o técnico não estiver ativo no quadro da empresa.
        if not technician.is_active:
            messages.error(request, "Este técnico não está ativo no quadro da empresa e não pode receber novas ordens de serviço.")
            return redirect('technician_management')

        # Bloquear se o técnico estiver ausente/fora da fábrica.
        if technician.is_ausente:
            messages.error(request, f"Técnico {technician.nome} está marcado como '{technician.get_status_display()}' e não pode receber novas ordens. Altere a disponibilidade antes de alocar.")
            return redirect('technician_management')

        # Bloquear APENAS se o técnico já tiver uma alocação EM_ATENDIMENTO ativa.
        # Ter apenas serviços pausados é permitido — o técnico pode receber nova ordem.
        if technician.active_allocation is not None:
            messages.error(request, f"Técnico {technician.nome} já possui um atendimento ativo. Pause-o antes de iniciar outro.")
            return redirect('technician_management')

        form = StartServiceForm(request.POST)
        if form.is_valid():
            allocation = form.save(commit=False)
            allocation.tecnico = technician
            allocation.data_inicio = timezone.now()
            allocation.status = 'EM_ATENDIMENTO'
            allocation.usuario_operador = request.user
            allocation.save()

            # Status do técnico sempre reflete a alocação ativa
            technician.status = 'EM_ATENDIMENTO'
            technician.save()

            messages.success(request, f"Serviço iniciado com sucesso para {technician.nome}.")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Erro: {error}")

    return redirect('technician_management')


# Action: Pause Service (pausa a alocação ativa do técnico)
@login_required
def pause_service(request, technician_id):
    if request.method == 'POST':
        technician = get_object_or_404(Technician, id=technician_id)

        # Verificação de permissão: apenas TECNICO comum tem restrição ao próprio card.
        if not _user_is_lider_ou_operador(request.user):
            tecnico_proprio = _get_technician_proprio(request.user)
            if not tecnico_proprio or tecnico_proprio.id != technician.id:
                messages.error(request, "Acesso negado. Você só pode pausar serviços no seu próprio card.")
                return redirect('technician_management')

        if technician.status != 'EM_ATENDIMENTO':
            messages.error(request, f"Técnico {technician.nome} não está em atendimento ativo.")
            return redirect('technician_management')

        active_alloc = technician.active_allocation
        if not active_alloc:
            messages.error(request, f"Nenhuma alocação ativa encontrada para {technician.nome}.")
            return redirect('technician_management')

        form = PauseServiceForm(request.POST)
        if form.is_valid():
            now_time = timezone.now()
            motivo = form.cleaned_data['motivo_pausa']

            # Cria registro de histórico de pausa
            HistoricoPausa.objects.create(
                alocacao=active_alloc,
                data_pausa=now_time,
                motivo_pausa=motivo
            )

            # Marca a alocação como EM_PAUSA (mantendo campos legados)
            active_alloc.data_pausa = now_time
            active_alloc.motivo_pausa = motivo
            active_alloc.status = 'EM_PAUSA'
            active_alloc.save()

            # Atualiza status do técnico: EM_PAUSA se não houver outra ativa
            if technician.active_allocation is None:
                technician.status = 'EM_PAUSA'
                technician.save()

            messages.warning(request, f"Serviço pausado para {technician.nome}.")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Erro: {error}")

    return redirect('technician_management')


# Action: Resume Service (retoma o ÚNICO serviço pausado — caso simples)
@login_required
def resume_service(request, technician_id):
    if request.method == 'POST':
        technician = get_object_or_404(Technician, id=technician_id)

        # Verificação de permissão: apenas TECNICO comum tem restrição ao próprio card.
        if not _user_is_lider_ou_operador(request.user):
            tecnico_proprio = _get_technician_proprio(request.user)
            if not tecnico_proprio or tecnico_proprio.id != technician.id:
                messages.error(request, "Acesso negado. Você só pode retomar serviços no seu próprio card.")
                return redirect('technician_management')

        if technician.status != 'EM_PAUSA':
            messages.error(request, f"Técnico {technician.nome} não está em pausa.")
            return redirect('technician_management')

        # Pega a única pausada (caso simples sem alocação ativa)
        paused_allocs = technician.paused_allocations
        if not paused_allocs.exists():
            messages.error(request, f"Nenhuma alocação pausada encontrada para {technician.nome}.")
            return redirect('technician_management')

        # Retoma a alocação pausada mais antiga
        alloc = paused_allocs.first()

        # Localiza o último registro de HistoricoPausa onde data_retorno é nulo e preenche
        pausa_aberta = alloc.pausas.filter(data_retorno__isnull=True).order_by('-data_pausa').first()
        if pausa_aberta:
            pausa_aberta.data_retorno = timezone.now()
            pausa_aberta.save()

        alloc.data_pausa = None
        alloc.motivo_pausa = None
        alloc.status = 'EM_ATENDIMENTO'
        alloc.save()

        technician.status = 'EM_ATENDIMENTO'
        technician.save()

        messages.success(request, f"Serviço retomado por {technician.nome}.")

    return redirect('technician_management')


# Action: Resume Paused Allocation (troca de contexto — ativa alocação específica por ID)
@login_required
def resume_paused_allocation(request, allocation_id):
    if request.method == 'POST':
        alloc_to_resume = get_object_or_404(Allocation, id=allocation_id)
        technician = alloc_to_resume.tecnico

        # Verificação de permissão: apenas TECNICO comum tem restrição ao próprio card.
        if not _user_is_lider_ou_operador(request.user):
            tecnico_proprio = _get_technician_proprio(request.user)
            if not tecnico_proprio or tecnico_proprio.id != technician.id:
                messages.error(request, "Acesso negado. Você só pode retomar alocações do seu próprio card.")
                return redirect('technician_management')

        if alloc_to_resume.status != 'EM_PAUSA' or alloc_to_resume.data_fim is not None:
            messages.error(request, "Esta alocação não está em pausa ou já foi encerrada.")
            return redirect('technician_management')
        
        # Se houver uma alocação ativa, ela vai para EM_PAUSA (troca de contexto automática)
        current_active = technician.active_allocation
        if current_active:
            now_time = timezone.now()
            motivo = "Interrompido para retomada de outro serviço."
            HistoricoPausa.objects.create(
                alocacao=current_active,
                data_pausa=now_time,
                motivo_pausa=motivo
            )
            current_active.data_pausa = now_time
            current_active.motivo_pausa = motivo
            current_active.status = 'EM_PAUSA'
            current_active.save()
        
        # Ativa a alocação selecionada
        pausa_aberta = alloc_to_resume.pausas.filter(data_retorno__isnull=True).order_by('-data_pausa').first()
        if pausa_aberta:
            pausa_aberta.data_retorno = timezone.now()
            pausa_aberta.save()
            
        alloc_to_resume.data_pausa = None
        alloc_to_resume.motivo_pausa = None
        alloc_to_resume.status = 'EM_ATENDIMENTO'
        alloc_to_resume.save()
        
        technician.status = 'EM_ATENDIMENTO'
        technician.save()
        
        messages.success(request, f"Alocação retomada: {alloc_to_resume.maquina.nome if alloc_to_resume.maquina else 'Sem máquina'} para {technician.nome}.")
        
    return redirect('technician_management')


# Action: Finish Service (finaliza a alocação ATIVA do técnico)
@login_required
def finish_service(request, technician_id):
    if request.method == 'POST':
        technician = get_object_or_404(Technician, id=technician_id)

        # Verificação de permissão: apenas TECNICO comum tem restrição ao próprio card.
        if not _user_is_lider_ou_operador(request.user):
            tecnico_proprio = _get_technician_proprio(request.user)
            if not tecnico_proprio or tecnico_proprio.id != technician.id:
                messages.error(request, "Acesso negado. Você só pode concluir serviços no seu próprio card.")
                return redirect('technician_management')
        
        active_alloc = technician.active_allocation
        if not active_alloc:
            messages.error(request, f"Nenhuma alocação ativa encontrada para {technician.nome}.")
            return redirect('technician_management')
            
        form = FinishServiceForm(request.POST, request.FILES)
        if form.is_valid():
            now_time = timezone.now()
            active_alloc.data_fim = now_time
            active_alloc.observacao_conclusao = form.cleaned_data['observacao_conclusao']
            if 'foto_anexo' in request.FILES:
                active_alloc.foto_anexo = request.FILES['foto_anexo']
            active_alloc.status = 'CONCLUIDO'
            
            # Garante que se houver uma pausa aberta (sem data de retorno), preenche a data_retorno
            pausa_aberta = active_alloc.pausas.filter(data_retorno__isnull=True).order_by('-data_pausa').first()
            if pausa_aberta:
                pausa_aberta.data_retorno = now_time
                pausa_aberta.save()
                
            active_alloc.save()
            
            # Fechamento de Ordem de Serviço vinculada se este for o último técnico concluindo
            if active_alloc.ordem_servico:
                os_obj = active_alloc.ordem_servico
                outras_ativas = os_obj.allocations.filter(data_fim__isnull=True).exclude(id=active_alloc.id)
                if not outras_ativas.exists():
                    if 'foto_conclusao' in request.FILES:
                        os_obj.foto_conclusao = request.FILES['foto_conclusao']
                    if 'foto_verso' in request.FILES:
                        os_obj.foto_verso = request.FILES['foto_verso']
                    if request.POST.get('lider_assinatura_nome'):
                        os_obj.lider_assinatura_nome = request.POST['lider_assinatura_nome'].strip()
                    if request.POST.get('causa'):
                        os_obj.causa = request.POST['causa'].strip()
                    if request.POST.get('descricao_servico_realizado'):
                        os_obj.descricao_servico_realizado = request.POST['descricao_servico_realizado'].strip()
                    if request.POST.get('observacao_fechamento'):
                        os_obj.observacao_fechamento = request.POST['observacao_fechamento'].strip()
                    
                    pecas_txt = request.POST.get('pecas_utilizadas_texto', '').strip()
                    if pecas_txt:
                        for p_line in pecas_txt.splitlines():
                            if p_line.strip():
                                OrdemServicoPeca.objects.create(ordem_servico=os_obj, descricao=p_line.strip(), quantidade=1.0)

                    os_obj.data_hora_fim_conserto = os_obj.data_hora_fim_conserto or now_time
                    os_obj.data_hora_fim_ocorrencia = os_obj.data_hora_fim_ocorrencia or now_time
                    os_obj.data_conclusao = now_time
                    os_obj.status = 'CONCLUIDA'
                    os_obj.save()

            # Recalcula status do técnico
            if technician.active_allocation is None:
                remaining_paused = technician.paused_allocations.exists()
                technician.status = 'EM_PAUSA' if remaining_paused else 'OCIOSO'
                technician.save()
            
            messages.success(request, f"Serviço finalizado com sucesso por {technician.nome}.")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Erro: {error}")
            return redirect(f'/management/?open_modal=finish_tech&tech_id={technician_id}')
                    
    return redirect('technician_management')


# Action: Finish Allocation (finaliza uma alocação específica por ID — ativa OU pausada)
@login_required
def finish_allocation(request, allocation_id):
    if request.method == 'POST':
        alloc = get_object_or_404(Allocation, id=allocation_id)
        technician = alloc.tecnico

        # Verificação de permissão: apenas TECNICO comum tem restrição ao próprio card.
        if not _user_is_lider_ou_operador(request.user):
            tecnico_proprio = _get_technician_proprio(request.user)
            if not tecnico_proprio or tecnico_proprio.id != technician.id:
                messages.error(request, "Acesso negado. Você só pode finalizar alocações do seu próprio card.")
                return redirect('technician_management')
        
        if alloc.data_fim is not None:
            messages.error(request, "Esta alocação já foi encerrada.")
            return redirect('technician_management')
        
        form = FinishServiceForm(request.POST, request.FILES)
        if form.is_valid():
            now_time = timezone.now()
            alloc.data_fim = now_time
            alloc.observacao_conclusao = form.cleaned_data['observacao_conclusao']
            if 'foto_anexo' in request.FILES:
                alloc.foto_anexo = request.FILES['foto_anexo']
                
            # Garante que se houver uma pausa aberta (sem data de retorno), preenche a data_retorno
            pausa_aberta = alloc.pausas.filter(data_retorno__isnull=True).order_by('-data_pausa').first()
            if pausa_aberta:
                pausa_aberta.data_retorno = now_time
                pausa_aberta.save()
                
            alloc.status = 'CONCLUIDO'
            alloc.save()

            # Fechamento de Ordem de Serviço vinculada se este for o último técnico concluindo
            if alloc.ordem_servico:
                os_obj = alloc.ordem_servico
                outras_ativas = os_obj.allocations.filter(data_fim__isnull=True).exclude(id=alloc.id)
                if not outras_ativas.exists():
                    if 'foto_conclusao' in request.FILES:
                        os_obj.foto_conclusao = request.FILES['foto_conclusao']
                    if 'foto_verso' in request.FILES:
                        os_obj.foto_verso = request.FILES['foto_verso']
                    if request.POST.get('lider_assinatura_nome'):
                        os_obj.lider_assinatura_nome = request.POST['lider_assinatura_nome'].strip()
                    if request.POST.get('causa'):
                        os_obj.causa = request.POST['causa'].strip()
                    if request.POST.get('descricao_servico_realizado'):
                        os_obj.descricao_servico_realizado = request.POST['descricao_servico_realizado'].strip()
                    if request.POST.get('observacao_fechamento'):
                        os_obj.observacao_fechamento = request.POST['observacao_fechamento'].strip()
                    
                    pecas_txt = request.POST.get('pecas_utilizadas_texto', '').strip()
                    if pecas_txt:
                        for p_line in pecas_txt.splitlines():
                            if p_line.strip():
                                OrdemServicoPeca.objects.create(ordem_servico=os_obj, descricao=p_line.strip(), quantidade=1.0)

                    os_obj.data_hora_fim_conserto = os_obj.data_hora_fim_conserto or now_time
                    os_obj.data_hora_fim_ocorrencia = os_obj.data_hora_fim_ocorrencia or now_time
                    os_obj.data_conclusao = now_time
                    os_obj.status = 'CONCLUIDA'
                    os_obj.save()
            
            # Recalcula status do técnico com base nas alocações abertas restantes
            if technician.active_allocation is not None:
                technician.status = 'EM_ATENDIMENTO'
            elif technician.paused_allocations.exists():
                technician.status = 'EM_PAUSA'
            else:
                technician.status = 'OCIOSO'
            technician.save()
            
            messages.success(request, f"Alocação de {alloc.maquina.nome if alloc.maquina else 'serviço'} finalizada com sucesso.")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Erro: {error}")
            return redirect(f'/management/?open_modal=finish_alloc&alloc_id={allocation_id}')
                    
    return redirect('technician_management')



# Action: Add Progress Update (adiciona nota de progresso parcial a uma alocação)
@login_required
def add_allocation_progress_update(request, allocation_id):
    if request.method == 'POST':
        alloc = get_object_or_404(Allocation, id=allocation_id)
        technician = alloc.tecnico

        # Verificação de permissão: apenas TECNICO comum tem restrição ao próprio card.
        if not _user_is_lider_ou_operador(request.user):
            tecnico_proprio = _get_technician_proprio(request.user)
            if not tecnico_proprio or tecnico_proprio.id != technician.id:
                messages.error(request, "Acesso negado. Você só pode registrar atualizações de progresso nas suas próprias alocações.")
                return redirect('technician_management')

        descricao = request.POST.get("descricao", "").strip()
        if not descricao:
            messages.error(request, "A descrição da atualização de progresso não pode ficar em branco.")
            return redirect('technician_management')

        AllocationProgressUpdate.objects.create(
            allocation=alloc,
            autor=request.user,
            descricao=descricao
        )

        messages.success(request, f"Atualização de progresso registrada com sucesso na alocação #{alloc.id}.")
        return redirect('technician_management')

    return redirect('technician_management')


# Action: Set Availability / Absence status
@login_required
def set_availability(request, technician_id):
    """Permite ao OPERADOR definir o status de escala/ausência do técnico.

    Técnicos com perfil TECNICO NÃO podem alterar escalas.

    Status aceitos:
        OCIOSO          → retorna o técnico ao fluxo normal
        AUSENTE_FOLGA   → Folga/Escala
        AUSENTE_FERIAS  → Férias
        AUSENTE_MEDICO  → Licença Médica/Afastamento
        EXTERNO_PLANTAO → Plantão fora da fábrica

    Regras de negócio:
        - Ao marcar ausência, o técnico NÃO recebe novas ordens.
        - Serviços pausados existentes são MANTIDOS congelados no histórico.
        - O status EM_ATENDIMENTO não pode ser definido manualmente aqui;
          ele é controlado exclusivamente pelas views start/pause/finish.
    """
    # OPERADOR e TECNICO_LIDER podem alterar escala. Técnico comum não.
    if not _user_is_lider_ou_operador(request.user):
        messages.error(request, "Acesso negado. Somente Técnicos Líderes e Operadores podem alterar escalas e disponibilidade.")
        return redirect('technician_management')
    if request.method == 'POST':
        technician = get_object_or_404(Technician, id=technician_id)
        if not technician.is_active:
            messages.error(request, "Este técnico não está ativo no quadro da empresa.")
            return redirect('technician_management')

        novo_status = request.POST.get('novo_status', '').strip()

        STATUS_PERMITIDOS = {'OCIOSO'} | set(Technician.STATUS_AUSENCIA)

        if novo_status not in STATUS_PERMITIDOS:
            messages.error(request, "Status de disponibilidade inválido.")
            return redirect('technician_management')

        # Impede alterar técnico que está EM_ATENDIMENTO para ausência diretamente.
        # O operador deve primeiro pausar/encerrar o serviço ativo.
        if technician.status == 'EM_ATENDIMENTO' and novo_status != 'OCIOSO':
            messages.error(
                request,
                f"{technician.nome} está em atendimento ativo. Pause ou finalize o serviço antes de marcar ausência."
            )
            return redirect('technician_management')

        # Se o técnico tem apenas pausados e está voltando para OCIOSO, mantém pausados
        # (não alteramos as alocações — apenas o status do técnico)
        label_novo = dict(Technician.STATUS_CHOICES).get(novo_status, novo_status)
        technician.status = novo_status
        technician.save()

        # Registra a alteração de escala no histórico de auditoria
        HistoricoEscala.objects.create(
            tecnico=technician,
            status_definido=novo_status,
            usuario_responsavel=request.user if request.user.is_authenticated else None,
        )

        if novo_status == 'OCIOSO':
            messages.success(request, f"{technician.nome} retornou como Disponível (Ocioso).")
        else:
            messages.warning(request, f"{technician.nome} marcado como '{label_novo}'. Não receberá novas ordens até retornar como Ocioso.")

    return redirect('technician_management')


# ----------------------------------------------------
# 3. TELA D: DASHBOARD DE ANÁLISE (GESTÃO)
# ----------------------------------------------------
@operador_required
def dashboard(request):
    # ── Filtro de Período (GET) ─────────────────────────────────────────────
    # Captura parâmetros data_inicio e data_final da query string.
    # Valida o formato DD/MM/YYYY ou YYYY-MM-DD (input type="date" envia YYYY-MM-DD).
    # Fallback: últimos 30 dias se os parâmetros forem ausentes ou inválidos.
    today = timezone.localdate()
    default_inicio = today - datetime.timedelta(days=30)

    data_inicio_str = request.GET.get('data_inicio', '').strip()
    data_final_str  = request.GET.get('data_final',  '').strip()

    try:
        data_inicio = datetime.date.fromisoformat(data_inicio_str) if data_inicio_str else default_inicio
    except ValueError:
        data_inicio = default_inicio
        data_inicio_str = ''

    try:
        data_final = datetime.date.fromisoformat(data_final_str) if data_final_str else today
    except ValueError:
        data_final = today
        data_final_str = ''

    # Garante ordem correta (inicio <= fim)
    if data_inicio > data_final:
        data_inicio, data_final = data_final, data_inicio

    # Strings formatadas para popular os inputs do formulário (formato YYYY-MM-DD)
    data_inicio_str = data_inicio.isoformat()
    data_final_str  = data_final.isoformat()

    # ── KPI cards (snapshot do status ATUAL dos técnicos — sem filtro temporal) ──
    total_techs  = Technician.objects.filter(is_active=True).count()
    active_techs = Technician.objects.filter(is_active=True, status='EM_ATENDIMENTO').count()
    paused_techs = Technician.objects.filter(is_active=True, status='EM_PAUSA').count()
    idle_techs   = Technician.objects.filter(is_active=True, status='OCIOSO').count()
    absent_techs = Technician.objects.filter(
        is_active=True,
        status__in=list(Technician.STATUS_AUSENCIA)
    ).count()

    # Distinct machines currently undergoing maintenance (data_fim is null)
    # machines_in_maintenance = Machine.objects.filter(
    #     allocations__data_fim__isnull=True
    # ).distinct().count()

    # ── Base queryset filtrado pelo período com prefetches ─────────────────
    alloc_filtrado = Allocation.objects.filter(
        data_inicio__date__range=[data_inicio, data_final]
    ).select_related('tecnico', 'maquina', 'maquina__setor').prefetch_related('pausas')

    now = timezone.now()
    
    total_bruto_segundos = 0.0
    total_liquido_segundos = 0.0
    concluidas_net_durations = []
    
    # Agrupamentos para gráficos
    maquina_segundos = {}
    maquina_chamados = {}
    criticidade_segundos = {
        'BAIXA': 0.0,
        'MEDIA': 0.0,
        'ALTA': 0.0,
    }
    
    # Para o gráfico de setores (Volume de atendimentos por setor) - mantendo compatibilidade
    sectors = Sector.objects.all()
    alloc_by_sector = {s.nome: 0 for s in sectors}
    
    for alloc in alloc_filtrado:
        # Cálculo de tempo bruto
        if alloc.data_fim:
            gross_seconds = (alloc.data_fim - alloc.data_inicio).total_seconds()
        else:
            gross_seconds = (now - alloc.data_inicio).total_seconds()
        
        gross_seconds = max(0.0, gross_seconds)
        
        # Cálculo de pausas
        pause_seconds = 0.0
        for p in alloc.pausas.all():
            if p.data_retorno:
                p_end = p.data_retorno
            else:
                p_end = alloc.data_fim or now
            
            p_dur = (p_end - p.data_pausa).total_seconds()
            p_dur = max(0.0, p_dur)
            pause_seconds += p_dur
            
        net_seconds = max(0.0, gross_seconds - pause_seconds)
        
        # Acumuladores
        total_bruto_segundos += gross_seconds
        total_liquido_segundos += net_seconds
        
        if alloc.status == 'CONCLUIDO':
            concluidas_net_durations.append(net_seconds)
            
        # Agrupamento Criticidade (horas de manutenção)
        if alloc.maquina:
            crit = alloc.maquina.criticidade
            if crit in criticidade_segundos:
                criticidade_segundos[crit] += gross_seconds
            else:
                criticidade_segundos[crit] = criticidade_segundos.get(crit, 0.0) + gross_seconds
            
            # Agrupamento Máquinas Ofensoras (Top 5 Máquinas) - ignorando projetos e fábrica
            m_nome = alloc.maquina.nome
            m_nome_lower = m_nome.lower()
            if not ('projeto' in m_nome_lower or 'fabrica' in m_nome_lower or 'fábrica' in m_nome_lower):
                maquina_segundos[m_nome] = maquina_segundos.get(m_nome, 0.0) + gross_seconds
                maquina_chamados[m_nome] = maquina_chamados.get(m_nome, 0) + 1
            
            # Volume de atendimentos por setor
            if alloc.maquina.setor:
                s_nome = alloc.maquina.setor.nome
                if s_nome in alloc_by_sector:
                    alloc_by_sector[s_nome] += 1
                else:
                    alloc_by_sector[s_nome] = alloc_by_sector.get(s_nome, 0) + 1

    # 1. MTTR por Equipamento (excluindo projetos/fábrica)
    maquina_concluidas_net = {}
    maquina_concluidas_count = {}
    for alloc in alloc_filtrado:
        if alloc.status == 'CONCLUIDO' and alloc.maquina:
            m_nome = alloc.maquina.nome
            m_nome_lower = m_nome.lower()
            if not ('projeto' in m_nome_lower or 'fabrica' in m_nome_lower or 'fábrica' in m_nome_lower):
                if alloc.data_fim:
                    gross_seconds = (alloc.data_fim - alloc.data_inicio).total_seconds()
                else:
                    gross_seconds = (now - alloc.data_inicio).total_seconds()
                gross_seconds = max(0.0, gross_seconds)
                
                pause_seconds = 0.0
                for p in alloc.pausas.all():
                    if p.data_retorno:
                        p_end = p.data_retorno
                    else:
                        p_end = alloc.data_fim or now
                    p_dur = (p_end - p.data_pausa).total_seconds()
                    pause_seconds += max(0.0, p_dur)
                    
                net_seconds = max(0.0, gross_seconds - pause_seconds)
                
                maquina_concluidas_net[m_nome] = maquina_concluidas_net.get(m_nome, 0.0) + net_seconds
                maquina_concluidas_count[m_nome] = maquina_concluidas_count.get(m_nome, 0) + 1

    mttr_equipamentos = {}
    for m_nome, net_sec in maquina_concluidas_net.items():
        count = maquina_concluidas_count[m_nome]
        if count > 0:
            mttr_equipamentos[m_nome] = round((net_sec / count) / 60.0, 1) # em minutos

    # Ordenar por MTTR decrescente
    sorted_mttr_equip = sorted(mttr_equipamentos.items(), key=lambda x: x[1], reverse=True)
    mttr_equip_labels = [item[0] for item in sorted_mttr_equip]
    mttr_equip_values = [item[1] for item in sorted_mttr_equip]

    # 2. Índice de Eficiência Operacional
    if total_bruto_segundos > 0:
        eficiencia_percent = round((total_liquido_segundos / total_bruto_segundos) * 100, 1)
        eficiencia_display = f"{eficiencia_percent}%"
    else:
        eficiencia_display = "N/A"

    # 3. Taxa de Utilização da Equipe (Simplificada)
    total_horas_liquidas = total_liquido_segundos / 3600.0
    total_dias = (data_final - data_inicio).days + 1
    capacidade_horas = total_dias * 8.0 * total_techs
    if capacidade_horas > 0:
        utilizacao_percent = round((total_horas_liquidas / capacidade_horas) * 100, 1)
        utilizacao_display = f"{utilizacao_percent}%"
    else:
        utilizacao_display = "N/A"

    # 4. Gráfico 1: Pareto de Serviços Executados (agrupado por atividade_observacao)
    servicos_segundos = {}
    for alloc in alloc_filtrado:
        if alloc.status == 'CONCLUIDO':
            desc = (alloc.atividade_observacao or 'Sem descrição').strip()
            if not desc:
                desc = 'Sem descrição'
            
            # Cálculo de tempo bruto
            if alloc.data_fim:
                gross_seconds = (alloc.data_fim - alloc.data_inicio).total_seconds()
            else:
                gross_seconds = (now - alloc.data_inicio).total_seconds()
            gross_seconds = max(0.0, gross_seconds)
            
            # Cálculo de pausas
            pause_seconds = 0.0
            for p in alloc.pausas.all():
                if p.data_retorno:
                    p_end = p.data_retorno
                else:
                    p_end = alloc.data_fim or now
                p_dur = (p_end - p.data_pausa).total_seconds()
                pause_seconds += max(0.0, p_dur)
                
            net_seconds = max(0.0, gross_seconds - pause_seconds)
            servicos_segundos[desc] = servicos_segundos.get(desc, 0.0) + net_seconds

    sorted_servicos = sorted(servicos_segundos.items(), key=lambda x: x[1], reverse=True)[:15]
    servico_labels = [item[0] for item in sorted_servicos]
    servico_values = [round(item[1] / 3600.0, 1) for item in sorted_servicos] # em horas

    # 5. Gráfico 2: Top 5 Máquinas Ofensoras (tempo bruto de manutenção)
    sorted_maquinas = sorted(maquina_segundos.items(), key=lambda x: x[1], reverse=True)[:5]
    ofensoras_labels = [item[0] for item in sorted_maquinas]
    ofensoras_durations = [round(item[1] / 3600.0, 1) for item in sorted_maquinas] # em horas
    ofensoras_counts = [maquina_chamados[item[0]] for item in sorted_maquinas]

    # 6. Gráfico 3: Distribuição por Criticidade (Horas gastas em manutenção)
    crit_labels = ['Baixa', 'Média', 'Alta']
    crit_values = [
        round(criticidade_segundos.get('BAIXA', 0.0) / 3600.0, 1),
        round(criticidade_segundos.get('MEDIA', 0.0) / 3600.0, 1),
        round(criticidade_segundos.get('ALTA', 0.0) / 3600.0, 1)
    ]

    # 7. Pie/Doughnut Chart Data: Distribution of tech status
    status_distribution = {
        'Ocioso': idle_techs,
        'Em Atendimento': active_techs,
        'Em Pausa': paused_techs,
        'Ausente/Externo': absent_techs,
    }

    context = {
        'total_techs':  total_techs,
        'active_techs': active_techs,
        'paused_techs': paused_techs,
        'idle_techs':   idle_techs,
        'absent_techs': absent_techs,

        # Datas do filtro
        'data_inicio_str': data_inicio_str,
        'data_final_str':  data_final_str,

        # Novos KPIs
        'eficiencia_display': eficiencia_display,
        'utilizacao_display': utilizacao_display,

        # Serializações para gráficos
        'status_labels': json.dumps(list(status_distribution.keys())),
        'status_values': json.dumps(list(status_distribution.values())),

        'servico_labels': json.dumps(servico_labels),
        'servico_values': json.dumps(servico_values),

        'ofensoras_labels': json.dumps(ofensoras_labels),
        'ofensoras_durations': json.dumps(ofensoras_durations),
        'ofensoras_counts': json.dumps(ofensoras_counts),

        'crit_labels': json.dumps(crit_labels),
        'crit_values': json.dumps(crit_values),

        'mttr_equip_labels': json.dumps(mttr_equip_labels),
        'mttr_equip_values': json.dumps(mttr_equip_values),

        'sector_labels': json.dumps(list(alloc_by_sector.keys())),
        'sector_values': json.dumps(list(alloc_by_sector.values())),
    }

    # 8. Desempenho por Técnico (concluídas no período)
    tech_concluidas_net = {}
    tech_concluidas_count = {}
    for alloc in alloc_filtrado:
        if alloc.status == 'CONCLUIDO' and alloc.tecnico:
            t_nome = alloc.tecnico.nome
            
            # Cálculo de tempo bruto
            if alloc.data_fim:
                gross_seconds = (alloc.data_fim - alloc.data_inicio).total_seconds()
            else:
                gross_seconds = (now - alloc.data_inicio).total_seconds()
            gross_seconds = max(0.0, gross_seconds)
            
            # Cálculo de pausas
            pause_seconds = 0.0
            for p in alloc.pausas.all():
                if p.data_retorno:
                    p_end = p.data_retorno
                else:
                    p_end = alloc.data_fim or now
                p_dur = (p_end - p.data_pausa).total_seconds()
                pause_seconds += max(0.0, p_dur)
                
            net_seconds = max(0.0, gross_seconds - pause_seconds)
            
            tech_concluidas_net[t_nome] = tech_concluidas_net.get(t_nome, 0.0) + net_seconds
            tech_concluidas_count[t_nome] = tech_concluidas_count.get(t_nome, 0) + 1

    tech_data = []
    for t_nome, count in tech_concluidas_count.items():
        net_sec = tech_concluidas_net.get(t_nome, 0.0)
        mttr_min = round((net_sec / count) / 60.0, 1) if count > 0 else 0.0
        tech_data.append((t_nome, count, mttr_min))

    # Ordenar por volume decrescente
    tech_data_sorted = sorted(tech_data, key=lambda x: x[1], reverse=True)
    tech_desempenho_labels = [item[0] for item in tech_data_sorted]
    tech_desempenho_volumes = [item[1] for item in tech_data_sorted]
    tech_desempenho_mttrs = [item[2] for item in tech_data_sorted]

    context['tech_desempenho_labels'] = json.dumps(tech_desempenho_labels)
    context['tech_desempenho_volumes'] = json.dumps(tech_desempenho_volumes)
    context['tech_desempenho_mttrs'] = json.dumps(tech_desempenho_mttrs)

    return render(request, 'maintenance/dashboard.html', context)


# ----------------------------------------------------
# 4. TELA B: CADASTRO E CONFIGURAÇÕES (CRUDs)
# ----------------------------------------------------
@operador_required
def crud_list(request):
    sectors = Sector.objects.all().order_by('nome')
    machines = Machine.objects.all().order_by('nome')
    technicians = Technician.objects.all().order_by('nome')
    
    context = {
        'sectors': sectors,
        'machines': machines,
        'technicians': technicians,
    }
    return render(request, 'maintenance/crud_list.html', context)


# Sector CRUD
@operador_required
def sector_create(request):
    if request.method == 'POST':
        form = SectorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Setor cadastrado com sucesso.")
            return redirect('crud_list')
    else:
        form = SectorForm()
    return render(request, 'maintenance/sector_form.html', {'form': form, 'title': 'Cadastrar Setor'})

@operador_required
def sector_edit(request, pk):
    sector = get_object_or_404(Sector, pk=pk)
    if request.method == 'POST':
        form = SectorForm(request.POST, instance=sector)
        if form.is_valid():
            form.save()
            messages.success(request, "Setor atualizado com sucesso.")
            return redirect('crud_list')
    else:
        form = SectorForm(instance=sector)
    return render(request, 'maintenance/sector_form.html', {'form': form, 'title': 'Editar Setor'})

@operador_required
def sector_delete(request, pk):
    sector = get_object_or_404(Sector, pk=pk)
    if request.method == 'POST':
        sector.delete()
        messages.success(request, "Setor excluído com sucesso.")
        return redirect('crud_list')
    return render(request, 'maintenance/crud_confirm_delete.html', {'object': sector, 'type': 'Setor'})


# Machine CRUD
@operador_required
def machine_create(request):
    if request.method == 'POST':
        form = MachineForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Máquina cadastrada com sucesso.")
            return redirect('crud_list')
    else:
        form = MachineForm()
    return render(request, 'maintenance/machine_form.html', {'form': form, 'title': 'Cadastrar Máquina'})

@operador_required
def machine_edit(request, pk):
    machine = get_object_or_404(Machine, pk=pk)
    if request.method == 'POST':
        form = MachineForm(request.POST, instance=machine)
        if form.is_valid():
            form.save()
            messages.success(request, "Máquina atualizada com sucesso.")
            return redirect('crud_list')
    else:
        form = MachineForm(instance=machine)
    return render(request, 'maintenance/machine_form.html', {'form': form, 'title': 'Editar Máquina'})

@operador_required
def machine_delete(request, pk):
    machine = get_object_or_404(Machine, pk=pk)
    if request.method == 'POST':
        machine.delete()
        messages.success(request, "Máquina excluída com sucesso.")
        return redirect('crud_list')
    return render(request, 'maintenance/crud_confirm_delete.html', {'object': machine, 'type': 'Máquina'})


# Technician CRUD
@operador_required
def technician_create(request):
    if request.method == 'POST':
        form = TechnicianForm(request.POST)
        if form.is_valid():
            technician = form.save(commit=False)
            # Processa criação de usuário se username foi fornecido
            username = form.cleaned_data.get('username_login', '').strip()
            senha = form.cleaned_data.get('senha_acesso', '').strip()
            perfil = form.cleaned_data.get('perfil_acesso', 'TECNICO') or 'TECNICO'
            if username and senha:
                user = User.objects.create_user(username=username, password=senha)
                technician.user = user

            if technician.user:
                # Sincroniza grupo nativo do Django
                from django.contrib.auth.models import Group
                technician.user.groups.clear()
                if perfil == 'TECNICO':
                    technician.user.groups.add(Group.objects.get(name='Tecnicos'))
                elif perfil == 'TECNICO_LIDER':
                    technician.user.groups.add(Group.objects.get(name='Tecnicos_Lideres'))
                elif perfil == 'OPERADOR':
                    technician.user.groups.add(Group.objects.get(name='Operadores'))
            technician.perfil = perfil
            technician.save()
            messages.success(request, "Técnico cadastrado com sucesso." + (
                f" Usuário '{username}' criado e vinculado." if username and senha else ""
            ))
            return redirect('crud_list')
    else:
        form = TechnicianForm()
    return render(request, 'maintenance/technician_form.html', {'form': form, 'title': 'Cadastrar Técnico'})

@operador_required
def technician_edit(request, pk):
    technician = get_object_or_404(Technician, pk=pk)
    if request.method == 'POST':
        form = TechnicianForm(request.POST, instance=technician)
        if form.is_valid():
            technician = form.save(commit=False)
            # Processa criação/atualização de usuário se username foi fornecido
            username = form.cleaned_data.get('username_login', '').strip()
            senha = form.cleaned_data.get('senha_acesso', '').strip()
            perfil = form.cleaned_data.get('perfil_acesso', 'TECNICO') or 'TECNICO'
            if username:
                if technician.user:
                    # Autoriza alteração do usuário já vinculado
                    technician.user.username = username
                    if senha:
                        technician.user.set_password(senha)
                    technician.user.save()
                elif senha:
                    # Cria um novo usuário e vincula
                    user = User.objects.create_user(username=username, password=senha)
                    technician.user = user

            if technician.user:
                # Sincroniza grupo nativo do Django
                from django.contrib.auth.models import Group
                technician.user.groups.clear()
                if perfil == 'TECNICO':
                    technician.user.groups.add(Group.objects.get(name='Tecnicos'))
                elif perfil == 'TECNICO_LIDER':
                    technician.user.groups.add(Group.objects.get(name='Tecnicos_Lideres'))
                elif perfil == 'OPERADOR':
                    technician.user.groups.add(Group.objects.get(name='Operadores'))
            technician.perfil = perfil
            technician.save()
            msg_extra = ""
            if username:
                msg_extra = f" Usuário '{username}' atualizado/vinculado com perfil {perfil}."
            messages.success(request, "Técnico atualizado com sucesso." + msg_extra)
            return redirect('crud_list')
    else:
        form = TechnicianForm(instance=technician)
        # Pré-preencher username e perfil se já houver usuário vinculado
        if technician.user:
            form.initial['username_login'] = technician.user.username
        form.initial['perfil_acesso'] = technician.perfil or 'TECNICO'
    return render(request, 'maintenance/technician_form.html', {'form': form, 'title': 'Editar Técnico', 'technician': technician})

@operador_required
def technician_delete(request, pk):
    technician = get_object_or_404(Technician, pk=pk)
    if request.method == 'POST':
        technician.delete()
        messages.success(request, "Técnico excluído com sucesso.")
        return redirect('crud_list')
    return render(request, 'maintenance/crud_confirm_delete.html', {'object': technician, 'type': 'Técnico'})


# ----------------------------------------------------
# 5. EXPORTAÇÃO DE RELATÓRIO EXCEL
# ----------------------------------------------------
@operador_required
def exportar_relatorio_excel(request):
    """
    Gera e retorna um arquivo Excel (.xlsx) com o relatório detalhado de alocações.
    Respeita os parâmetros GET data_inicio e data_final (formato YYYY-MM-DD) para
    filtrar somente o período selecionado no Dashboard — o mesmo filtro unificado
    que restringe os gráficos da tela.

    Colunas exportadas (13 no total):
      A  Técnico          B  Matrícula        C  Setor
      D  Máquina          E  Criticidade      F  Operador
      G  Status           H  Data de Início   I  Data de Término
      J  Tempo Total(min) K  Obs. Inicial     L  Obs. de Conclusão
      M  Histórico de Pausas (pausas concatenadas, quebra de linha)

    Requer: openpyxl >= 3.1.0
    """
    # ── Filtro de Período (mesmos parâmetros GET do Dashboard) ─────────────
    today = timezone.localdate()
    default_inicio = today - datetime.timedelta(days=30)

    data_inicio_str = request.GET.get('data_inicio', '').strip()
    data_final_str  = request.GET.get('data_final',  '').strip()

    try:
        data_inicio = datetime.date.fromisoformat(data_inicio_str) if data_inicio_str else default_inicio
    except ValueError:
        data_inicio = default_inicio

    try:
        data_final = datetime.date.fromisoformat(data_final_str) if data_final_str else today
    except ValueError:
        data_final = today

    if data_inicio > data_final:
        data_inicio, data_final = data_final, data_inicio

    # ── Busca alocações do período com todos os dados relacionados ──────────
    allocations = Allocation.objects.filter(
        data_inicio__date__range=[data_inicio, data_final]
    ).select_related(
        'tecnico', 'maquina', 'maquina__setor', 'usuario_operador'
    ).prefetch_related('pausas').order_by('-data_inicio')

    # ── Workbook ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Alocações"

    # ── Estilos ───────────────────────────────────────────────────────────
    header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    alt_fill     = PatternFill(start_color='EBF0F8', end_color='EBF0F8', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    thin_border  = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    CRIT_LABELS = {'BAIXA': 'Baixa', 'MEDIA': 'Média', 'ALTA': 'Alta'}
    STATUS_LABELS = {
        'EM_ATENDIMENTO': 'Em Atendimento',
        'EM_PAUSA': 'Em Pausa',
        'CONCLUIDO': 'Concluído',
    }
    BR_TZ = timezone.get_current_timezone()

    # ── Título da planilha (13 colunas: A1:M1) ────────────────────────────
    periodo_label = f"{data_inicio.strftime('%d/%m/%Y')} a {data_final.strftime('%d/%m/%Y')}"
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = (
        f"Relatório de Manutenção — Período: {periodo_label} — "
        f"Exportado em {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
    )
    title_cell.font      = Font(name='Calibri', bold=True, size=13, color='1E3A5F')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # ── Cabeçalhos (13 colunas) ───────────────────────────────────────────
    HEADERS = [
        'Técnico', 'Matrícula', 'Setor', 'Máquina', 'Criticidade',
        'Operador', 'Status', 'Data de Início', 'Data de Término',
        'Tempo Total (min)', 'Obs. Inicial', 'Obs. de Conclusão',
        'Histórico de Pausas',
    ]
    COL_WIDTHS = [22, 14, 18, 22, 12, 20, 16, 20, 20, 18, 35, 35, 45]

    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 30

    # ── Linhas de dados ───────────────────────────────────────────────────
    for row_idx, alloc in enumerate(allocations, start=3):
        is_alt   = (row_idx % 2 == 0)
        row_fill = alt_fill if is_alt else None

        # Dados básicos derivados
        setor_nome   = alloc.maquina.setor.nome if alloc.maquina and alloc.maquina.setor else '—'
        maquina_nome = alloc.maquina.nome        if alloc.maquina else '—'
        crit_label   = CRIT_LABELS.get(alloc.maquina.criticidade, '—') if alloc.maquina else '—'
        status_label = STATUS_LABELS.get(alloc.status, alloc.status)

        # Operador responsável pela alocação
        if alloc.usuario_operador:
            nome_completo = alloc.usuario_operador.get_full_name()
            operador_label = nome_completo if nome_completo.strip() else alloc.usuario_operador.username
        else:
            operador_label = '—'

        # Datas formatadas
        data_inicio_fmt = (
            timezone.localtime(alloc.data_inicio, BR_TZ).strftime('%d/%m/%Y %H:%M')
            if alloc.data_inicio else '—'
        )
        data_fim_fmt = (
            timezone.localtime(alloc.data_fim, BR_TZ).strftime('%d/%m/%Y %H:%M')
            if alloc.data_fim else '—'
        )

        # Tempo total de atendimento (descontando pausas)
        if alloc.data_inicio:
            fim = alloc.data_fim or timezone.now()
            total_seconds = int((fim - alloc.data_inicio).total_seconds())
            for pausa in alloc.pausas.all():
                if pausa.data_retorno:
                    total_seconds -= int((pausa.data_retorno - pausa.data_pausa).total_seconds())
                else:
                    pausa_fim = alloc.data_fim or timezone.now()
                    total_seconds -= int((pausa_fim - pausa.data_pausa).total_seconds())
            tempo_min = round(max(0, total_seconds) / 60, 1)
        else:
            tempo_min = '—'

        # Observação inicial
        obs_inicial = alloc.atividade_observacao or '—'

        # Observação de conclusão
        obs_conclusao = alloc.observacao_conclusao or '—'

        # Histórico de pausas — concatenado com quebras de linha
        pausas_linhas = []
        for pausa in alloc.pausas.all().order_by('data_pausa'):
            p_inicio = timezone.localtime(pausa.data_pausa, BR_TZ).strftime('%d/%m/%Y %H:%M')
            if pausa.data_retorno:
                p_retorno = timezone.localtime(pausa.data_retorno, BR_TZ).strftime('%d/%m/%Y %H:%M')
            else:
                p_retorno = 'Em aberto'
            motivo = (pausa.motivo_pausa or '').strip()
            pausas_linhas.append(f"↓ {p_inicio}  →  {p_retorno} | {motivo}")
        historico_pausas = '\n'.join(pausas_linhas) if pausas_linhas else '—'

        row_data = [
            alloc.tecnico.nome,   # A
            alloc.tecnico.matricula,  # B
            setor_nome,           # C
            maquina_nome,         # D
            crit_label,           # E
            operador_label,       # F
            status_label,         # G
            data_inicio_fmt,      # H
            data_fim_fmt,         # I
            tempo_min,            # J
            obs_inicial,          # K
            obs_conclusao,        # L
            historico_pausas,     # M
        ]

        # Colunas de texto longo: A(1) C(3) D(4) F(6) G(7) K(11) L(12) M(13)
        LEFT_COLS = {1, 3, 4, 6, 7, 11, 12, 13}

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = left_align if col_idx in LEFT_COLS else center_align

        # Altura dinâmica: linhas com histórico de pausas recebem mais altura
        n_pausas = len(pausas_linhas)
        ws.row_dimensions[row_idx].height = max(20, 18 * max(1, n_pausas))

    # Congela cabeçalhos
    ws.freeze_panes = 'A3'

    # ── Geração do arquivo em memória ─────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    data_str   = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')
    periodo_fn = f"{data_inicio.strftime('%Y%m%d')}-{data_final.strftime('%Y%m%d')}"
    filename   = f'relatorio_manutencao_{periodo_fn}_{data_str}.xlsx'

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# PWA: View para servir o Service Worker a partir da raiz do site
# ─────────────────────────────────────────────────────────────────────────────
def service_worker_view(request):
    """Serve o arquivo service-worker.js com Content-Type e Service-Worker-Allowed
    adequados para que o Service Worker tenha escopo sobre toda a aplicação (/).
    
    Não requer autenticação para que o SW possa ser registrado na tela de login.
    """
    import os
    from django.conf import settings

    sw_path = os.path.join(
        settings.BASE_DIR, 'maintenance', 'static', 'maintenance', 'service-worker.js'
    )

    try:
        with open(sw_path, 'r', encoding='utf-8') as f:
            sw_content = f.read()
    except FileNotFoundError:
        return HttpResponse('// Service Worker not found', content_type='application/javascript', status=404)

    response = HttpResponse(sw_content, content_type='application/javascript')
    response['Service-Worker-Allowed'] = '/'
    response['Cache-Control'] = 'no-cache'
    return response


# ----------------------------------------------------
# 6. PASSAGEM DE TURNO (RELATÓRIO)
# ----------------------------------------------------
@tecnico_or_operador_required
def relatorio_turno(request):
    tecnico = _get_technician_proprio(request.user)
    if not tecnico:
        messages.error(request, "Seu usuário não possui um perfil de Técnico vinculado para gerar o relatório de turno.")
        return redirect('technician_management')

    from datetime import timedelta
    limite_tempo = timezone.now() - timedelta(hours=12)
    allocations = Allocation.objects.filter(
        tecnico=tecnico
    ).filter(
        Q(data_inicio__gte=limite_tempo) | Q(data_fim__gte=limite_tempo)
    ).select_related('maquina')

    if request.method == 'POST':
        texto = request.POST.get('texto_relatorio', '').strip()
        destino = request.POST.get('destino', 'meu_numero').strip()
        
        if destino == 'meu_numero':
            numero_destino = (tecnico.whatsapp or '').strip()
            if not numero_destino:
                messages.warning(request, "Relatório salvo, mas o técnico não possui número de WhatsApp cadastrado.")
                return redirect('relatorio_turno')
        else:
            if not WhatsAppGroup.objects.filter(jid=destino, is_active=True).exists():
                messages.error(request, "Destino inválido selecionado.")
                return redirect('relatorio_turno')
            numero_destino = destino
            
        import requests
        from django.conf import settings
        try:
            payload = {
                'numero': numero_destino,
                'mensagem': texto
            }
            whatsapp_url = getattr(settings, 'WHATSAPP_SERVICE_URL', 'http://localhost:3000/send')
            response = requests.post(whatsapp_url, json=payload, timeout=10)
            if response.status_code in [200, 202]:
                messages.success(request, "Relatório enviado com sucesso via WhatsApp!")
            elif response.status_code == 429:
                messages.warning(request, "Muitas requisições enviadas em curto período. Por favor, aguarde um momento antes de tentar novamente.")
            elif response.status_code == 503:
                try:
                    res_json = response.json()
                    error_msg = res_json.get('error', '')
                    if 'Serviço temporariamente indisponível' in error_msg:
                        messages.warning(request, "Serviço temporariamente indisponível. Por favor, tente novamente mais tarde.")
                    else:
                        messages.warning(request, "Relatório salvo, mas o servidor de WhatsApp está offline.")
                except ValueError:
                    messages.warning(request, "Relatório salvo, mas o servidor de WhatsApp está offline.")
            else:
                messages.warning(request, "Relatório salvo, mas o servidor de WhatsApp está offline.")
        except requests.exceptions.RequestException:
            messages.warning(request, "Relatório salvo, mas o servidor de WhatsApp está offline.")
            
        return redirect('relatorio_turno')

    # Lógica de Construção do Texto (String)
    texto_linhas = []
    texto_linhas.append("Boa noite")
    texto_linhas.append("Passagem de turno")
    texto_linhas.append(f"Técnico: {tecnico.nome}")
    texto_linhas.append("")

    # Corpo (Concluídos)
    concluidas = allocations.filter(status='CONCLUIDO')
    for alloc in concluidas:
        nome_maquina = alloc.maquina.nome if alloc.maquina else "Sem máquina"
        obs = (alloc.observacao_conclusao or "").strip()
        if not obs:
            obs = (alloc.atividade_observacao or "").strip()
        if not obs:
            obs = "Serviço sem descrição"
        texto_linhas.append(f"* {nome_maquina} - {obs}")

    texto_linhas.append("")

    # Rodapé (Pendências)
    pendentes = allocations.filter(status__in=['EM_ATENDIMENTO', 'EM_PAUSA'])
    if not pendentes.exists():
        texto_linhas.append("Sem pendências para o próximo turno")
    else:
        texto_linhas.append("Pendências para o próximo turno:")
        for alloc in pendentes:
            nome_maquina = alloc.maquina.nome if alloc.maquina else "Ocioso"
            if alloc.status == 'EM_PAUSA':
                motivo = (alloc.motivo_pausa or "").strip()
                status_desc = f"Em Pausa - {motivo}" if motivo else "Em Pausa"
            else:
                status_desc = "Em Atendimento"
            texto_linhas.append(f"* {nome_maquina} - {status_desc}")

    texto_precompilado = "\n".join(texto_linhas)

    grupos = WhatsAppGroup.objects.filter(is_active=True)

    context = {
        'tecnico': tecnico,
        'texto_precompilado': texto_precompilado,
        'grupos_whatsapp': grupos,
    }
    return render(request, 'maintenance/relatorio_turno.html', context)


# ----------------------------------------------------
# 4. DOWNLOAD E EXIBIÇÃO PROTEGIDA DE ANEXOS DE ALOCAÇÃO
# ----------------------------------------------------
@login_required
def serve_allocation_attachment(request, allocation_id):
    """Serve o anexo de foto de uma alocação de forma autenticada e autorizada."""
    allocation = get_object_or_404(Allocation, id=allocation_id)
    user = request.user

    # 1. Regra para perfil de Produção (sem acesso à manutenção)
    if user.groups.filter(name="Liderança de Produção").exists() and not _user_has_maintenance_access(user):
        return HttpResponseForbidden("Acesso negado. Perfil não autorizado para visualizar anexos de manutenção.")

    # 2. Regra de autorização baseada no perfil e relação com a alocação
    if _user_is_lider_ou_operador(user):
        pass  # Operadores, Staff, Superusers e Técnicos Líderes têm acesso total
    else:
        # Técnico comum: restrito às suas próprias alocações
        tecnico_proprio = _get_technician_proprio(user)
        if not tecnico_proprio or tecnico_proprio.id != allocation.tecnico_id:
            return HttpResponseForbidden("Acesso negado. Você não possui permissão para acessar este anexo.")

    # 3. Confirmar a existência do campo foto_anexo
    if not allocation.foto_anexo:
        raise Http404("Nenhum anexo associado a esta alocação.")

    # 4. Confirmar a existência física do arquivo no armazenamento sem gerar 500
    try:
        if not allocation.foto_anexo.storage.exists(allocation.foto_anexo.name):
            raise Http404("Arquivo de anexo não encontrado no armazenamento.")
        file_handle = allocation.foto_anexo.open('rb')
    except (FileNotFoundError, OSError, ValueError):
        raise Http404("Arquivo de anexo não encontrado no armazenamento.")

    # 5. Determinar Content-Type e nome seguro do arquivo
    content_type, _ = mimetypes.guess_type(allocation.foto_anexo.name)
    if not content_type:
        content_type = 'application/octet-stream'

    filename = os.path.basename(allocation.foto_anexo.name)
    safe_filename = f"anexo_alocacao_{allocation.id}_{filename}"

    # 6. Retornar FileResponse com cabeçalhos de segurança
    response = FileResponse(file_handle, content_type=content_type)
    response['Content-Disposition'] = f'inline; filename="{safe_filename}"'
    response['Cache-Control'] = 'private, no-store'
    response['X-Content-Type-Options'] = 'nosniff'

    return response


@login_required
def extrair_dados_os_foto_api(request):
    """
    Endpoint de API interno para processar imagem da folha física de OS via Gemini Vision.
    Recebe requisição POST com arquivo multipart 'foto_os'.
    """
    if request.method != "POST":
        return JsonResponse({"sucesso": False, "mensagem": "Método não permitido. Utilize POST."}, status=405)

    foto_os = request.FILES.get("foto_os")
    if not foto_os:
        return JsonResponse({"sucesso": False, "mensagem": "Nenhum arquivo de imagem 'foto_os' foi enviado."}, status=400)

    from .services.os_ocr_service import extrair_dados_os_por_foto
    resultado = extrair_dados_os_por_foto(foto_os)

    return JsonResponse(resultado)


@login_required
def api_verificar_numero_os(request):
    """
    Verifica instantaneamente se um número de OS física já existe no banco.
    Usado para feedback anti-duplicidade em tempo real no frontend.
    """
    numero = request.GET.get("numero", "").strip().upper()
    if not numero:
        return JsonResponse({"existe": False})

    os_existente = OrdemServico.objects.filter(numero_os__iexact=numero).first()
    if os_existente:
        maquina_str = os_existente.descricao_equipamento or (os_existente.maquina.nome if os_existente.maquina else "Não especificada")
        return JsonResponse({
            "existe": True,
            "os": {
                "id": os_existente.id,
                "numero": os_existente.numero_os,
                "data": os_existente.data_abertura.strftime('%d/%m/%Y %H:%M'),
                "status": os_existente.get_status_display(),
                "solicitante": os_existente.solicitante or "",
                "maquina": maquina_str,
            }
        })
    return JsonResponse({"existe": False})


@os_creation_required
def os_create(request):
    """
    Tela de abertura de nova Ordem de Serviço física com suporte a captura de foto,
    reconhecimento assistido por IA e prevenção estrita de duplicidade.
    Suporta o parâmetro vincular_alocacao para vincular imediatamente a um atendimento emergencial.
    """
    vincular_alocacao_id = request.POST.get('vincular_alocacao') or request.GET.get('vincular_alocacao')

    if request.method == "POST":
        form = OrdemServicoCreateForm(request.POST, request.FILES)
        if form.is_valid():
            os_obj = form.save(commit=False)
            os_obj.criado_por = request.user
            if not os_obj.data_hora_inicio_ocorrencia:
                os_obj.data_hora_inicio_ocorrencia = timezone.now()
            os_obj.status = 'PENDENTE'
            os_obj.save()

            if vincular_alocacao_id:
                try:
                    alloc_to_link = Allocation.objects.get(id=vincular_alocacao_id)
                    alloc_to_link.ordem_servico = os_obj
                    alloc_to_link.save()
                    os_obj.status = 'EM_ANDAMENTO'
                    if not os_obj.data_hora_inicio_conserto:
                        os_obj.data_hora_inicio_conserto = alloc_to_link.data_inicio
                    os_obj.save()
                    messages.success(
                        request,
                        f"Ordem de Serviço nº {os_obj.numero_os} aberta e vinculada com sucesso ao atendimento do técnico {alloc_to_link.tecnico.nome}!"
                    )
                    return redirect('technician_management')
                except Allocation.DoesNotExist:
                    pass

            messages.success(
                request,
                f"Ordem de Serviço nº {os_obj.numero_os} aberta com sucesso! Foto da folha física registrada."
            )
            return redirect('technician_management')
        else:
            messages.error(request, "Por favor, corrija os erros indicados no formulário abaixo.")
    else:
        # Inicializa com solicitante padrão baseado no usuário logado
        solicitante_inicial = request.user.get_full_name() or request.user.username
        initial_data = {
            'solicitante': solicitante_inicial,
            'data_hora_inicio_ocorrencia': timezone.localtime().strftime('%Y-%m-%dT%H:%M'),
            'parou_maquina': True,
            'criticidade': 'MEDIA',
            'tipo_manutencao': 'CORRETIVA',
        }
        form = OrdemServicoCreateForm(initial=initial_data)

    machines_json = [
        {"id": m.id, "nome": m.nome, "setor_id": m.setor_id, "setor_nome": m.setor.nome}
        for m in Machine.objects.select_related('setor').all()
    ]

    return render(request, "maintenance/os_create.html", {
        "form": form,
        "machines_json": json.dumps(machines_json),
        "vincular_alocacao_id": vincular_alocacao_id,
    })



@login_required
def os_board(request):
    """
    Quadro visual de Ordens de Serviço (Kanban/Abas):
    - Aba 1: Pendentes (Na fila para atendimento)
    - Aba 2: Em Andamento (Atendimento com suporte a multi-técnicos)
    - Aba 3: Concluídas (Histórico recente com fotos de conclusão)
    - Aba 4: Canceladas
    """
    search = request.GET.get('busca', '').strip()
    setor_id = request.GET.get('setor', '').strip()
    criticidade = request.GET.get('criticidade', '').strip()
    active_tab = request.GET.get('tab', 'pendentes').strip()

    qs = OrdemServico.objects.select_related('maquina', 'setor', 'tecnico_designado', 'criado_por').prefetch_related('allocations__tecnico', 'pecas_utilizadas')

    if search:
        qs = qs.filter(
            Q(numero_os__icontains=search) |
            Q(solicitante__icontains=search) |
            Q(descricao_falha__icontains=search) |
            Q(descricao_equipamento__icontains=search) |
            Q(tag__icontains=search) |
            Q(motivo__icontains=search) |
            Q(maquina__nome__icontains=search) |
            Q(setor__nome__icontains=search)
        )

    if setor_id:
        qs = qs.filter(Q(setor_id=setor_id) | Q(maquina__setor_id=setor_id))

    if criticidade:
        qs = qs.filter(criticidade=criticidade)

    # Separação por status
    pendentes = qs.filter(status='PENDENTE').order_by(
        models.Case(
            models.When(criticidade='ALTA', then=models.Value(1)),
            models.When(criticidade='MEDIA', then=models.Value(2)),
            default=models.Value(3),
            output_field=models.IntegerField(),
        ),
        'data_abertura'
    )

    em_andamento = qs.filter(status='EM_ANDAMENTO').order_by(
        models.Case(
            models.When(criticidade='ALTA', then=models.Value(1)),
            models.When(criticidade='MEDIA', then=models.Value(2)),
            default=models.Value(3),
            output_field=models.IntegerField(),
        ),
        '-data_abertura'
    )

    concluidas = qs.filter(status='CONCLUIDA').order_by('-data_conclusao', '-data_abertura')[:60]
    canceladas = qs.filter(status='CANCELADA').order_by('-data_abertura')[:30]

    # Contagens gerais (sem filtro de busca para os badges das abas)
    counts = {
        'pendentes': OrdemServico.objects.filter(status='PENDENTE').count(),
        'em_andamento': OrdemServico.objects.filter(status='EM_ANDAMENTO').count(),
        'concluidas': OrdemServico.objects.filter(status='CONCLUIDA').count(),
        'canceladas': OrdemServico.objects.filter(status='CANCELADA').count(),
    }

    # Técnicos disponíveis para modais
    active_technicians = Technician.objects.filter(is_active=True).order_by('nome')
    available_technicians = Technician.objects.filter(is_active=True).exclude(status__startswith='AUSENTE').order_by('nome')
    sectors = Sector.objects.all().order_by('nome')
    technician_proprio = _get_technician_proprio(request.user)

    return render(request, "maintenance/os_board.html", {
        "pendentes": pendentes,
        "em_andamento": em_andamento,
        "concluidas": concluidas,
        "canceladas": canceladas,
        "counts": counts,
        "active_tab": active_tab,
        "search": search,
        "setor_id": int(setor_id) if setor_id.isdigit() else "",
        "criticidade": criticidade,
        "sectors": sectors,
        "active_technicians": active_technicians,
        "available_technicians": available_technicians,
        "technician_proprio": technician_proprio,
        "technician_proprio_id": technician_proprio.id if technician_proprio else None,
        "user_can_manage": _user_is_lider_ou_operador(request.user),
        "user_can_create_os": _user_can_create_os(request.user),
    })


@login_required
def os_assign_technician(request, os_id):
    """
    Atribui ou altera o técnico designado para a Ordem de Serviço física.
    """
    if request.method != "POST":
        return redirect('os_board')

    if not (_user_is_lider_ou_operador(request.user) or _user_can_create_os(request.user)):
        messages.error(request, "Acesso restrito. Você não possui permissão para atribuir técnicos.")
        return redirect('os_board')

    os_obj = get_object_or_404(OrdemServico, id=os_id)
    technician_id = request.POST.get('technician_id', '').strip()

    if not technician_id:
        os_obj.tecnico_designado = None
        os_obj.save()
        messages.success(request, f"Atribuição de técnico removida da OS #{os_obj.numero_os}.")
    else:
        tech = get_object_or_404(Technician, id=technician_id)
        if tech.is_ausente:
            messages.warning(request, f"O técnico {tech.nome} está ausente ({tech.get_status_display()}) e não pode ser designado.")
        else:
            os_obj.tecnico_designado = tech
            os_obj.save()
            messages.success(request, f"Técnico {tech.nome} designado para a OS #{os_obj.numero_os}.")

    return redirect(f"{reverse('os_board')}?tab=pendentes")


@login_required
def os_start_service(request, os_id):
    """
    Inicia o atendimento de uma OS física pela fila:
    - Cria alocação individual
    - Muda status da OS para EM_ANDAMENTO
    - Muda status do técnico para EM_ATENDIMENTO
    - Valida ausência e concorrência estrita (apenas 1 atendimento ativo por técnico)
    """
    if request.method != "POST":
        return redirect('os_board')

    os_obj = get_object_or_404(OrdemServico, id=os_id)

    # Identificar técnico
    tech_id = request.POST.get('technician_id', '').strip()
    if tech_id and _user_is_lider_ou_operador(request.user):
        tech = get_object_or_404(Technician, id=tech_id)
    else:
        tech = _get_technician_proprio(request.user)

    if not tech:
        messages.error(request, "Nenhum técnico selecionado ou vinculado ao seu usuário.")
        return redirect('os_board')

    # Validação 1: Ausência
    if tech.is_ausente:
        messages.error(request, f"O técnico {tech.nome} está ausente ({tech.get_status_display()}) e não pode iniciar atendimentos.")
        return redirect('os_board')

    # Validação 2: Concorrência estrita (apenas 1 alocação EM_ATENDIMENTO ativa por vez)
    active_alloc = tech.active_allocation
    if active_alloc:
        maq_nome = active_alloc.maquina.nome if active_alloc.maquina else "outra máquina"
        messages.error(
            request, 
            f"O técnico {tech.nome} já está em atendimento na {maq_nome}. É necessário pausar ou finalizar o atendimento atual antes de iniciar uma nova OS."
        )
        return redirect('os_board')

    # Criação da Alocação
    Allocation.objects.create(
        tecnico=tech,
        maquina=os_obj.maquina,
        ordem_servico=os_obj,
        atividade_observacao=f"OS #{os_obj.numero_os} - {os_obj.descricao_falha or os_obj.motivo or 'Atendimento de Manutenção'}",
        data_inicio=timezone.now(),
        status='EM_ATENDIMENTO',
        usuario_operador=request.user
    )

    # Atualiza técnico e OS
    tech.status = 'EM_ATENDIMENTO'
    tech.save()

    os_obj.status = 'EM_ANDAMENTO'
    if not os_obj.data_hora_inicio_conserto:
        os_obj.data_hora_inicio_conserto = timezone.now()
    if not os_obj.tecnico_designado:
        os_obj.tecnico_designado = tech
    os_obj.save()

    messages.success(request, f"Atendimento da OS #{os_obj.numero_os} iniciado pelo técnico {tech.nome} com sucesso!")
    return redirect('technician_management')


@login_required
def os_join_team(request, os_id):
    """
    Permite que um 2º ou 3º técnico entre na equipe de uma OS já em andamento:
    - Cria alocação individual vinculada à mesma OS
    - Permite trabalho conjunto mantendo apontamentos de tempo e pausas individuais
    """
    if request.method != "POST":
        return redirect('os_board')

    os_obj = get_object_or_404(OrdemServico, id=os_id)

    # Identificar técnico
    tech_id = request.POST.get('technician_id', '').strip()
    if tech_id and _user_is_lider_ou_operador(request.user):
        tech = get_object_or_404(Technician, id=tech_id)
    else:
        tech = _get_technician_proprio(request.user)

    if not tech:
        messages.error(request, "Nenhum técnico selecionado ou vinculado ao seu usuário.")
        return redirect('os_board')

    # Validação 1: Ausência
    if tech.is_ausente:
        messages.error(request, f"O técnico {tech.nome} está ausente ({tech.get_status_display()}) e não pode entrar na equipe.")
        return redirect('os_board')

    # Validação 2: Já alocado nesta mesma OS
    if os_obj.allocations.filter(tecnico=tech, data_fim__isnull=True).exists():
        messages.warning(request, f"O técnico {tech.nome} já possui uma alocação ativa ou pausada nesta OS.")
        return redirect(f"{reverse('os_board')}?tab=em_andamento")

    # Validação 3: Concorrência
    active_alloc = tech.active_allocation
    if active_alloc:
        maq_nome = active_alloc.maquina.nome if active_alloc.maquina else "outra máquina"
        messages.error(
            request, 
            f"O técnico {tech.nome} já está em atendimento na {maq_nome}. É necessário pausar ou finalizar o atendimento atual antes de entrar nesta OS."
        )
        return redirect(f"{reverse('os_board')}?tab=em_andamento")

    # Criação da Alocação de Trabalho em Equipe
    Allocation.objects.create(
        tecnico=tech,
        maquina=os_obj.maquina,
        ordem_servico=os_obj,
        atividade_observacao=f"OS #{os_obj.numero_os} (Trabalho em Equipe) - {os_obj.descricao_falha or os_obj.motivo or 'Apoio Manutenção'}",
        data_inicio=timezone.now(),
        status='EM_ATENDIMENTO',
        usuario_operador=request.user
    )

    tech.status = 'EM_ATENDIMENTO'
    tech.save()

    os_obj.status = 'EM_ANDAMENTO'
    os_obj.save()

    messages.success(request, f"Técnico {tech.nome} entrou na equipe da OS #{os_obj.numero_os} com sucesso!")
    return redirect('technician_management')


@login_required
def os_cancel(request, os_id):
    """
    Cancela uma Ordem de Serviço física (desde que não possua atendimentos em andamento).
    """
    if request.method != "POST":
        return redirect('os_board')

    if not (_user_is_lider_ou_operador(request.user) or _user_can_create_os(request.user)):
        messages.error(request, "Acesso restrito. Você não possui permissão para cancelar Ordens de Serviço.")
        return redirect('os_board')

    os_obj = get_object_or_404(OrdemServico, id=os_id)

    # Bloqueia se houver alocações ativas
    if os_obj.allocations.filter(data_fim__isnull=True).exists():
        messages.error(
            request, 
            f"Não é possível cancelar a OS #{os_obj.numero_os} pois há técnicos com atendimento em andamento. Pause ou conclua os atendimentos primeiro."
        )
        return redirect(f"{reverse('os_board')}?tab=em_andamento")

    os_obj.status = 'CANCELADA'
    os_obj.save()

    messages.success(request, f"Ordem de Serviço #{os_obj.numero_os} cancelada com sucesso.")
    return redirect(f"{reverse('os_board')}?tab=canceladas")


@login_required
def os_detail(request, pk):
    """
    Tela de detalhes, auditoria visual e histórico completo da Ordem de Serviço física:
    - Comparativo lado a lado da foto de abertura x foto de conclusão assinada
    - Auditoria de mão de obra de todos os técnicos que atuaram
    - Apontamentos de tempos líquidos, homem-hora, tempo de máquina parada
    - Histórico de peças e notas parciais de progresso
    """
    os_obj = get_object_or_404(
        OrdemServico.objects.select_related('maquina', 'setor', 'tecnico_designado', 'criado_por')
        .prefetch_related('allocations__tecnico', 'allocations__pausas', 'pecas_utilizadas'),
        pk=pk
    )
    return render(request, "maintenance/os_detail.html", {"os": os_obj})


@login_required
def link_allocation_os(request, allocation_id):
    """
    Permite vincular um atendimento emergencial (que foi iniciado avulso no sistema)
    a uma folha de Ordem de Serviço física recebida posteriormente.
    """
    if request.method != "POST":
        return redirect('technician_management')

    alloc = get_object_or_404(Allocation, id=allocation_id)
    os_id = request.POST.get('os_id', '').strip()

    if not os_id:
        messages.error(request, "Selecione uma Ordem de Serviço para vincular.")
        return redirect('technician_management')

    os_obj = get_object_or_404(OrdemServico, id=os_id)
    alloc.ordem_servico = os_obj
    alloc.save()

    os_obj.status = 'EM_ANDAMENTO'
    if not os_obj.data_hora_inicio_conserto:
        os_obj.data_hora_inicio_conserto = alloc.data_inicio
    os_obj.save()

    messages.success(request, f"Atendimento da máquina {alloc.maquina.nome if alloc.maquina else ''} vinculado à OS #{os_obj.numero_os} com sucesso!")
    return redirect('technician_management')




