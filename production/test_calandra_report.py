import io
from datetime import datetime, timedelta, time as dt_time
from unittest.mock import patch, MagicMock
import openpyxl

from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth.models import User, Group
from django.utils import timezone
from django.db import connections

from production.models import (
    ScadaDataPoint,
    ScadaPointValue,
    ScadaPointValueAnnotation,
)
from production.services_calandra import (
    CalandraHistoricalService,
    CALANDRA_VARIABLES_CONFIG,
)
from production.tests import init_scada_test_tables


class CalandraReportTestCase(TestCase):
    """
    Suíte completa de testes automatizados para a Central de Relatórios de Máquinas,
    Relatório Histórico da Calandra e Exportação Excel.
    """

    databases = {"default", "scada"}

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        init_scada_test_tables()

    def setUp(self):
        self.client = Client()

        # Grupos padrão
        self.grp_lp, _ = Group.objects.get_or_create(name="Liderança de Produção")
        self.grp_op, _ = Group.objects.get_or_create(name="Operadores")
        self.grp_pcp, _ = Group.objects.get_or_create(name="PCP")
        self.grp_tec, _ = Group.objects.get_or_create(name="Tecnicos")
        self.grp_tv, _ = Group.objects.get_or_create(name="Visualizador")

        # Usuários para matriz de permissões com nomes exclusivos
        self.user_lp = User.objects.create_user("calandra_user_lp", "lp_cal@test.com", "pass123")
        self.user_lp.groups.add(self.grp_lp)

        self.user_op = User.objects.create_user("calandra_user_op", "op_cal@test.com", "pass123")
        self.user_op.groups.add(self.grp_op)

        self.user_pcp = User.objects.create_user("calandra_user_pcp", "pcp_cal@test.com", "pass123")
        self.user_pcp.groups.add(self.grp_pcp)

        self.user_admin = User.objects.create_superuser("calandra_user_admin", "admin_cal@test.com", "pass123")

        self.user_tec = User.objects.create_user("calandra_user_tec", "tec_cal@test.com", "pass123")
        self.user_tec.groups.add(self.grp_tec)

        self.user_tv = User.objects.create_user("calandra_user_tv", "tv_cal@test.com", "pass123")
        self.user_tv.groups.add(self.grp_tv)

        # Popular DataPoints da Calandra no banco de teste 'scada' sem afetar outros testes
        with connections["scada"].cursor() as cursor:
            cursor.execute("DELETE FROM pointvalues WHERE dataPointId >= 100 AND dataPointId <= 200;")
            cursor.execute("DELETE FROM datapoints WHERE id >= 100 AND id <= 200;")

            self.dp_map = {}
            for idx, var in enumerate(CALANDRA_VARIABLES_CONFIG, start=100):
                cursor.execute(
                    "INSERT INTO datapoints (id, xid, dataSourceId, pointName, plcAlarmLevel) VALUES (%s, %s, %s, %s, %s);",
                    (idx, var["tag_name"], 1, var["tag_name"], 0)
                )
                self.dp_map[var["key"]] = idx

    # ─────────────────────────────────────────────────────────────────────────
    # 1. TESTES DE PERMISSÕES E ACESSO
    # ─────────────────────────────────────────────────────────────────────────

    def test_hub_access_authorized_users(self):
        """Usuários autorizados (Líder Produção, Operador, PCP, Admin) acessam a Central de Relatórios."""
        for user in [self.user_lp, self.user_op, self.user_pcp, self.user_admin]:
            self.client.force_login(user)
            resp = self.client.get(reverse("production:machine_reports_hub"))
            self.assertEqual(resp.status_code, 200)
            self.assertContains(resp, "Central de Relatórios de Máquinas")
            self.assertContains(resp, "Calandra")

    def test_hub_access_unauthorized_users(self):
        """Usuários não autorizados (Técnico Manutenção pura, TV) são bloqueados."""
        for user in [self.user_tec, self.user_tv]:
            self.client.force_login(user)
            resp = self.client.get(reverse("production:machine_reports_hub"))
            self.assertEqual(resp.status_code, 302)

    def test_calandra_report_access_permissions(self):
        """Acesso à tela de relatório da Calandra é restrito a perfis autorizados."""
        self.client.force_login(self.user_lp)
        resp = self.client.get(reverse("production:calandra_report"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Calandra")
        self.assertContains(resp, "Auditoria de Condições de Processo")

        self.client.force_login(self.user_tec)
        resp_blocked = self.client.get(reverse("production:calandra_report"))
        self.assertEqual(resp_blocked.status_code, 302)

    def test_calandra_excel_export_permissions(self):
        """Acesso ao endpoint de exportação Excel é restrito a perfis autorizados."""
        self.client.force_login(self.user_lp)
        resp = self.client.get(reverse("production:calandra_export_excel"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        self.client.force_login(self.user_tec)
        resp_blocked = self.client.get(reverse("production:calandra_export_excel"))
        self.assertEqual(resp_blocked.status_code, 302)

    # ─────────────────────────────────────────────────────────────────────────
    # 2. TESTES DE FILTRO DE PERÍODO
    # ─────────────────────────────────────────────────────────────────────────

    def test_period_filter_presets(self):
        """Valida os atalhos de período: hoje, ontem, 7d, 30d."""
        now = timezone.localtime()

        # Hoje
        s_dt, e_dt, act, err = CalandraHistoricalService.parse_period_filters(periodo="hoje")
        self.assertEqual(act, "hoje")
        self.assertIsNone(err)
        self.assertEqual(s_dt.date(), now.date())

        # Ontem
        s_dt, e_dt, act, err = CalandraHistoricalService.parse_period_filters(periodo="ontem")
        self.assertEqual(act, "ontem")
        self.assertIsNone(err)
        self.assertEqual(s_dt.date(), (now - timedelta(days=1)).date())
        self.assertEqual(e_dt.date(), (now - timedelta(days=1)).date())

        # 7d
        s_dt, e_dt, act, err = CalandraHistoricalService.parse_period_filters(periodo="7d")
        self.assertEqual(act, "7d")
        self.assertIsNone(err)
        self.assertEqual(s_dt.date(), (now - timedelta(days=7)).date())

        # 30d
        s_dt, e_dt, act, err = CalandraHistoricalService.parse_period_filters(periodo="30d")
        self.assertEqual(act, "30d")
        self.assertIsNone(err)
        self.assertEqual(s_dt.date(), (now - timedelta(days=30)).date())

    def test_period_filter_custom_and_validation(self):
        """Valida período personalizado e tratamento de datas invertidas ou excessivas."""
        s_dt, e_dt, act, err = CalandraHistoricalService.parse_period_filters(
            periodo="personalizado",
            data_inicio="2026-08-10",
            hora_inicio="08:00",
            data_final="2026-08-15",
            hora_final="18:00",
        )
        self.assertEqual(act, "personalizado")
        self.assertIsNone(err)
        self.assertEqual(s_dt.strftime("%Y-%m-%d %H:%M"), "2026-08-10 08:00")
        self.assertEqual(e_dt.strftime("%Y-%m-%d %H:%M"), "2026-08-15 18:00")

        # Início maior que fim -> emite aviso e recupera hoje
        s_inv, e_inv, act_inv, err_inv = CalandraHistoricalService.parse_period_filters(
            periodo="personalizado",
            data_inicio="2026-08-20",
            data_final="2026-08-10",
        )
        self.assertIsNotNone(err_inv)
        self.assertEqual(act_inv, "hoje")

        # Período excessivo (> 31 dias) -> trunca para 31 dias com aviso
        s_exc, e_exc, act_exc, err_exc = CalandraHistoricalService.parse_period_filters(
            periodo="personalizado",
            data_inicio="2026-06-01",
            data_final="2026-08-01",
        )
        self.assertIsNotNone(err_exc)
        self.assertIn("31 dias", err_exc)
        self.assertEqual((e_exc - s_exc).days, 31)

    # ─────────────────────────────────────────────────────────────────────────
    # 3. TESTES DE HISTÓRICO, FORWARD-FILL E PASSADAS
    # ─────────────────────────────────────────────────────────────────────────

    def test_passada_formatting(self):
        """Valida a conversão semântica da variável PASSADA em rótulos contextuais padronizados."""
        self.assertEqual(CalandraHistoricalService.format_passada_label(1), "PASSADA 1 — 1ª face")
        self.assertEqual(CalandraHistoricalService.format_passada_label("1"), "PASSADA 1 — 1ª face")
        self.assertEqual(CalandraHistoricalService.format_passada_label(1.0), "PASSADA 1 — 1ª face")
        self.assertEqual(CalandraHistoricalService.format_passada_label(2), "PASSADA 2 — 2ª face / face oposta")
        self.assertEqual(CalandraHistoricalService.format_passada_label("2"), "PASSADA 2 — 2ª face / face oposta")
        self.assertEqual(CalandraHistoricalService.format_passada_label(None), "Não informada")

    def test_effective_process_rule_6_scenarios(self):
        """
        Valida os 6 cenários obrigatórios da regra de Processo Efetivo (Velocidade + Metragem):
        1. vel > 0 + metragem aumentando -> True
        2. vel = 0 + metragem parada -> False
        3. vel > 0 + metragem estagnada/sem avanço -> False
        4. vel = 0 + metragem aumentando -> False
        5. reinício/reset de metragem -> True (com avanço posterior)
        6. pequenas flutuações/ruído sem produção real -> False
        """
        # Cenário 1: vel > 0 + metragem aumentando
        tl_1 = [
            {"ts": 1000, "values": {"vel_calandra": 12.0, "metragem_bobinada": 100.0}},
            {"ts": 2000, "values": {"vel_calandra": 12.5, "metragem_bobinada": 120.0}},
        ]
        CalandraHistoricalService.detect_effective_process(tl_1)
        self.assertTrue(tl_1[0]["is_effective"])
        self.assertTrue(tl_1[1]["is_effective"])

        # Cenário 2: vel = 0 + metragem parada
        tl_2 = [
            {"ts": 1000, "values": {"vel_calandra": 0.0, "metragem_bobinada": 120.0}},
            {"ts": 2000, "values": {"vel_calandra": 0.0, "metragem_bobinada": 120.0}},
        ]
        CalandraHistoricalService.detect_effective_process(tl_2)
        self.assertFalse(tl_2[0]["is_effective"])
        self.assertFalse(tl_2[1]["is_effective"])

        # Cenário 3: vel > 0 + metragem sem avanço (estagnada)
        tl_3 = [
            {"ts": 1000, "values": {"vel_calandra": 10.0, "metragem_bobinada": 50.0}},
            {"ts": 2000, "values": {"vel_calandra": 10.0, "metragem_bobinada": 50.0}},
            {"ts": 3000, "values": {"vel_calandra": 10.0, "metragem_bobinada": 50.0}},
        ]
        CalandraHistoricalService.detect_effective_process(tl_3)
        self.assertFalse(tl_3[0]["is_effective"])
        self.assertFalse(tl_3[1]["is_effective"])
        self.assertFalse(tl_3[2]["is_effective"])

        # Cenário 4: vel = 0 + metragem aumentando (ruído/estática)
        tl_4 = [
            {"ts": 1000, "values": {"vel_calandra": 0.0, "metragem_bobinada": 50.0}},
            {"ts": 2000, "values": {"vel_calandra": 0.0, "metragem_bobinada": 70.0}},
        ]
        CalandraHistoricalService.detect_effective_process(tl_4)
        self.assertFalse(tl_4[0]["is_effective"])
        self.assertFalse(tl_4[1]["is_effective"])

        # Cenário 5: reinício/reset de metragem (troca de bobina com produção contínua)
        tl_5 = [
            {"ts": 1000, "values": {"vel_calandra": 15.0, "metragem_bobinada": 850.0}},
            {"ts": 2000, "values": {"vel_calandra": 15.0, "metragem_bobinada": 0.0}},  # Reset de bobina
            {"ts": 3000, "values": {"vel_calandra": 15.0, "metragem_bobinada": 25.0}}, # Avanço
        ]
        CalandraHistoricalService.detect_effective_process(tl_5)
        self.assertTrue(tl_5[0]["is_effective"])
        self.assertTrue(tl_5[1]["is_effective"])
        self.assertTrue(tl_5[2]["is_effective"])

        # Cenário 6: pequenas flutuações/ruídos (v <= 0.05)
        tl_6 = [
            {"ts": 1000, "values": {"vel_calandra": 0.02, "metragem_bobinada": 10.0}},
            {"ts": 2000, "values": {"vel_calandra": 0.03, "metragem_bobinada": 10.0}},
        ]
        CalandraHistoricalService.detect_effective_process(tl_6)
        self.assertFalse(tl_6[0]["is_effective"])
        self.assertFalse(tl_6[1]["is_effective"])

    def test_audit_cards_stats_computation(self):
        """Valida o cálculo correto de média, mínimo e máximo dos 4 cards sobre processo efetivo."""
        timeline = [
            # Ponto 1: Em processo efetivo
            {
                "ts": 1000,
                "is_effective": True,
                "values": {
                    "vel_calandra": 10.0, "metragem_bobinada": 10.0,
                    "temp_borracha_saida_extrusao": 80.0,
                    "temp_borracha_ent_calandra": 40.0,
                    "temp_borracha_saida_calandra": 75.0,
                    "temp_cilindro_inf": 60.0,
                    "temp_cilindro_inter": 65.0,
                    "temp_cilindro_sup": 70.0,
                    "carga_desbobinador": 50.0,
                    "carga_quebra_trama": 200.0,
                    "carga_pos_calandra": 220.0,
                    "carga_bobinamento": 180.0,
                    "temp_furador": 30.0,
                    "temp_aquecedor": 45.0,
                    "temp_tcu_extrusora": 55.0,
                    "temp_geladeira": 18.0,
                }
            },
            # Ponto 2: Em processo efetivo
            {
                "ts": 2000,
                "is_effective": True,
                "values": {
                    "vel_calandra": 12.0, "metragem_bobinada": 30.0,
                    "temp_borracha_saida_extrusao": 90.0,
                    "temp_borracha_ent_calandra": 42.0,
                    "temp_borracha_saida_calandra": 85.0,
                    "temp_cilindro_inf": 64.0,
                    "temp_cilindro_inter": 67.0,
                    "temp_cilindro_sup": 72.0,
                    "carga_desbobinador": 70.0,
                    "carga_quebra_trama": 210.0,
                    "carga_pos_calandra": 230.0,
                    "carga_bobinamento": 190.0,
                    "temp_furador": 32.0,
                    "temp_aquecedor": 47.0,
                    "temp_tcu_extrusora": 57.0,
                    "temp_geladeira": 22.0,
                }
            },
            # Ponto 3: Máquina parada (NÃO deve distorcer as médias dos cards!)
            {
                "ts": 3000,
                "is_effective": False,
                "values": {
                    "vel_calandra": 0.0, "metragem_bobinada": 30.0,
                    "temp_borracha_saida_extrusao": 20.0, # Valor frio de parada
                    "temp_cilindro_inf": 25.0,
                    "carga_desbobinador": 0.0,
                    "temp_geladeira": 10.0,
                }
            },
        ]

        stats = CalandraHistoricalService.compute_effective_process_stats(timeline)
        self.assertTrue(stats["has_effective_process"])
        self.assertEqual(stats["effective_points_count"], 2)

        # Card 1: Temp Borracha (Médias de [80, 90] = 85.0, Mín 80.0, Máx 90.0)
        c1 = stats["temp_borracha"]["saida_extrusao"]
        self.assertEqual(c1["avg"], 85.0)
        self.assertEqual(c1["min"], 80.0)
        self.assertEqual(c1["max"], 90.0)

        # Card 2: Cilindros (Inf: [60, 64] -> 62.0; Inter: [65, 67] -> 66.0; Sup: [70, 72] -> 71.0)
        c2 = stats["temp_cilindros"]
        self.assertEqual(c2["cilindro_inf"]["avg"], 62.0)
        self.assertEqual(c2["cilindro_sup"]["avg"], 71.0)
        # Delta Cilindros = 71.0 - 62.0 = 9.0
        self.assertEqual(c2["delta_cilindros"], 9.0)

        # Card 3: Cargas (Desbobinador: [50, 70] -> 60.0)
        c3 = stats["cargas"]["desbobinador"]
        self.assertEqual(c3["avg"], 60.0)
        self.assertEqual(c3["min"], 50.0)
        self.assertEqual(c3["max"], 70.0)

        # Card 4: Temperaturas Auxiliares (Geladeira: [18, 22] -> 20.0)
        c4 = stats["temp_auxiliares"]
        self.assertEqual(c4["furador"]["avg"], 31.0)
        self.assertEqual(c4["aquecedor"]["avg"], 46.0)
        self.assertEqual(c4["tcu_extrusora"]["avg"], 56.0)
        self.assertEqual(c4["geladeira"]["avg"], 20.0)
        self.assertEqual(c4["geladeira"]["min"], 18.0)
        self.assertEqual(c4["geladeira"]["max"], 22.0)

    def test_audit_cards_no_effective_process(self):
        """Valida que período sem processo efetivo retorna has_effective_process=False sem divisão por zero."""
        timeline = [
            {"ts": 1000, "is_effective": False, "values": {"vel_calandra": 0.0, "temp_borracha_saida_extrusao": 25.0}},
            {"ts": 2000, "is_effective": False, "values": {"vel_calandra": 0.0, "temp_borracha_saida_extrusao": 24.0}},
        ]
        stats = CalandraHistoricalService.compute_effective_process_stats(timeline)
        self.assertFalse(stats["has_effective_process"])
        self.assertEqual(stats["effective_points_count"], 0)
        self.assertIsNone(stats["temp_borracha"]["saida_extrusao"]["avg"])

    def test_passada_window_context(self):
        """Valida o contexto semântico de passada ativa na janela."""
        t_p1 = [{"passada_val": 1.0}, {"passada_val": 1.0}]
        self.assertEqual(CalandraHistoricalService.get_passada_window_context(t_p1), "PASSADA 1 — 1ª face")

        t_p2 = [{"passada_val": 2.0}, {"passada_val": 2.0}]
        self.assertEqual(CalandraHistoricalService.get_passada_window_context(t_p2), "PASSADA 2 — 2ª face / face oposta")

        t_both = [{"passada_val": 1.0}, {"passada_val": 2.0}]
        self.assertEqual(CalandraHistoricalService.get_passada_window_context(t_both), "PASSADA 1 e PASSADA 2")

    def test_forward_fill_state_synchronization(self):
        """
        Valida que leituras assíncronas em diferentes timestamps são sincronizadas
        corretamente sem células vazias e refletem o estado conhecido no instante.
        """
        base_time = timezone.now()
        t0 = int((base_time - timedelta(minutes=10)).timestamp() * 1000)
        t1 = int((base_time - timedelta(minutes=8)).timestamp() * 1000)
        t2 = int((base_time - timedelta(minutes=5)).timestamp() * 1000)
        t3 = int((base_time - timedelta(minutes=2)).timestamp() * 1000)

        dp_passada = self.dp_map["passada"]
        dp_vel = self.dp_map["vel_calandra"]
        dp_carga_bob = self.dp_map["carga_bobinamento"]
        dp_temp_ext = self.dp_map["temp_borracha_saida_extrusao"]

        with connections["scada"].cursor() as cursor:
            # Em t0 (anterior ao filtro): seed inicial
            cursor.execute("INSERT INTO pointvalues (dataPointId, dataType, pointValue, ts) VALUES (%s, %s, %s, %s);", (dp_passada, 2, 1.0, t0))
            cursor.execute("INSERT INTO pointvalues (dataPointId, dataType, pointValue, ts) VALUES (%s, %s, %s, %s);", (dp_vel, 3, 10.5, t0))

            # Em t1: atualização da carga
            cursor.execute("INSERT INTO pointvalues (dataPointId, dataType, pointValue, ts) VALUES (%s, %s, %s, %s);", (dp_carga_bob, 3, 200.0, t1))

            # Em t2: atualização da temperatura
            cursor.execute("INSERT INTO pointvalues (dataPointId, dataType, pointValue, ts) VALUES (%s, %s, %s, %s);", (dp_temp_ext, 3, 85.2, t2))

            # Em t3: mudança de PASSADA 1 -> PASSADA 2
            cursor.execute("INSERT INTO pointvalues (dataPointId, dataType, pointValue, ts) VALUES (%s, %s, %s, %s);", (dp_passada, 2, 2.0, t3))

        start_dt = base_time - timedelta(minutes=9)
        end_dt = base_time

        history = CalandraHistoricalService.get_synchronized_history(start_dt, end_dt)
        timeline = history["timeline"]

        self.assertEqual(len(timeline), 3)  # t1, t2, t3

        # Ponto 1 (t1): deve herdar vel=10.5 e passada=1 do seed, e conter carga_bob=200.0
        row1 = timeline[0]
        self.assertIn("PASSADA 1", row1["passada_label"])
        self.assertEqual(row1["values"]["vel_calandra"], 10.5)
        self.assertEqual(row1["values"]["carga_bobinamento"], 200.0)

        # Ponto 2 (t2): deve manter vel=10.5, carga_bob=200.0, passada=1 e atualizar temp=85.2
        row2 = timeline[1]
        self.assertIn("PASSADA 1", row2["passada_label"])
        self.assertEqual(row2["values"]["vel_calandra"], 10.5)
        self.assertEqual(row2["values"]["carga_bobinamento"], 200.0)
        self.assertEqual(row2["values"]["temp_borracha_saida_extrusao"], 85.2)

        # Ponto 3 (t3): transição para PASSADA 2 (2ª face / face oposta)
        row3 = timeline[2]
        self.assertIn("PASSADA 2", row3["passada_label"])
        self.assertEqual(row3["values"]["vel_calandra"], 10.5)
        self.assertEqual(row3["values"]["temp_borracha_saida_extrusao"], 85.2)

    def test_chart_datasets_structure(self):
        """Valida que os 6 datasets de gráficos contêm as chaves e séries corretas (com Geladeira incluída em Auxiliares)."""
        start_dt = timezone.now() - timedelta(hours=1)
        end_dt = timezone.now()

        history = CalandraHistoricalService.get_synchronized_history(start_dt, end_dt)
        charts = history["chart_datasets"]

        self.assertIn("chart_1_producao", charts)
        self.assertIn("chart_2_cargas", charts)
        self.assertIn("chart_3_espessuras", charts)
        self.assertIn("chart_4_temp_borracha", charts)
        self.assertIn("chart_5_temp_cilindros", charts)
        self.assertIn("chart_6_temp_auxiliares", charts)

        # Gráfico 1: Produção
        self.assertIn("velocidade", charts["chart_1_producao"])
        self.assertIn("metragem", charts["chart_1_producao"])
        self.assertIn("passada", charts["chart_1_producao"])

        # Gráfico 2: Cargas
        self.assertIn("bobinamento", charts["chart_2_cargas"])
        self.assertIn("desbobinador", charts["chart_2_cargas"])
        self.assertIn("pos_calandra", charts["chart_2_cargas"])
        self.assertIn("quebra_trama", charts["chart_2_cargas"])

        # Gráfico 3: Espessuras
        self.assertIn("esq_sup", charts["chart_3_espessuras"])
        self.assertIn("dir_sup", charts["chart_3_espessuras"])

        # Gráfico 4: Temp Borracha
        self.assertIn("saida_extrusao", charts["chart_4_temp_borracha"])
        self.assertIn("ent_calandra", charts["chart_4_temp_borracha"])
        self.assertIn("saida_calandra", charts["chart_4_temp_borracha"])

        # Gráfico 5: Temp Cilindros
        self.assertIn("cilindro_inf", charts["chart_5_temp_cilindros"])
        self.assertIn("cilindro_inter", charts["chart_5_temp_cilindros"])
        self.assertIn("cilindro_sup", charts["chart_5_temp_cilindros"])

        # Gráfico 6: Temp Auxiliares (Furador, Aquecedor, TCU Extrusora e Geladeira)
        self.assertIn("furador", charts["chart_6_temp_auxiliares"])
        self.assertIn("aquecedor", charts["chart_6_temp_auxiliares"])
        self.assertIn("tcu_extrusora", charts["chart_6_temp_auxiliares"])
        self.assertIn("geladeira", charts["chart_6_temp_auxiliares"])

    # ─────────────────────────────────────────────────────────────────────────
    # 4. TESTES DE GERAÇÃO EXCEL (.XLSX)
    # ─────────────────────────────────────────────────────────────────────────

    def test_excel_export_structure_and_types(self):
        """Valida que o arquivo Excel gerado segue rigorosamente as regras de cabeçalho, tipos e colunas (22 colunas)."""
        base_time = timezone.now()
        t1 = int((base_time - timedelta(minutes=5)).timestamp() * 1000)
        t2 = int((base_time - timedelta(minutes=2)).timestamp() * 1000)

        with connections["scada"].cursor() as cursor:
            cursor.execute("INSERT INTO pointvalues (dataPointId, dataType, pointValue, ts) VALUES (%s, %s, %s, %s);", (self.dp_map["passada"], 2, 1.0, t1))
            cursor.execute("INSERT INTO pointvalues (dataPointId, dataType, pointValue, ts) VALUES (%s, %s, %s, %s);", (self.dp_map["vel_calandra"], 3, 12.5, t1))
            cursor.execute("INSERT INTO pointvalues (dataPointId, dataType, pointValue, ts) VALUES (%s, %s, %s, %s);", (self.dp_map["metragem_bobinada"], 3, 540.0, t1))
            cursor.execute("INSERT INTO pointvalues (dataPointId, dataType, pointValue, ts) VALUES (%s, %s, %s, %s);", (self.dp_map["temp_geladeira"], 3, 18.5, t1))
            cursor.execute("INSERT INTO pointvalues (dataPointId, dataType, pointValue, ts) VALUES (%s, %s, %s, %s);", (self.dp_map["passada"], 2, 2.0, t2))
            cursor.execute("INSERT INTO pointvalues (dataPointId, dataType, pointValue, ts) VALUES (%s, %s, %s, %s);", (self.dp_map["vel_calandra"], 3, 14.0, t2))
            cursor.execute("INSERT INTO pointvalues (dataPointId, dataType, pointValue, ts) VALUES (%s, %s, %s, %s);", (self.dp_map["temp_geladeira"], 3, 19.0, t2))

        start_dt = base_time - timedelta(minutes=10)
        end_dt = base_time

        excel_bytes = CalandraHistoricalService.generate_excel_report(start_dt, end_dt)
        self.assertTrue(len(excel_bytes) > 0)

        # Ler Excel com openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        # Validar cabeçalho
        self.assertEqual(ws.cell(row=1, column=1).value, "DATA/HORA")
        self.assertEqual(ws.cell(row=1, column=2).value, "PASSADA")
        self.assertEqual(ws.cell(row=1, column=3).value, "METRAGEM BOBINADA (m)")
        self.assertEqual(ws.cell(row=1, column=4).value, "VEL. CALANDRA (m/min)")
        # Última coluna: TEMP. GELADEIRA (°C)
        self.assertEqual(ws.cell(row=1, column=22).value, "TEMP. GELADEIRA (°C)")

        # Total de colunas: 1 (Data/Hora) + 1 (Passada) + 20 (demais variáveis) = 22 colunas
        self.assertEqual(ws.max_column, 22)

        # Linhas de dados: 2 registros
        self.assertEqual(ws.max_row, 3)  # 1 cabeçalho + 2 dados

        # Linha 1 de dados (row=2)
        row1_passada = ws.cell(row=2, column=2).value
        self.assertIn("PASSADA 1", row1_passada)
        row1_metragem = ws.cell(row=2, column=3).value
        self.assertEqual(row1_metragem, 540.0)
        self.assertIsInstance(row1_metragem, (int, float))  # Numérico real!
        row1_gel = ws.cell(row=2, column=22).value
        self.assertEqual(row1_gel, 18.5)

        # Linha 2 de dados (row=3)
        row2_passada = ws.cell(row=3, column=2).value
        self.assertIn("PASSADA 2", row2_passada)
        row2_vel = ws.cell(row=3, column=4).value
        self.assertEqual(row2_vel, 14.0)
        self.assertIsInstance(row2_vel, (int, float))
        row2_gel = ws.cell(row=3, column=22).value
        self.assertEqual(row2_gel, 19.0)

        # Congelamento e autofiltro
        self.assertEqual(ws.freeze_panes, "A2")

    # ─────────────────────────────────────────────────────────────────────────
    # 5. TESTES DE SEGURANÇA (READ-ONLY DO SCADA)
    # ─────────────────────────────────────────────────────────────────────────

    def test_scada_database_unmodified(self):
        """Valida que nenhuma alteração ou escrita ocorreu no banco 'scada'."""
        with connections["scada"].cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM datapoints;")
            count_before = cursor.fetchone()[0]

        # Executar fluxo completo de consulta e exportação
        start_dt = timezone.now() - timedelta(days=1)
        end_dt = timezone.now()
        _ = CalandraHistoricalService.get_synchronized_history(start_dt, end_dt)
        _ = CalandraHistoricalService.generate_excel_report(start_dt, end_dt)

        with connections["scada"].cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM datapoints;")
            count_after = cursor.fetchone()[0]

        self.assertEqual(count_before, count_after)

    # ─────────────────────────────────────────────────────────────────────────
    # 6. TESTES DA TELA DE CONFIGURAÇÃO DE XIDs DA CALANDRA
    # ─────────────────────────────────────────────────────────────────────────

    def test_xid_calandra_config_permissions(self):
        """Acesso à tela de configuração de XIDs da Calandra é exclusivo para superusuários."""
        self.client.force_login(self.user_admin)
        resp = self.client.get(reverse("production:xid_calandra_config"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Configuração de XIDs da Calandra")

        # Usuários não superuser são bloqueados
        for u in [self.user_lp, self.user_op, self.user_tec, self.user_tv]:
            self.client.force_login(u)
            resp_blocked = self.client.get(reverse("production:xid_calandra_config"))
            self.assertEqual(resp_blocked.status_code, 302)

    def test_xid_calandra_config_save_and_dynamic_override(self):
        """Valida que salvar XIDs customizados atualiza o banco default e o serviço lê o override."""
        self.client.force_login(self.user_admin)

        post_data = {
            "action": "save_calandra_xids",
            "xid_passada": "CUSTOM_TAG_PASSADA",
            "xid_vel_calandra": "CUSTOM_TAG_VEL",
        }
        for var in CALANDRA_VARIABLES_CONFIG:
            if var["key"] not in ["passada", "vel_calandra"]:
                post_data[f"xid_{var['key']}"] = var["tag_name"]

        resp = self.client.post(reverse("production:xid_calandra_config"), post_data, follow=True)
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Configurações de XIDs da Calandra salvas com sucesso!")

        # Validar override no serviço
        configs = CalandraHistoricalService.get_variables_config()
        cfg_map = {c["key"]: c["tag_name"] for c in configs}
        self.assertEqual(cfg_map["passada"], "CUSTOM_TAG_PASSADA")
        self.assertEqual(cfg_map["vel_calandra"], "CUSTOM_TAG_VEL")

        # Testar restauração de padrões
        resp_restore = self.client.post(reverse("production:xid_calandra_config"), {"action": "restore_defaults"}, follow=True)
        self.assertEqual(resp_restore.status_code, 200)
        self.assertContains(resp_restore, "restauradas para os padrões canônicos")

        configs_restored = CalandraHistoricalService.get_variables_config()
        cfg_restored_map = {c["key"]: c["tag_name"] for c in configs_restored}
        self.assertEqual(cfg_restored_map["passada"], "CALANDRA_meta - PASSADA")
        self.assertEqual(cfg_restored_map["vel_calandra"], "CALANDRA - VEL_CALANDRA (m/min)")

    # ─────────────────────────────────────────────────────────────────────────
    # 7. TESTES ESPECÍFICOS DA NOVA VARIÁVEL GELADEIRA & RENDERIZAÇÃO
    # ─────────────────────────────────────────────────────────────────────────

    def test_geladeira_variable_attributes(self):
        """Valida que a 21ª variável Geladeira possui todas as propriedades canônicas exigidas."""
        configs = CalandraHistoricalService.get_variables_config()
        self.assertEqual(len(configs), 21)

        gel_cfg = next((c for c in configs if c["key"] == "temp_geladeira"), None)
        self.assertIsNotNone(gel_cfg)
        self.assertIn("GELADEIRA", gel_cfg["tag_name"])
        self.assertEqual(gel_cfg["label"], "Geladeira")
        self.assertEqual(gel_cfg["excel_header"], "TEMP. GELADEIRA (°C)")
        self.assertEqual(gel_cfg["unit"], "°C")
        self.assertEqual(gel_cfg["data_type"], 3)
        self.assertEqual(gel_cfg["order"], 21)
        self.assertTrue(gel_cfg["is_numeric"])
        self.assertEqual(gel_cfg["group"], "temperaturas_processo")

    def test_calandra_report_renders_geladeira_and_21_vars(self):
        """Valida que a página de relatório renderiza o Card de Geladeira, Gráfico, Tabela e contagem de 21 XIDs."""
        self.client.force_login(self.user_lp)
        resp = self.client.get(reverse("production:calandra_report"))
        self.assertEqual(resp.status_code, 200)

        # Badge e contagem
        self.assertContains(resp, "21 XIDs")
        # Card 4
        self.assertContains(resp, "Geladeira")
        self.assertContains(resp, "c4_gel_avg")
        # Gráfico 6
        self.assertContains(resp, "chartTempAuxiliares")
        # Tabela
        self.assertContains(resp, "T. Geladeira (°C)")


