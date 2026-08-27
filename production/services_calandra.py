import io
import time
import math
from datetime import datetime, date, time as dt_time, timedelta
from typing import Dict, List, Any, Optional, Tuple
from django.utils import timezone
from django.db.models import Max, Q
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import ScadaDataPoint, ScadaPointValue, ScadaPointValueAnnotation
from .services import scada_reader


# ==============================================================================
# CANONICAL INVENTORY OF THE 20 CALANDRA VARIABLES
# ==============================================================================

CALANDRA_VARIABLES_CONFIG: List[Dict[str, Any]] = [
    # ── 1. Produção / Contexto ──
    {
        "key": "passada",
        "tag_name": "CALANDRA_meta - PASSADA",
        "label": "Passada",
        "excel_header": "PASSADA",
        "group": "producao",
        "group_label": "Produção",
        "unit": "",
        "data_type": 2,  # Multistate / Contextual
        "order": 1,
        "is_numeric": False,
    },
    {
        "key": "metragem_bobinada",
        "tag_name": "CALANDRA - METRAGEM_BOBINADA",
        "label": "Metragem Bobinada",
        "excel_header": "METRAGEM BOBINADA (m)",
        "group": "producao",
        "group_label": "Produção",
        "unit": "m",
        "data_type": 3,
        "order": 2,
        "is_numeric": True,
    },
    {
        "key": "vel_calandra",
        "tag_name": "CALANDRA - VEL_CALANDRA (m/min)",
        "label": "Velocidade da Calandra",
        "excel_header": "VEL. CALANDRA (m/min)",
        "group": "producao",
        "group_label": "Produção",
        "unit": "m/min",
        "data_type": 3,
        "order": 3,
        "is_numeric": True,
    },

    # ── 2. Cargas / Tensões ──
    {
        "key": "carga_bobinamento",
        "tag_name": "CALANDRA - CARGA_BOBINAMENTO (Kg)",
        "label": "Carga Bobinamento",
        "excel_header": "CARGA BOBINAMENTO (kg)",
        "group": "cargas",
        "group_label": "Cargas",
        "unit": "kg",
        "data_type": 3,
        "order": 4,
        "is_numeric": True,
    },
    {
        "key": "carga_desbobinador",
        "tag_name": "CALANDRA - CARGA_DESBOBINADOR (Kg)",
        "label": "Carga Desbobinador",
        "excel_header": "CARGA DESBOBINADOR (kg)",
        "group": "cargas",
        "group_label": "Cargas",
        "unit": "kg",
        "data_type": 3,
        "order": 5,
        "is_numeric": True,
    },
    {
        "key": "carga_pos_calandra",
        "tag_name": "CALANDRA - CARGA_POS-CALANDRA (Kg)",
        "label": "Carga Pós-Calandra",
        "excel_header": "CARGA PÓS-CALANDRA (kg)",
        "group": "cargas",
        "group_label": "Cargas",
        "unit": "kg",
        "data_type": 3,
        "order": 6,
        "is_numeric": True,
    },
    {
        "key": "carga_quebra_trama",
        "tag_name": "CALANDRA - CARGA_QUEBRA-TRAMA (Kg)",
        "label": "Carga Quebra-Trama",
        "excel_header": "CARGA QUEBRA-TRAMA (kg)",
        "group": "cargas",
        "group_label": "Cargas",
        "unit": "kg",
        "data_type": 3,
        "order": 7,
        "is_numeric": True,
    },

    # ── 3. Espessuras ──
    {
        "key": "espessura_esq_sup",
        "tag_name": "CALANDRA - ESPESSURA_LADO ESQ SUPERIOR",
        "label": "Espessura Esq. Superior",
        "excel_header": "ESPESSURA ESQ. SUPERIOR (mm)",
        "group": "espessuras",
        "group_label": "Espessuras",
        "unit": "mm",
        "data_type": 3,
        "order": 8,
        "is_numeric": True,
    },
    {
        "key": "espessura_dir_sup",
        "tag_name": "CALANDRA - ESPESSURA_LADO DIR SUPERIOR",
        "label": "Espessura Dir. Superior",
        "excel_header": "ESPESSURA DIR. SUPERIOR (mm)",
        "group": "espessuras",
        "group_label": "Espessuras",
        "unit": "mm",
        "data_type": 3,
        "order": 9,
        "is_numeric": True,
    },
    {
        "key": "espessura_dir_inf",
        "tag_name": "CALANDRA - ESPESSURA_LADO DIR INFERIOR",
        "label": "Espessura Dir. Inferior",
        "excel_header": "ESPESSURA DIR. INFERIOR (mm)",
        "group": "espessuras",
        "group_label": "Espessuras",
        "unit": "mm",
        "data_type": 3,
        "order": 10,
        "is_numeric": True,
    },
    {
        "key": "espessura_esq_inf",
        "tag_name": "CALANDRA - ESPESSURA_LADO ESQ INFERIOR",
        "label": "Espessura Esq. Inferior",
        "excel_header": "ESPESSURA ESQ. INFERIOR (mm)",
        "group": "espessuras",
        "group_label": "Espessuras",
        "unit": "mm",
        "data_type": 3,
        "order": 11,
        "is_numeric": True,
    },

    # ── 4. Temperaturas da Borracha ──
    {
        "key": "temp_borracha_saida_extrusao",
        "tag_name": "CALANDRA - TEMPERATURA_BORRACHA_SAIDA_EXTRUSAO (°C)",
        "label": "Saída Extrusão",
        "excel_header": "TEMP. BORRACHA SAÍDA EXTRUSÃO (°C)",
        "group": "temperatura_borracha",
        "group_label": "Temperatura da Borracha",
        "unit": "°C",
        "data_type": 3,
        "order": 12,
        "is_numeric": True,
    },
    {
        "key": "temp_borracha_ent_calandra",
        "tag_name": "CALANDRA - TEMPERATURA_BORRACHA_ENT_CALANDRA (°C)",
        "label": "Entrada Calandra",
        "excel_header": "TEMP. BORRACHA ENTRADA CALANDRA (°C)",
        "group": "temperatura_borracha",
        "group_label": "Temperatura da Borracha",
        "unit": "°C",
        "data_type": 3,
        "order": 13,
        "is_numeric": True,
    },
    {
        "key": "temp_borracha_saida_calandra",
        "tag_name": "CALANDRA - TEMPERATURA_BORRACHA_SAIDA_CALANDRA (°C)",
        "label": "Saída Calandra",
        "excel_header": "TEMP. BORRACHA SAÍDA CALANDRA (°C)",
        "group": "temperatura_borracha",
        "group_label": "Temperatura da Borracha",
        "unit": "°C",
        "data_type": 3,
        "order": 14,
        "is_numeric": True,
    },

    # ── 5. Temperaturas do Processo / Equipamento ──
    {
        "key": "temp_cilindro_inf",
        "tag_name": "CALANDRA_TEMPERATURA - CILINDRO_INFERIOR (°C)",
        "label": "Cilindro Inferior",
        "excel_header": "TEMP. CILINDRO INFERIOR (°C)",
        "group": "temperaturas_processo",
        "group_label": "Temperaturas do Processo",
        "unit": "°C",
        "data_type": 3,
        "order": 15,
        "is_numeric": True,
    },
    {
        "key": "temp_cilindro_inter",
        "tag_name": "CALANDRA_TEMPERATURA - CILINDRO_INTERMEDIÁRIO (°C)",
        "label": "Cilindro Intermediário",
        "excel_header": "TEMP. CILINDRO INTERMEDIÁRIO (°C)",
        "group": "temperaturas_processo",
        "group_label": "Temperaturas do Processo",
        "unit": "°C",
        "data_type": 3,
        "order": 16,
        "is_numeric": True,
    },
    {
        "key": "temp_cilindro_sup",
        "tag_name": "CALANDRA_TEMPERATURA - CILINDRO_SUPERIOR (°C)",
        "label": "Cilindro Superior",
        "excel_header": "TEMP. CILINDRO SUPERIOR (°C)",
        "group": "temperaturas_processo",
        "group_label": "Temperaturas do Processo",
        "unit": "°C",
        "data_type": 3,
        "order": 17,
        "is_numeric": True,
    },
    {
        "key": "temp_furador",
        "tag_name": "CALANDRA_TEMPERATURA - FURADOR (°C)",
        "label": "Furador",
        "excel_header": "TEMP. FURADOR (°C)",
        "group": "temperaturas_processo",
        "group_label": "Temperaturas do Processo",
        "unit": "°C",
        "data_type": 3,
        "order": 18,
        "is_numeric": True,
    },
    {
        "key": "temp_aquecedor",
        "tag_name": "CALANDRA_TEMPERATURA - AQUECEDOR",
        "label": "Aquecedor",
        "excel_header": "TEMP. AQUECEDOR (°C)",
        "group": "temperaturas_processo",
        "group_label": "Temperaturas do Processo",
        "unit": "°C",
        "data_type": 3,
        "order": 19,
        "is_numeric": True,
    },
    {
        "key": "temp_tcu_extrusora",
        "tag_name": "CALANDRA_TEMPERATURA - TCU_EXTRUSORA (°C)",
        "label": "TCU Extrusora",
        "excel_header": "TEMP. TCU EXTRUSORA (°C)",
        "group": "temperaturas_processo",
        "group_label": "Temperaturas do Processo",
        "unit": "°C",
        "data_type": 3,
        "order": 20,
        "is_numeric": True,
    },
]


class CalandraHistoricalService:
    """
    Serviço especializado para consulta histórica, sincronização temporal (forward-fill),
    visualização gráfica agrupada e exportação Excel das variáveis da CALANDRA.
    """

    MAX_QUERY_DAYS = 31  # Proteção contra períodos excessivos ao SCADA

    @classmethod
    def get_variables_config(cls) -> List[Dict[str, Any]]:
        """
        Retorna as 20 configurações de variáveis da Calandra.
        Verifica no banco 'default' (ProductionGlobalParameter) se o usuário customizou
        algum XID via Central de Configuração SCADA. Caso contrário, utiliza o padrão canônico.
        """
        from .models import ProductionGlobalParameter
        configs = [dict(c) for c in CALANDRA_VARIABLES_CONFIG]
        try:
            db_params = {
                p.chave: p.xid
                for p in ProductionGlobalParameter.objects.filter(chave__startswith="calandra_")
                if p.xid and p.xid.strip()
            }
            if db_params:
                for c in configs:
                    db_key = f"calandra_{c['key']}"
                    if db_key in db_params:
                        c["tag_name"] = db_params[db_key].strip()
        except Exception:
            pass
        return configs

    @classmethod
    def format_passada_label(cls, val: Any) -> str:
        """
        Formata o valor da variável PASSADA em rótulo contextual:
        - 1 / '1': PASSADA 1 — 1ª face
        - 2 / '2': PASSADA 2 — 2ª face / face oposta
        """
        if val is None or val == "":
            return "Não informada"
        s = str(val).strip()
        try:
            f = float(s)
            int_v = int(f)
            if int_v == 1:
                return "PASSADA 1 — 1ª face"
            elif int_v == 2:
                return "PASSADA 2 — 2ª face / face oposta"
            return f"PASSADA {int_v}"
        except (ValueError, TypeError):
            if s in ("1", "1.0"):
                return "PASSADA 1 — 1ª face"
            elif s in ("2", "2.0"):
                return "PASSADA 2 — 2ª face / face oposta"
            return s

    @classmethod
    def get_passada_window_context(cls, timeline: List[Dict[str, Any]]) -> str:
        """
        Identifica o contexto semântico da PASSADA presente na janela da timeline:
        - Se contiver apenas PASSADA 1: 'PASSADA 1 — 1ª face'
        - Se contiver apenas PASSADA 2: 'PASSADA 2 — 2ª face / face oposta'
        - Se contiver ambas: 'PASSADA 1 e PASSADA 2'
        - Caso contrário: 'Não informada' ou descrição adequada
        """
        if not timeline:
            return "Sem dados"

        passadas_encontradas = set()
        for item in timeline:
            p_val = item.get("passada_val")
            if p_val is not None and str(p_val).strip() != "":
                try:
                    f = float(p_val)
                    passadas_encontradas.add(int(f))
                except (ValueError, TypeError):
                    passadas_encontradas.add(str(p_val).strip())

        if 1 in passadas_encontradas and 2 in passadas_encontradas:
            return "PASSADA 1 e PASSADA 2"
        elif 1 in passadas_encontradas or "1" in passadas_encontradas:
            return "PASSADA 1 — 1ª face"
        elif 2 in passadas_encontradas or "2" in passadas_encontradas:
            return "PASSADA 2 — 2ª face / face oposta"
        elif len(passadas_encontradas) == 1:
            val = list(passadas_encontradas)[0]
            return cls.format_passada_label(val)
        return "Não informada"

    @classmethod
    def parse_period_filters(
        cls,
        periodo: Optional[str] = None,
        data_inicio: Optional[str] = None,
        hora_inicio: Optional[str] = None,
        data_final: Optional[str] = None,
        hora_final: Optional[str] = None,
    ) -> Tuple[datetime, datetime, str, Optional[str]]:
        """
        Interpreta filtros de período retornando (start_dt, end_dt, periodo_ativo, error_msg).
        Respeita o timezone do Django (America/Sao_Paulo).
        """
        now = timezone.localtime()
        today_date = now.date()
        error_msg = None

        if periodo == "ontem":
            yesterday = today_date - timedelta(days=1)
            start_dt = timezone.make_aware(datetime.combine(yesterday, dt_time(0, 0, 0)))
            end_dt = timezone.make_aware(datetime.combine(yesterday, dt_time(23, 59, 59, 999000)))
            return start_dt, end_dt, "ontem", None

        elif periodo == "7d":
            start_date = today_date - timedelta(days=7)
            start_dt = timezone.make_aware(datetime.combine(start_date, dt_time(0, 0, 0)))
            end_dt = now
            return start_dt, end_dt, "7d", None

        elif periodo == "30d":
            start_date = today_date - timedelta(days=30)
            start_dt = timezone.make_aware(datetime.combine(start_date, dt_time(0, 0, 0)))
            end_dt = now
            return start_dt, end_dt, "30d", None

        elif periodo == "personalizado" or (data_inicio and data_final):
            try:
                # Tratar data_inicio
                d_ini = datetime.strptime(data_inicio.strip(), "%Y-%m-%d").date()
                if hora_inicio and hora_inicio.strip():
                    h_ini = datetime.strptime(hora_inicio.strip(), "%H:%M").time()
                else:
                    h_ini = dt_time(0, 0, 0)
                start_dt = timezone.make_aware(datetime.combine(d_ini, h_ini))

                # Tratar data_final
                d_fim = datetime.strptime(data_final.strip(), "%Y-%m-%d").date()
                if hora_final and hora_final.strip():
                    h_fim = datetime.strptime(hora_final.strip(), "%H:%M").time()
                else:
                    h_fim = dt_time(23, 59, 59, 999000)
                end_dt = timezone.make_aware(datetime.combine(d_fim, h_fim))

                if start_dt > end_dt:
                    error_msg = "A data/hora inicial não pode ser maior que a data/hora final."
                    start_dt = timezone.make_aware(datetime.combine(today_date, dt_time(0, 0, 0)))
                    end_dt = now
                    return start_dt, end_dt, "hoje", error_msg

                # Proteção contra períodos excessivos (> 31 dias)
                if (end_dt - start_dt).days > cls.MAX_QUERY_DAYS:
                    error_msg = f"O intervalo máximo permitido para consulta é de {cls.MAX_QUERY_DAYS} dias."
                    start_dt = end_dt - timedelta(days=cls.MAX_QUERY_DAYS)

                return start_dt, end_dt, "personalizado", error_msg
            except Exception as e:
                error_msg = f"Parâmetros de data/hora inválidos: {e}"
                start_dt = timezone.make_aware(datetime.combine(today_date, dt_time(0, 0, 0)))
                end_dt = now
                return start_dt, end_dt, "hoje", error_msg

        # Padrão: Hoje
        start_dt = timezone.make_aware(datetime.combine(today_date, dt_time(0, 0, 0)))
        end_dt = now
        return start_dt, end_dt, "hoje", None

    @classmethod
    def resolve_calandra_datapoints(cls) -> Dict[str, Dict[str, Any]]:
        """
        Localiza os DataPoints do Scada-LTS para as 20 variáveis.
        Busca de forma flexível por `xid` ou `pointName` para garantir compatibilidade.
        Retorna dicionário {var_key: {'dp_id': int, 'xid': str, 'point_name': str, 'config': dict}}.
        """
        configs = cls.get_variables_config()
        tag_names = [c["tag_name"] for c in configs]

        resolved: Dict[str, Dict[str, Any]] = {}
        try:
            dps = list(
                ScadaDataPoint.objects.using("scada")
                .filter(Q(xid__in=tag_names) | Q(point_name__in=tag_names))
                .values("id", "xid", "point_name")
            )
            dp_by_xid = {dp["xid"]: dp for dp in dps}
            dp_by_name = {dp["point_name"]: dp for dp in dps if dp.get("point_name")}

            for c in configs:
                tag = c["tag_name"]
                dp_match = dp_by_xid.get(tag) or dp_by_name.get(tag)
                if dp_match:
                    resolved[c["key"]] = {
                        "dp_id": dp_match["id"],
                        "xid": dp_match["xid"],
                        "point_name": dp_match["point_name"],
                        "config": c,
                    }
                else:
                    # Registra entrada mesmo que datapoint não exista ainda no banco
                    resolved[c["key"]] = {
                        "dp_id": None,
                        "xid": tag,
                        "point_name": tag,
                        "config": c,
                    }
        except Exception:
            for c in configs:
                resolved[c["key"]] = {
                    "dp_id": None,
                    "xid": c["tag_name"],
                    "point_name": c["tag_name"],
                    "config": c,
                }

        return resolved

    @classmethod
    def detect_effective_process(cls, timeline: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Classifica cada ponto da timeline como processo efetivo (is_effective = True/False)
        baseado na combinação de velocidade da calandra e avanço real da metragem bobinada.
        
        Regras aplicadas:
        1. Velocidade: v > 0.05 m/min (elimina ruído de parada).
        2. Metragem: deve haver avanço positivo no bloco contínuo de produção.
        3. Resets de metragem (ex: troca de rolo/bobina com avanço posterior) são reconhecidos.
        4. Períodos estagnados, testes em vazio (v > 0 sem metragem avançar) ou paradas (v <= 0) são marcados como False.
        """
        if not timeline:
            return timeline

        n = len(timeline)
        speeds: List[float] = []
        meters: List[Optional[float]] = []

        for item in timeline:
            v_val = item["values"].get("vel_calandra")
            m_val = item["values"].get("metragem_bobinada")
            try:
                v_f = float(v_val) if v_val is not None and str(v_val).strip() != "" else 0.0
            except (ValueError, TypeError):
                v_f = 0.0
            try:
                m_f = float(m_val) if m_val is not None and str(m_val).strip() != "" else None
            except (ValueError, TypeError):
                m_f = None
            speeds.append(v_f)
            meters.append(m_f)

        is_effective_flags = [False] * n

        i = 0
        while i < n:
            if speeds[i] > 0.05:
                start_blk = i
                while i < n and speeds[i] > 0.05:
                    i += 1
                end_blk = i  # Bloco contíguo de velocidade positiva: [start_blk:end_blk]

                blk_meters = [meters[k] for k in range(start_blk, end_blk) if meters[k] is not None]

                if not blk_meters:
                    for k in range(start_blk, end_blk):
                        is_effective_flags[k] = False
                elif len(blk_meters) == 1:
                    prev_m = meters[start_blk - 1] if start_blk > 0 else None
                    curr_m = blk_meters[0]
                    if prev_m is not None and curr_m > prev_m:
                        is_effective_flags[start_blk] = True
                    elif curr_m > 0:
                        is_effective_flags[start_blk] = True
                    else:
                        is_effective_flags[start_blk] = False
                else:
                    # Múltiplos pontos no bloco. Verificar se há avanço positivo de metragem
                    has_advance = False
                    for k in range(start_blk, end_blk - 1):
                        m_curr = meters[k]
                        m_nxt = meters[k + 1]
                        if m_curr is not None and m_nxt is not None:
                            if m_nxt > m_curr:
                                has_advance = True
                                break
                            elif m_nxt < m_curr and m_nxt >= 0:
                                # Reset de bobina / nova metragem com produção contínua
                                has_advance = True
                                break

                    if not has_advance and start_blk > 0 and meters[start_blk - 1] is not None and blk_meters[0] > meters[start_blk - 1]:
                        has_advance = True

                    for k in range(start_blk, end_blk):
                        is_effective_flags[k] = has_advance
            else:
                is_effective_flags[i] = False
                i += 1

        for idx, item in enumerate(timeline):
            item["is_effective"] = is_effective_flags[idx]

        return timeline

    @classmethod
    def compute_effective_process_stats(cls, timeline: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Calcula as estatísticas descritivas (Média, Mínimo e Máximo) para os 4 cards principais
        de auditoria, considerando estritamente os pontos classificados como Processo Efetivo.
        """
        effective_rows = [item for item in timeline if item.get("is_effective") is True]
        has_effective = len(effective_rows) > 0

        def _calc_var_stat(var_key: str) -> Dict[str, Optional[float]]:
            if not has_effective:
                return {"avg": None, "min": None, "max": None, "count": 0}
            vals: List[float] = []
            for r in effective_rows:
                v = r["values"].get(var_key)
                if v is not None and str(v).strip() != "":
                    try:
                        vals.append(float(v))
                    except (ValueError, TypeError):
                        pass
            if not vals:
                return {"avg": None, "min": None, "max": None, "count": 0}
            return {
                "avg": round(sum(vals) / len(vals), 1),
                "min": round(min(vals), 1),
                "max": round(max(vals), 1),
                "count": len(vals),
            }

        # Card 1: Temperatura da Borracha (°C)
        temp_borracha = {
            "saida_extrusao": _calc_var_stat("temp_borracha_saida_extrusao"),
            "ent_calandra": _calc_var_stat("temp_borracha_ent_calandra"),
            "saida_calandra": _calc_var_stat("temp_borracha_saida_calandra"),
        }

        # Card 2: Temperatura dos Cilindros (°C)
        cil_inf = _calc_var_stat("temp_cilindro_inf")
        cil_inter = _calc_var_stat("temp_cilindro_inter")
        cil_sup = _calc_var_stat("temp_cilindro_sup")
        
        delta_cilindros = None
        cil_avgs = [c["avg"] for c in (cil_inf, cil_inter, cil_sup) if c["avg"] is not None]
        if len(cil_avgs) >= 2:
            delta_cilindros = round(max(cil_avgs) - min(cil_avgs), 1)

        temp_cilindros = {
            "cilindro_inf": cil_inf,
            "cilindro_inter": cil_inter,
            "cilindro_sup": cil_sup,
            "delta_cilindros": delta_cilindros,
        }

        # Card 3: Cargas do Processo (kg)
        cargas = {
            "desbobinador": _calc_var_stat("carga_desbobinador"),
            "quebra_trama": _calc_var_stat("carga_quebra_trama"),
            "pos_calandra": _calc_var_stat("carga_pos_calandra"),
            "bobinamento": _calc_var_stat("carga_bobinamento"),
        }

        # Card 4: Temperaturas Auxiliares (°C)
        temp_auxiliares = {
            "furador": _calc_var_stat("temp_furador"),
            "aquecedor": _calc_var_stat("temp_aquecedor"),
            "tcu_extrusora": _calc_var_stat("temp_tcu_extrusora"),
        }

        return {
            "has_effective_process": has_effective,
            "effective_points_count": len(effective_rows),
            "total_points_count": len(timeline),
            "temp_borracha": temp_borracha,
            "temp_cilindros": temp_cilindros,
            "cargas": cargas,
            "temp_auxiliares": temp_auxiliares,
        }

    @classmethod
    def get_synchronized_history(
        cls,
        start_dt: datetime,
        end_dt: datetime,
    ) -> Dict[str, Any]:
        """
        Executa a consulta histórica indexada e executa o algoritmo de forward-fill temporal.
        Aplica a classificação de Processo Efetivo e calcula as estatísticas para auditoria.
        """
        start_ms = int(start_dt.timestamp() * 1000)
        end_ms = int(end_dt.timestamp() * 1000)

        resolved_dps = cls.resolve_calandra_datapoints()
        dp_id_to_key: Dict[int, str] = {}
        active_dp_ids: List[int] = []
        variables_missing: List[str] = []

        for key, info in resolved_dps.items():
            if info["dp_id"]:
                dp_id_to_key[info["dp_id"]] = key
                active_dp_ids.append(info["dp_id"])
            else:
                variables_missing.append(info["config"]["label"])

        if not active_dp_ids:
            empty_charts = cls._empty_chart_datasets()
            return {
                "timeline": [],
                "raw_points_count": 0,
                "variables_found_count": 0,
                "variables_missing": variables_missing,
                "chart_datasets": empty_charts,
                "card_stats": cls.compute_effective_process_stats([]),
                "effective_points_count": 0,
                "passada_context": "Sem dados",
                "start_dt": start_dt,
                "end_dt": end_dt,
            }

        # 1. Obter estado inicial anterior ao start_ms para cada DP (seed do forward-fill)
        initial_state: Dict[str, Any] = {}
        for key in resolved_dps:
            initial_state[key] = None

        try:
            prior_max_ts_records = (
                ScadaPointValue.objects.using("scada")
                .filter(data_point_id__in=active_dp_ids, ts__lt=start_ms)
                .values("data_point_id")
                .annotate(max_ts=Max("ts"))
            )
            prior_map = {r["data_point_id"]: r["max_ts"] for r in prior_max_ts_records if r.get("max_ts")}
            if prior_map:
                q_prior = Q()
                for dp_id, max_ts in prior_map.items():
                    q_prior |= Q(data_point_id=dp_id, ts=max_ts)

                prior_values = (
                    ScadaPointValue.objects.using("scada")
                    .filter(q_prior)
                    .select_related("annotation")
                )
                for pv in prior_values:
                    v_key = dp_id_to_key.get(pv.data_point_id)
                    if v_key:
                        norm_val, _ = scada_reader.normalize_value(pv.data_type, pv.point_value, getattr(pv, "annotation", None))
                        initial_state[v_key] = norm_val
        except Exception:
            pass

        # 2. Consultar todos os registros no intervalo [start_ms, end_ms] indexados por ts
        try:
            point_values_qs = (
                ScadaPointValue.objects.using("scada")
                .filter(data_point_id__in=active_dp_ids, ts__gte=start_ms, ts__lte=end_ms)
                .select_related("annotation")
                .order_by("ts", "id")
            )
            raw_records = list(point_values_qs)
        except Exception:
            raw_records = []

        raw_points_count = len(raw_records)

        # 3. Algoritmo de Forward-Fill Temporal
        events_by_ts: Dict[int, Dict[str, Any]] = {}
        for pv in raw_records:
            v_key = dp_id_to_key.get(pv.data_point_id)
            if not v_key:
                continue
            norm_val, _ = scada_reader.normalize_value(pv.data_type, pv.point_value, getattr(pv, "annotation", None))
            if pv.ts not in events_by_ts:
                events_by_ts[pv.ts] = {}
            events_by_ts[pv.ts][v_key] = norm_val

        current_state = initial_state.copy()
        timeline: List[Dict[str, Any]] = []

        # Se houver estado inicial mas nenhum evento no intervalo, emitir ponto no início da janela
        if not events_by_ts and any(v is not None for v in current_state.values()):
            row_dt = timezone.localtime(timezone.datetime.fromtimestamp(start_ms / 1000, tz=timezone.utc))
            passada_val = current_state.get("passada")
            timeline.append({
                "ts": start_ms,
                "datetime": row_dt,
                "datetime_str": row_dt.strftime("%d/%m/%Y %H:%M:%S"),
                "passada_val": passada_val,
                "passada_label": cls.format_passada_label(passada_val),
                "values": current_state.copy(),
            })

        for ts in sorted(events_by_ts.keys()):
            updates = events_by_ts[ts]
            current_state.update(updates)
            row_dt = timezone.localtime(timezone.datetime.fromtimestamp(ts / 1000, tz=timezone.utc))
            passada_val = current_state.get("passada")

            timeline.append({
                "ts": ts,
                "datetime": row_dt,
                "datetime_str": row_dt.strftime("%d/%m/%Y %H:%M:%S"),
                "passada_val": passada_val,
                "passada_label": cls.format_passada_label(passada_val),
                "values": current_state.copy(),
            })

        # 4. Classificação de Processo Efetivo
        cls.detect_effective_process(timeline)

        # 5. Cálculo das estatísticas de auditoria dos 4 cards
        card_stats = cls.compute_effective_process_stats(timeline)
        passada_context = cls.get_passada_window_context(timeline)

        # 6. Preparar datasets para os 6 gráficos
        chart_datasets = cls._build_chart_datasets(timeline)

        return {
            "timeline": timeline,
            "raw_points_count": raw_points_count,
            "variables_found_count": len(active_dp_ids),
            "variables_missing": variables_missing,
            "chart_datasets": chart_datasets,
            "card_stats": card_stats,
            "effective_points_count": card_stats["effective_points_count"],
            "passada_context": passada_context,
            "start_dt": start_dt,
            "end_dt": end_dt,
        }

    @classmethod
    def _empty_chart_datasets(cls) -> Dict[str, Any]:
        return {
            "chart_1_producao": {"labels": [], "timestamps": [], "velocidade": [], "metragem": [], "passada": [], "passada_labels": [], "is_effective": []},
            "chart_2_cargas": {"labels": [], "timestamps": [], "desbobinador": [], "quebra_trama": [], "pos_calandra": [], "bobinamento": []},
            "chart_3_espessuras": {"labels": [], "timestamps": [], "esq_sup": [], "dir_sup": [], "dir_inf": [], "esq_inf": []},
            "chart_4_temp_borracha": {"labels": [], "timestamps": [], "saida_extrusao": [], "ent_calandra": [], "saida_calandra": []},
            "chart_5_temp_cilindros": {"labels": [], "timestamps": [], "cilindro_inf": [], "cilindro_inter": [], "cilindro_sup": []},
            "chart_6_temp_auxiliares": {"labels": [], "timestamps": [], "furador": [], "aquecedor": [], "tcu_extrusora": []},
            # Aliases legados para retrocompatibilidade
            "chart_a_producao": {"labels": [], "velocidade": [], "metragem": [], "passada": []},
            "chart_b_cargas": {"labels": [], "bobinamento": [], "desbobinador": [], "pos_calandra": [], "quebra_trama": []},
            "chart_c_espessuras": {"labels": [], "esq_sup": [], "dir_sup": [], "dir_inf": [], "esq_inf": []},
            "chart_d_temp_borracha": {"labels": [], "saida_extrusao": [], "ent_calandra": [], "saida_calandra": []},
            "chart_e_temp_processo": {"labels": [], "cilindro_inf": [], "cilindro_inter": [], "cilindro_sup": [], "furador": [], "aquecedor": [], "tcu_extrusora": []},
        }

    @classmethod
    def _build_chart_datasets(cls, timeline: List[Dict[str, Any]], max_chart_points: int = 1500) -> Dict[str, Any]:
        """
        Monta as séries de dados temporais para os 6 gráficos no Chart.js.
        Preserva timestamps e flags de processo efetivo para seleção coordenada no cliente.
        """
        if not timeline:
            return cls._empty_chart_datasets()

        total = len(timeline)
        sampled_timeline: List[Dict[str, Any]] = []

        if total <= max_chart_points:
            sampled_timeline = timeline
        else:
            step = total / max_chart_points
            for i in range(max_chart_points):
                idx = min(int(i * step), total - 1)
                sampled_timeline.append(timeline[idx])

            # Garantir inclusão de transições de passada
            for idx in range(1, total):
                if timeline[idx].get("passada_val") != timeline[idx - 1].get("passada_val"):
                    if timeline[idx] not in sampled_timeline:
                        sampled_timeline.append(timeline[idx])

            sampled_timeline.sort(key=lambda x: x["ts"])

        # Vetores comuns
        labels = [item["datetime_str"] for item in sampled_timeline]
        timestamps = [item["ts"] for item in sampled_timeline]
        passada_vals = [item.get("passada_val") for item in sampled_timeline]
        passada_lbls = [item.get("passada_label", cls.format_passada_label(p)) for item, p in zip(sampled_timeline, passada_vals)]
        is_effective_list = [item.get("is_effective", False) for item in sampled_timeline]

        def _get_float(item: Dict[str, Any], key: str) -> Optional[float]:
            v = item["values"].get(key)
            if v is None or str(v).strip() == "":
                return None
            try:
                return round(float(v), 2)
            except (ValueError, TypeError):
                return None

        # Séries individuais
        vel_list = [_get_float(item, "vel_calandra") for item in sampled_timeline]
        metr_list = [_get_float(item, "metragem_bobinada") for item in sampled_timeline]

        bob_list = [_get_float(item, "carga_bobinamento") for item in sampled_timeline]
        desb_list = [_get_float(item, "carga_desbobinador") for item in sampled_timeline]
        pos_list = [_get_float(item, "carga_pos_calandra") for item in sampled_timeline]
        queb_list = [_get_float(item, "carga_quebra_trama") for item in sampled_timeline]

        esq_sup = [_get_float(item, "espessura_esq_sup") for item in sampled_timeline]
        dir_sup = [_get_float(item, "espessura_dir_sup") for item in sampled_timeline]
        dir_inf = [_get_float(item, "espessura_dir_inf") for item in sampled_timeline]
        esq_inf = [_get_float(item, "espessura_esq_inf") for item in sampled_timeline]

        t_saida_extr = [_get_float(item, "temp_borracha_saida_extrusao") for item in sampled_timeline]
        t_ent_cal = [_get_float(item, "temp_borracha_ent_calandra") for item in sampled_timeline]
        t_saida_cal = [_get_float(item, "temp_borracha_saida_calandra") for item in sampled_timeline]

        t_cil_inf = [_get_float(item, "temp_cilindro_inf") for item in sampled_timeline]
        t_cil_inter = [_get_float(item, "temp_cilindro_inter") for item in sampled_timeline]
        t_cil_sup = [_get_float(item, "temp_cilindro_sup") for item in sampled_timeline]

        t_fur = [_get_float(item, "temp_furador") for item in sampled_timeline]
        t_aquec = [_get_float(item, "temp_aquecedor") for item in sampled_timeline]
        t_tcu = [_get_float(item, "temp_tcu_extrusora") for item in sampled_timeline]

        return {
            # Gráfico 1: Produção (Velocidade + Metragem + Passada)
            "chart_1_producao": {
                "labels": labels,
                "timestamps": timestamps,
                "velocidade": vel_list,
                "metragem": metr_list,
                "passada": passada_vals,
                "passada_labels": passada_lbls,
                "is_effective": is_effective_list,
            },
            # Gráfico 2: Cargas do Processo (kg)
            "chart_2_cargas": {
                "labels": labels,
                "timestamps": timestamps,
                "desbobinador": desb_list,
                "quebra_trama": queb_list,
                "pos_calandra": pos_list,
                "bobinamento": bob_list,
            },
            # Gráfico 3: Espessuras (mm)
            "chart_3_espessuras": {
                "labels": labels,
                "timestamps": timestamps,
                "esq_sup": esq_sup,
                "dir_sup": dir_sup,
                "dir_inf": dir_inf,
                "esq_inf": esq_inf,
            },
            # Gráfico 4: Temperatura da Borracha (°C)
            "chart_4_temp_borracha": {
                "labels": labels,
                "timestamps": timestamps,
                "saida_extrusao": t_saida_extr,
                "ent_calandra": t_ent_cal,
                "saida_calandra": t_saida_cal,
            },
            # Gráfico 5: Temperatura dos Cilindros (°C)
            "chart_5_temp_cilindros": {
                "labels": labels,
                "timestamps": timestamps,
                "cilindro_inf": t_cil_inf,
                "cilindro_inter": t_cil_inter,
                "cilindro_sup": t_cil_sup,
            },
            # Gráfico 6: Temperaturas Auxiliares (°C)
            "chart_6_temp_auxiliares": {
                "labels": labels,
                "timestamps": timestamps,
                "furador": t_fur,
                "aquecedor": t_aquec,
                "tcu_extrusora": t_tcu,
            },
            # Aliases legados (retrocompatibilidade)
            "chart_a_producao": {
                "labels": labels,
                "velocidade": vel_list,
                "metragem": metr_list,
                "passada": passada_vals,
                "passada_labels": passada_lbls,
            },
            "chart_b_cargas": {
                "labels": labels,
                "bobinamento": bob_list,
                "desbobinador": desb_list,
                "pos_calandra": pos_list,
                "quebra_trama": queb_list,
            },
            "chart_c_espessuras": {
                "labels": labels,
                "esq_sup": esq_sup,
                "dir_sup": dir_sup,
                "dir_inf": dir_inf,
                "esq_inf": esq_inf,
            },
            "chart_d_temp_borracha": {
                "labels": labels,
                "saida_extrusao": t_saida_extr,
                "ent_calandra": t_ent_cal,
                "saida_calandra": t_saida_cal,
            },
            "chart_e_temp_processo": {
                "labels": labels,
                "cilindro_inf": t_cil_inf,
                "cilindro_inter": t_cil_inter,
                "cilindro_sup": t_cil_sup,
                "furador": t_fur,
                "aquecedor": t_aquec,
                "tcu_extrusora": t_tcu,
            },
        }

    @classmethod
    def generate_excel_report(
        cls,
        start_dt: datetime,
        end_dt: datetime,
    ) -> bytes:
        """
        Gera arquivo Excel (.xlsx) contendo 100% dos dados históricos sincronizados brutos.
        Primeira coluna: DATA/HORA
        Segunda coluna: PASSADA
        Demais colunas: Variáveis da Calandra
        """
        history = cls.get_synchronized_history(start_dt, end_dt)
        timeline = history["timeline"]
        configs = cls.get_variables_config()

        # Variáveis excluindo 'passada' (que fica na coluna 2)
        other_vars = [c for c in configs if c["key"] != "passada"]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Histórico Calandra"
        ws.views.sheetView[0].showGridLines = True

        # Estilos
        font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        fill_header_gold = PatternFill(start_color="B8842E", end_color="B8842E", fill_type="solid")
        border_thin = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )
        font_data = Font(name="Segoe UI", size=10, color="0F172A")
        font_passada1 = Font(name="Segoe UI", size=10, bold=True, color="0369A1")
        font_passada2 = Font(name="Segoe UI", size=10, bold=True, color="B45309")
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        # Cabeçalhos
        headers = ["DATA/HORA", "PASSADA"] + [c["excel_header"] for c in other_vars]
        ws.append(headers)

        # Formatar Linha 1 (Cabeçalho)
        ws.row_dimensions[1].height = 28
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header_gold if col_num in (1, 2) else fill_header
            cell.alignment = align_center
            cell.border = border_thin

        # Preencher Dados
        for row_idx, item in enumerate(timeline, start=2):
            ws.row_dimensions[row_idx].height = 20
            row_dt = item["datetime"]
            passada_lbl = item["passada_label"]

            # Coluna 1: DATA/HORA (como datetime do Excel)
            c1 = ws.cell(row=row_idx, column=1, value=row_dt.strftime("%d/%m/%Y %H:%M:%S"))
            c1.font = font_data
            c1.alignment = align_center
            c1.border = border_thin

            # Coluna 2: PASSADA
            c2 = ws.cell(row=row_idx, column=2, value=passada_lbl)
            c2.font = font_passada1 if "PASSADA 1" in passada_lbl else (font_passada2 if "PASSADA 2" in passada_lbl else font_data)
            c2.alignment = align_center
            c2.border = border_thin

            # Demais colunas
            for c_idx, c_cfg in enumerate(other_vars, start=3):
                raw_v = item["values"].get(c_cfg["key"])
                cell = ws.cell(row=row_idx, column=c_idx)
                cell.border = border_thin
                cell.font = font_data

                if raw_v is None or raw_v == "":
                    cell.value = "-"
                    cell.alignment = align_center
                elif c_cfg["is_numeric"]:
                    try:
                        num_val = round(float(raw_v), 2)
                        cell.value = num_val
                        cell.alignment = align_right
                        cell.number_format = "#,##0.00" if isinstance(num_val, float) and not num_val.is_integer() else "#,##0"
                    except (ValueError, TypeError):
                        cell.value = str(raw_v)
                        cell.alignment = align_left
                else:
                    cell.value = str(raw_v)
                    cell.alignment = align_left

        # Configurações de Usabilidade
        ws.freeze_panes = "A2"
        if timeline:
            ws.auto_filter.ref = ws.dimensions

        # Auto-ajuste de largura de colunas
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
