import io
import os
import time
import math
from datetime import datetime, date, time as dt_time, timedelta
from typing import Dict, List, Any, Optional, Tuple
from django.utils import timezone
from django.db.models import Max, Q
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.chart.axis import DateAxis

from .models import ScadaDataPoint, ScadaPointValue, ScadaPointValueAnnotation
from .services import scada_reader


# ==============================================================================
# CANONICAL INVENTORY OF THE 9 CALDEIRA 2 / UTILITIES VARIABLES
# ==============================================================================

CALDEIRA_VARIABLES_CONFIG: List[Dict[str, Any]] = [
    # ── 1. Geração de Vapor ──
    {
        "key": "pressao_caldeira",
        "tag_name": "CALDEIRA 2 - PRESSCAL",
        "label": "Pressão da Caldeira 2",
        "friendly_name": "Pressão da Caldeira 2",
        "description": "Pressão de vapor medida na própria caldeira.",
        "excel_header": "PRESSÃO CALDEIRA 2 (bar)",
        "group": "geracao_vapor",
        "group_label": "Geração de Vapor",
        "unit": "bar",
        "data_type": 3,
        "order": 1,
        "nature": "Variável instantânea",
        "is_numeric": True,
    },
    {
        "key": "setpoint_pressao_alta",
        "tag_name": "VALVULA_VAPOR - setPress",
        "label": "Pressão de alta configurada",
        "friendly_name": "Pressão de alta configurada (Setpoint)",
        "description": "Setpoint configurado para manter as linhas de pressão de alta após a válvula de controle de vapor.",
        "excel_header": "SETPOINT PRESSÃO ALTA (bar)",
        "group": "geracao_vapor",
        "group_label": "Geração de Vapor",
        "unit": "bar",
        "data_type": 3,
        "order": 2,
        "nature": "Variável instantânea",
        "is_numeric": True,
    },

    # ── 2. Linhas de Alta ──
    {
        "key": "pressao_alta_prensas_1_7",
        "tag_name": "VALVULA_VAPOR - Press1_PRENSAS 1 A 7",
        "label": "Pressão alta — prensas 1 a 7",
        "friendly_name": "Pressão alta — prensas 1 a 7",
        "description": "Pressão de alta da linha de vapor que atende as prensas 1 a 7 (máx. 14 cavidades).",
        "excel_header": "PRESSÃO ALTA PRENSAS 1-7 (bar)",
        "group": "linhas_alta",
        "group_label": "Linhas de Alta",
        "unit": "bar",
        "data_type": 3,
        "order": 3,
        "nature": "Variável instantânea",
        "is_numeric": True,
    },
    {
        "key": "pressao_alta_prensas_8_12",
        "tag_name": "VALVULA_VAPOR - Press2_PRENSAS 8 A 12",
        "label": "Pressão alta — prensas 8 a 12",
        "friendly_name": "Pressão alta — prensas 8 a 12",
        "description": "Pressão de alta da linha de vapor que atende as prensas 8 a 12 (máx. 18 cavidades).",
        "excel_header": "PRESSÃO ALTA PRENSAS 8-12 (bar)",
        "group": "linhas_alta",
        "group_label": "Linhas de Alta",
        "unit": "bar",
        "data_type": 3,
        "order": 4,
        "nature": "Variável instantânea",
        "is_numeric": True,
    },

    # ── 3. Linhas de Baixa ──
    {
        "key": "pressao_baixa_prensas_1_7",
        "tag_name": "VALVULA_VAPOR - PressBX1_PRENSAS 1 A 7",
        "label": "Pressão baixa — prensas 1 a 7",
        "friendly_name": "Pressão baixa — prensas 1 a 7",
        "description": "Pressão de baixa da linha de vapor que atende as prensas 1 a 7.",
        "excel_header": "PRESSÃO BAIXA PRENSAS 1-7 (bar)",
        "group": "linhas_baixa",
        "group_label": "Linhas de Baixa",
        "unit": "bar",
        "data_type": 3,
        "order": 5,
        "nature": "Variável instantânea",
        "is_numeric": True,
    },
    {
        "key": "pressao_baixa_prensas_8_12",
        "tag_name": "VALVULA_VAPOR - PressBX2_PRENSAS 8 A 12",
        "label": "Pressão baixa — prensas 8 a 12",
        "friendly_name": "Pressão baixa — prensas 8 a 12",
        "description": "Pressão de baixa da linha de vapor que atende as prensas 8 a 12.",
        "excel_header": "PRESSÃO BAIXA PRENSAS 8-12 (bar)",
        "group": "linhas_baixa",
        "group_label": "Linhas de Baixa",
        "unit": "bar",
        "data_type": 3,
        "order": 6,
        "nature": "Variável instantânea",
        "is_numeric": True,
    },

    # ── 4. Condensado ──
    {
        "key": "volume_condensado",
        "tag_name": "Meta_Calculos_Prensas - VOLUME_CAL - Volume Condensado",
        "label": "Volume acumulado de condensado",
        "friendly_name": "Volume acumulado de condensado",
        "description": "Totalizador acumulado do volume de condensado medido na linha (consumido diretamente em litros).",
        "excel_header": "TOTALIZADOR CONDENSADO (L)",
        "group": "condensado",
        "group_label": "Condensado",
        "unit": "L",
        "data_type": 3,
        "order": 7,
        "nature": "Totalizador acumulativo",
        "is_numeric": True,
    },

    # ── 5. Utilidades Auxiliares ──
    {
        "key": "pressao_ar_comprimido",
        "tag_name": "VALVULA_VAPOR - AR_COMPRIMIDO_VULC",
        "label": "Pressão do ar comprimido da vulcanização",
        "friendly_name": "Pressão do ar comprimido da vulcanização",
        "description": "Pressão pneumática utilizada na conformação e descarga das prensas.",
        "excel_header": "AR COMPRIMIDO VULCANIZAÇÃO (bar)",
        "group": "utilidades_auxiliares",
        "group_label": "Utilidades Auxiliares",
        "unit": "bar",
        "data_type": 3,
        "order": 8,
        "nature": "Variável instantânea",
        "is_numeric": True,
    },
    {
        "key": "pressao_vacuo",
        "tag_name": "UNIDADE DE VÁCUO - PRESS_VACUO",
        "label": "Pressão da unidade de vácuo",
        "friendly_name": "Pressão da unidade de vácuo",
        "description": "Pressão da unidade de vácuo para drenagem do bladder e liberação de abertura das prensas.",
        "excel_header": "UNIDADE DE VÁCUO (bar)",
        "group": "utilidades_auxiliares",
        "group_label": "Utilidades Auxiliares",
        "unit": "bar",
        "data_type": 3,
        "order": 9,
        "nature": "Variável instantânea",
        "is_numeric": True,
    },
]


class CaldeiraHistoricalService:
    """
    Serviço analítico especializado para consulta histórica, sincronização temporal,
    estatísticas descritivas de vapor e utilidades, cálculo de condensado e exportação Excel
    da CALDEIRA 2.
    """

    MAX_QUERY_DAYS = 31

    @classmethod
    def get_variables_config(cls) -> List[Dict[str, Any]]:
        """
        Retorna as configurações das 9 variáveis da Caldeira.
        Verifica no banco 'default' (ProductionGlobalParameter) se o usuário customizou
        algum XID via Central de Configuração SCADA (chave com prefixo 'caldeira_').
        """
        from .models import ProductionGlobalParameter
        configs = [dict(c) for c in CALDEIRA_VARIABLES_CONFIG]
        try:
            db_params = {
                p.chave: p.xid
                for p in ProductionGlobalParameter.objects.filter(chave__startswith="caldeira_")
                if p.xid and p.xid.strip()
            }
            if db_params:
                for c in configs:
                    db_key = f"caldeira_{c['key']}"
                    if db_key in db_params:
                        c["tag_name"] = db_params[db_key].strip()
        except Exception:
            pass
        return configs

    @classmethod
    def parse_period_filters(
        cls,
        periodo: Optional[str] = None,
        data_inicio: Optional[str] = None,
        hora_inicio: Optional[str] = None,
        data_final: Optional[str] = None,
        hora_final: Optional[str] = None,
    ) -> Tuple[datetime, datetime, str, Optional[str]]:
        """
        Interpreta filtros de período retornando (start_dt, end_dt, periodo_ativo, error_msg).
        Respeita o timezone do Django (America/Sao_Paulo).
        """
        now = timezone.localtime()
        today_date = now.date()
        error_msg = None

        if periodo == "ontem":
            yesterday = today_date - timedelta(days=1)
            start_dt = timezone.make_aware(datetime.combine(yesterday, dt_time(0, 0, 0)))
            end_dt = timezone.make_aware(datetime.combine(yesterday, dt_time(23, 59, 59, 999000)))
            return start_dt, end_dt, "ontem", None

        elif periodo == "7d":
            start_date = today_date - timedelta(days=7)
            start_dt = timezone.make_aware(datetime.combine(start_date, dt_time(0, 0, 0)))
            end_dt = now
            return start_dt, end_dt, "7d", None

        elif periodo == "30d":
            start_date = today_date - timedelta(days=30)
            start_dt = timezone.make_aware(datetime.combine(start_date, dt_time(0, 0, 0)))
            end_dt = now
            return start_dt, end_dt, "30d", None

        elif periodo == "personalizado" or (data_inicio and data_final):
            try:
                d_ini = datetime.strptime(data_inicio.strip(), "%Y-%m-%d").date()
                if hora_inicio and hora_inicio.strip():
                    h_ini = datetime.strptime(hora_inicio.strip(), "%H:%M").time()
                else:
                    h_ini = dt_time(0, 0, 0)
                start_dt = timezone.make_aware(datetime.combine(d_ini, h_ini))

                d_fim = datetime.strptime(data_final.strip(), "%Y-%m-%d").date()
                if hora_final and hora_final.strip():
                    h_fim = datetime.strptime(hora_final.strip(), "%H:%M").time()
                else:
                    h_fim = dt_time(23, 59, 59, 999000)
                end_dt = timezone.make_aware(datetime.combine(d_fim, h_fim))

                if start_dt > end_dt:
                    error_msg = "A data/hora inicial não pode ser maior que a data/hora final."
                    start_dt = timezone.make_aware(datetime.combine(today_date, dt_time(0, 0, 0)))
                    end_dt = now
                    return start_dt, end_dt, "hoje", error_msg

                if (end_dt - start_dt).days > cls.MAX_QUERY_DAYS:
                    error_msg = f"O intervalo máximo permitido para consulta é de {cls.MAX_QUERY_DAYS} dias."
                    start_dt = end_dt - timedelta(days=cls.MAX_QUERY_DAYS)

                return start_dt, end_dt, "personalizado", error_msg
            except Exception as e:
                error_msg = f"Parâmetros de data/hora inválidos: {e}"
                start_dt = timezone.make_aware(datetime.combine(today_date, dt_time(0, 0, 0)))
                end_dt = now
                return start_dt, end_dt, "hoje", error_msg

        # Padrão: Hoje
        start_dt = timezone.make_aware(datetime.combine(today_date, dt_time(0, 0, 0)))
        end_dt = now
        return start_dt, end_dt, "hoje", None

    @classmethod
    def resolve_caldeira_datapoints(cls) -> Dict[str, Dict[str, Any]]:
        """
        Localiza os DataPoints do Scada-LTS para as 9 variáveis da Caldeira.
        Busca de forma flexível por `xid` ou `pointName`.
        """
        configs = cls.get_variables_config()
        tag_names = [c["tag_name"] for c in configs]

        resolved: Dict[str, Dict[str, Any]] = {}
        try:
            dps = list(
                ScadaDataPoint.objects.using("scada")
                .filter(Q(xid__in=tag_names) | Q(point_name__in=tag_names))
                .values("id", "xid", "point_name")
            )
            dp_by_xid = {dp["xid"]: dp for dp in dps}
            dp_by_name = {dp["point_name"]: dp for dp in dps if dp.get("point_name")}

            for c in configs:
                tag = c["tag_name"]
                dp_match = dp_by_xid.get(tag) or dp_by_name.get(tag)
                if dp_match:
                    resolved[c["key"]] = {
                        "dp_id": dp_match["id"],
                        "xid": dp_match["xid"],
                        "point_name": dp_match["point_name"],
                        "config": c,
                    }
                else:
                    resolved[c["key"]] = {
                        "dp_id": None,
                        "xid": tag,
                        "point_name": tag,
                        "config": c,
                    }
        except Exception:
            for c in configs:
                resolved[c["key"]] = {
                    "dp_id": None,
                    "xid": c["tag_name"],
                    "point_name": c["tag_name"],
                    "config": c,
                }

        return resolved

    @classmethod
    def compute_condensate_volume(
        cls,
        condensate_events: List[Dict[str, Any]],
        start_ms: int,
        end_ms: int,
        seed_reading: Optional[float] = None
    ) -> Dict[str, Any]:
        """
        Calcula o volume total de condensado no período a partir do totalizador acumulado VOLUME_CAL.
        Algoritmo:
        - Ordena as leituras válidas cronologicamente.
        - Para cada par consecutivo: delta = valor_atual - valor_anterior.
        - Se delta >= 0: adiciona ao volume do período.
        - Se delta < 0: registra reset/reinicialização do contador e inicia novo segmento sem subtrair.
        - Calcula a média horária (L/h) dividindo pelo intervalo coberto.
        """
        if not condensate_events:
            return {
                "has_data": False,
                "total_liters": 0.0,
                "formatted_liters": "Sem dados",
                "avg_liters_per_hour": None,
                "formatted_avg_lh": "Sem dados",
                "first_cumulative": None,
                "last_cumulative": None,
                "resets_count": 0,
                "samples_count": 0,
                "hours_covered": 0.0,
                "status_message": "Sem dados de condensado no período.",
                "hourly_series": [],
            }

        # Filtrar e ordenar por timestamp
        valid_events = []
        for ev in sorted(condensate_events, key=lambda x: x["ts"]):
            val = ev.get("val")
            if val is not None:
                try:
                    f_val = float(val)
                    if not math.isnan(f_val) and not math.isinf(f_val) and f_val >= 0:
                        valid_events.append({"ts": ev["ts"], "val": f_val})
                except (ValueError, TypeError):
                    pass

        if not valid_events:
            return {
                "has_data": False,
                "total_liters": 0.0,
                "formatted_liters": "Sem dados",
                "avg_liters_per_hour": None,
                "formatted_avg_lh": "Sem dados",
                "first_cumulative": None,
                "last_cumulative": None,
                "resets_count": 0,
                "samples_count": 0,
                "hours_covered": 0.0,
                "status_message": "Nenhuma leitura numérica válida de condensado.",
                "hourly_series": [],
            }

        if len(valid_events) == 1 and seed_reading is None:
            first_v = valid_events[0]["val"]
            return {
                "has_data": True,
                "total_liters": 0.0,
                "formatted_liters": "Dados insuficientes",
                "avg_liters_per_hour": None,
                "formatted_avg_lh": "Dados insuficientes",
                "first_cumulative": first_v,
                "last_cumulative": first_v,
                "resets_count": 0,
                "samples_count": 1,
                "hours_covered": 0.0,
                "status_message": "Apenas 1 leitura disponível. Dados insuficientes para calcular consumo por diferença.",
                "hourly_series": [],
            }

        total_volume = 0.0
        resets_count = 0
        deltas_list = []

        # Se houver seed reading antes do período, incluir no cálculo inicial
        all_readings = []
        if seed_reading is not None and not math.isnan(seed_reading) and seed_reading >= 0:
            all_readings.append({"ts": start_ms, "val": float(seed_reading)})
        all_readings.extend(valid_events)

        for i in range(1, len(all_readings)):
            prev_val = all_readings[i - 1]["val"]
            curr_val = all_readings[i]["val"]
            delta = curr_val - prev_val

            if delta >= 0:
                total_volume += delta
                deltas_list.append({
                    "ts": all_readings[i]["ts"],
                    "delta": delta,
                    "val": curr_val,
                })
            else:
                # Reset / Rollover detectado
                resets_count += 1
                deltas_list.append({
                    "ts": all_readings[i]["ts"],
                    "delta": 0.0,
                    "val": curr_val,
                    "is_reset": True,
                })

        first_cum = valid_events[0]["val"]
        last_cum = valid_events[-1]["val"]

        # Calcular tempo coberto
        first_ts = valid_events[0]["ts"]
        last_ts = valid_events[-1]["ts"]
        diff_ms = max(0, last_ts - first_ts)
        hours_covered = diff_ms / (1000.0 * 3600.0)

        avg_lh = None
        if hours_covered >= (1.0 / 60.0):  # Mínimo 1 minuto coberto para calcular taxa
            avg_lh = total_volume / hours_covered
        elif hours_covered > 0 and total_volume > 0:
            avg_lh = total_volume / hours_covered

        # Agrupar volume por hora para o Gráfico 4
        hourly_map: Dict[str, float] = {}
        for d in deltas_list:
            dt_loc = timezone.localtime(timezone.datetime.fromtimestamp(d["ts"] / 1000, tz=timezone.utc))
            hour_key = dt_loc.strftime("%d/%m %H:00")
            hourly_map[hour_key] = hourly_map.get(hour_key, 0.0) + d["delta"]

        hourly_series = [
            {"hour": h_label, "volume": round(vol, 1)}
            for h_label, vol in sorted(hourly_map.items())
        ]

        formatted_liters = f"{total_volume:,.1f} L".replace(",", "X").replace(".", ",").replace("X", ".")
        formatted_avg_lh = f"{avg_lh:,.1f} L/h".replace(",", "X").replace(".", ",").replace("X", ".") if avg_lh is not None else "Dados insuficientes"

        return {
            "has_data": True,
            "total_liters": round(total_volume, 2),
            "formatted_liters": formatted_liters,
            "avg_liters_per_hour": round(avg_lh, 2) if avg_lh is not None else None,
            "formatted_avg_lh": formatted_avg_lh,
            "first_cumulative": first_cum,
            "last_cumulative": last_cum,
            "resets_count": resets_count,
            "samples_count": len(valid_events),
            "hours_covered": round(hours_covered, 2),
            "status_message": f"Volume total calculado por diferenças consecutivas ({resets_count} reset(s) detectado(s))." if resets_count > 0 else "Cálculo contínuo sem resets.",
            "hourly_series": hourly_series,
            "deltas_by_ts": {d["ts"]: d["delta"] for d in deltas_list},
        }

    @classmethod
    def get_synchronized_history(
        cls,
        start_dt: datetime,
        end_dt: datetime,
    ) -> Dict[str, Any]:
        """
        Executa a consulta histórica indexada e executa o algoritmo de forward-fill temporal.
        Calcula todas as estatísticas analíticas de vapor, setpoint, desvios e condensado.
        """
        start_ms = int(start_dt.timestamp() * 1000)
        end_ms = int(end_dt.timestamp() * 1000)

        resolved_dps = cls.resolve_caldeira_datapoints()
        dp_id_to_key: Dict[int, str] = {}
        active_dp_ids: List[int] = []
        variables_missing: List[str] = []

        for key, info in resolved_dps.items():
            if info["dp_id"]:
                dp_id_to_key[info["dp_id"]] = key
                active_dp_ids.append(info["dp_id"])
            else:
                variables_missing.append(info["config"]["label"])

        if not active_dp_ids:
            empty_charts = cls._empty_chart_datasets()
            return {
                "timeline": [],
                "raw_points_count": 0,
                "variables_found_count": 0,
                "variables_missing": variables_missing,
                "chart_datasets": empty_charts,
                "stats": {},
                "card_stats": {},
                "condensate_stats": cls.compute_condensate_volume([], start_ms, end_ms),
                "start_dt": start_dt,
                "end_dt": end_dt,
                "has_data": False,
            }

        # 1. Obter estado inicial anterior ao start_ms para cada DP (seed do forward-fill)
        initial_state: Dict[str, Any] = {key: None for key in resolved_dps}
        seed_condensate_val: Optional[float] = None

        try:
            prior_max_ts_records = (
                ScadaPointValue.objects.using("scada")
                .filter(data_point_id__in=active_dp_ids, ts__lt=start_ms)
                .values("data_point_id")
                .annotate(max_ts=Max("ts"))
            )
            prior_map = {r["data_point_id"]: r["max_ts"] for r in prior_max_ts_records if r.get("max_ts")}
            if prior_map:
                q_prior = Q()
                for dp_id, max_ts in prior_map.items():
                    q_prior |= Q(data_point_id=dp_id, ts=max_ts)

                prior_values = (
                    ScadaPointValue.objects.using("scada")
                    .filter(q_prior)
                    .select_related("annotation")
                )
                for pv in prior_values:
                    v_key = dp_id_to_key.get(pv.data_point_id)
                    if v_key:
                        norm_val, _ = scada_reader.normalize_value(pv.data_type, pv.point_value, getattr(pv, "annotation", None))
                        initial_state[v_key] = norm_val
                        if v_key == "volume_condensado" and norm_val is not None:
                            try:
                                seed_condensate_val = float(norm_val)
                            except (ValueError, TypeError):
                                pass
        except Exception:
            pass

        # 2. Consultar registros no intervalo [start_ms, end_ms]
        try:
            point_values_qs = (
                ScadaPointValue.objects.using("scada")
                .filter(data_point_id__in=active_dp_ids, ts__gte=start_ms, ts__lte=end_ms)
                .select_related("annotation")
                .order_by("ts", "id")
            )
            raw_records = list(point_values_qs)
        except Exception:
            raw_records = []

        raw_points_count = len(raw_records)

        # 3. Algoritmo de Forward-Fill Temporal
        events_by_ts: Dict[int, Dict[str, Any]] = {}
        condensate_raw_events: List[Dict[str, Any]] = []

        for pv in raw_records:
            v_key = dp_id_to_key.get(pv.data_point_id)
            if not v_key:
                continue
            norm_val, _ = scada_reader.normalize_value(pv.data_type, pv.point_value, getattr(pv, "annotation", None))
            if pv.ts not in events_by_ts:
                events_by_ts[pv.ts] = {}
            events_by_ts[pv.ts][v_key] = norm_val

            if v_key == "volume_condensado":
                condensate_raw_events.append({"ts": pv.ts, "val": norm_val})

        current_state = initial_state.copy()
        timeline: List[Dict[str, Any]] = []

        # Se houver estado inicial mas nenhum evento no intervalo, emitir ponto no início da janela
        if not events_by_ts and any(v is not None for v in current_state.values()):
            row_dt = timezone.localtime(timezone.datetime.fromtimestamp(start_ms / 1000, tz=timezone.utc))
            timeline.append({
                "ts": start_ms,
                "datetime": row_dt,
                "datetime_str": row_dt.strftime("%d/%m/%Y %H:%M:%S"),
                "values": current_state.copy(),
            })

        for ts in sorted(events_by_ts.keys()):
            updates = events_by_ts[ts]
            current_state.update(updates)
            row_dt = timezone.localtime(timezone.datetime.fromtimestamp(ts / 1000, tz=timezone.utc))

            timeline.append({
                "ts": ts,
                "datetime": row_dt,
                "datetime_str": row_dt.strftime("%d/%m/%Y %H:%M:%S"),
                "values": current_state.copy(),
            })

        # 4. Cálculo de Condensado
        condensate_stats = cls.compute_condensate_volume(
            condensate_raw_events,
            start_ms=start_ms,
            end_ms=end_ms,
            seed_reading=seed_condensate_val
        )

        # 5. Cálculos Estatísticos Descritivos e Derivados
        stats = cls._compute_analytical_stats(timeline, condensate_stats)

        # 6. Preparar datasets para os 6 gráficos
        chart_datasets = cls._build_chart_datasets(timeline, condensate_stats)

        has_data = (raw_points_count > 0 or len(timeline) > 0)

        return {
            "timeline": timeline,
            "raw_points_count": raw_points_count,
            "variables_found_count": len(active_dp_ids),
            "variables_missing": variables_missing,
            "chart_datasets": chart_datasets,
            "stats": stats,
            "card_stats": stats.get("cards", {}),
            "condensate_stats": condensate_stats,
            "start_dt": start_dt,
            "end_dt": end_dt,
            "has_data": has_data,
        }

    @classmethod
    def _compute_analytical_stats(
        cls,
        timeline: List[Dict[str, Any]],
        condensate_stats: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Calcula estatísticas de variáveis instantâneas, desvios em relação ao setpoint,
        diferenças entre linhas e diferenças Caldeira × Linhas de alta.
        """
        if not timeline:
            return {"cards": {}, "variables": {}, "deviations": {}, "comparisons": {}}

        def _get_num(val: Any) -> Optional[float]:
            if val is None or str(val).strip() == "":
                return None
            try:
                f = float(val)
                return f if not math.isnan(f) and not math.isinf(f) else None
            except (ValueError, TypeError):
                return None

        # Estatísticas por variável individual
        var_stats: Dict[str, Dict[str, Any]] = {}
        for var_cfg in CALDEIRA_VARIABLES_CONFIG:
            v_key = var_cfg["key"]
            nums: List[float] = []
            first_val = None
            last_val = None
            first_ts = None
            last_ts = None

            for row in timeline:
                val = _get_num(row["values"].get(v_key))
                if val is not None:
                    nums.append(val)
                    if first_val is None:
                        first_val = val
                        first_ts = row["ts"]
                    last_val = val
                    last_ts = row["ts"]

            if nums:
                avg_v = sum(nums) / len(nums)
                min_v = min(nums)
                max_v = max(nums)
                amp_v = max_v - min_v
                var_stats[v_key] = {
                    "count": len(nums),
                    "avg": round(avg_v, 2),
                    "min": round(min_v, 2),
                    "max": round(max_v, 2),
                    "amplitude": round(amp_v, 2),
                    "first_val": first_val,
                    "last_val": last_val,
                    "first_ts": first_ts,
                    "last_ts": last_ts,
                    "has_data": True,
                }
            else:
                var_stats[v_key] = {
                    "count": 0,
                    "avg": None,
                    "min": None,
                    "max": None,
                    "amplitude": None,
                    "first_val": None,
                    "last_val": None,
                    "first_ts": None,
                    "last_ts": None,
                    "has_data": False,
                }

        # Cálculos Amostra a Amostra: Desvios de Setpoint e Diferenças
        desvios_1_7 = []
        desvios_8_12 = []
        dif_alta = []
        dif_baixa = []
        dif_cald_1 = []
        dif_cald_2 = []

        for row in timeline:
            vals = row["values"]
            p1 = _get_num(vals.get("pressao_alta_prensas_1_7"))
            p2 = _get_num(vals.get("pressao_alta_prensas_8_12"))
            sp = _get_num(vals.get("setpoint_pressao_alta"))
            pcal = _get_num(vals.get("pressao_caldeira"))
            pbx1 = _get_num(vals.get("pressao_baixa_prensas_1_7"))
            pbx2 = _get_num(vals.get("pressao_baixa_prensas_8_12"))

            # Desvio Setpoint
            if p1 is not None and sp is not None:
                d1 = p1 - sp
                desvios_1_7.append(d1)
                row["desvio_1_7"] = round(d1, 2)
            else:
                row["desvio_1_7"] = None

            if p2 is not None and sp is not None:
                d2 = p2 - sp
                desvios_8_12.append(d2)
                row["desvio_8_12"] = round(d2, 2)
            else:
                row["desvio_8_12"] = None

            # Diferença entre linhas de alta
            if p1 is not None and p2 is not None:
                da = p1 - p2
                dif_alta.append(da)
                row["diferenca_alta"] = round(da, 2)
            else:
                row["diferenca_alta"] = None

            # Diferença entre linhas de baixa
            if pbx1 is not None and pbx2 is not None:
                db = pbx1 - pbx2
                dif_baixa.append(db)
                row["diferenca_baixa"] = round(db, 2)
            else:
                row["diferenca_baixa"] = None

            # Diferença Caldeira × Linhas de Alta
            if pcal is not None and p1 is not None:
                dc1 = pcal - p1
                dif_cald_1.append(dc1)
                row["diferenca_caldeira_1"] = round(dc1, 2)
            else:
                row["diferenca_caldeira_1"] = None

            if pcal is not None and p2 is not None:
                dc2 = pcal - p2
                dif_cald_2.append(dc2)
                row["diferenca_caldeira_2"] = round(dc2, 2)
            else:
                row["diferenca_caldeira_2"] = None

        def _calc_series_stats(arr: List[float]) -> Dict[str, Any]:
            if not arr:
                return {"count": 0, "avg_signed": None, "avg_abs": None, "max_abs": None, "min": None, "max": None}
            abs_arr = [abs(x) for x in arr]
            return {
                "count": len(arr),
                "avg_signed": round(sum(arr) / len(arr), 2),
                "avg_abs": round(sum(abs_arr) / len(abs_arr), 2),
                "max_abs": round(max(abs_arr), 2),
                "min": round(min(arr), 2),
                "max": round(max(arr), 2),
            }

        deviations = {
            "linha_1_7": _calc_series_stats(desvios_1_7),
            "linha_8_12": _calc_series_stats(desvios_8_12),
        }

        comparisons = {
            "alta": _calc_series_stats(dif_alta),
            "baixa": _calc_series_stats(dif_baixa),
            "caldeira_linha_1": _calc_series_stats(dif_cald_1),
            "caldeira_linha_2": _calc_series_stats(dif_cald_2),
        }

        # Cards Gerenciais
        cards = {
            "pressao_caldeira": {
                "avg": var_stats["pressao_caldeira"]["avg"],
                "min": var_stats["pressao_caldeira"]["min"],
                "max": var_stats["pressao_caldeira"]["max"],
                "has_data": var_stats["pressao_caldeira"]["has_data"],
            },
            "pressao_alta_1_7": {
                "avg": var_stats["pressao_alta_prensas_1_7"]["avg"],
                "min": var_stats["pressao_alta_prensas_1_7"]["min"],
                "max": var_stats["pressao_alta_prensas_1_7"]["max"],
                "desvio_abs_medio": deviations["linha_1_7"]["avg_abs"],
                "has_data": var_stats["pressao_alta_prensas_1_7"]["has_data"],
            },
            "pressao_alta_8_12": {
                "avg": var_stats["pressao_alta_prensas_8_12"]["avg"],
                "min": var_stats["pressao_alta_prensas_8_12"]["min"],
                "max": var_stats["pressao_alta_prensas_8_12"]["max"],
                "desvio_abs_medio": deviations["linha_8_12"]["avg_abs"],
                "has_data": var_stats["pressao_alta_prensas_8_12"]["has_data"],
            },
            "setpoint_pressao_alta": {
                "avg": var_stats["setpoint_pressao_alta"]["avg"],
                "min": var_stats["setpoint_pressao_alta"]["min"],
                "max": var_stats["setpoint_pressao_alta"]["max"],
                "has_data": var_stats["setpoint_pressao_alta"]["has_data"],
            },
            "condensado": {
                "total_liters": condensate_stats.get("total_liters"),
                "formatted_liters": condensate_stats.get("formatted_liters"),
                "avg_liters_per_hour": condensate_stats.get("avg_liters_per_hour"),
                "formatted_avg_lh": condensate_stats.get("formatted_avg_lh"),
                "resets_count": condensate_stats.get("resets_count", 0),
                "has_data": condensate_stats.get("has_data", False),
            },
            "ar_comprimido": {
                "avg": var_stats["pressao_ar_comprimido"]["avg"],
                "min": var_stats["pressao_ar_comprimido"]["min"],
                "max": var_stats["pressao_ar_comprimido"]["max"],
                "has_data": var_stats["pressao_ar_comprimido"]["has_data"],
            },
            "vacuo": {
                "avg": var_stats["pressao_vacuo"]["avg"],
                "min": var_stats["pressao_vacuo"]["min"],
                "max": var_stats["pressao_vacuo"]["max"],
                "has_data": var_stats["pressao_vacuo"]["has_data"],
            },
        }

        return {
            "variables": var_stats,
            "deviations": deviations,
            "comparisons": comparisons,
            "cards": cards,
        }

    @classmethod
    def _empty_chart_datasets(cls) -> Dict[str, Any]:
        return {
            "chart_1_vapor": {"labels": [], "timestamps": [], "pressao_caldeira": [], "setpoint": [], "alta_1_7": [], "alta_8_12": []},
            "chart_2_setpoint_desvio": {"labels": [], "timestamps": [], "desvio_1_7": [], "desvio_8_12": []},
            "chart_3_baixa": {"labels": [], "timestamps": [], "baixa_1_7": [], "baixa_8_12": []},
            "chart_4_condensado_hora": {"labels": [], "volumes": []},
            "chart_5_ar_comprimido": {"labels": [], "timestamps": [], "ar_comprimido": []},
            "chart_6_vacuo": {"labels": [], "timestamps": [], "vacuo": []},
        }

    @classmethod
    def _build_chart_datasets(
        cls,
        timeline: List[Dict[str, Any]],
        condensate_stats: Dict[str, Any],
        max_chart_points: int = 1500
    ) -> Dict[str, Any]:
        """
        Monta datasets para renderização via Chart.js nos 6 gráficos solicitados.
        Aplica downsampling preservando pontos com dados.
        """
        if not timeline:
            empty = cls._empty_chart_datasets()
            empty["chart_4_condensado_hora"] = {
                "labels": [h["hour"] for h in condensate_stats.get("hourly_series", [])],
                "volumes": [h["volume"] for h in condensate_stats.get("hourly_series", [])],
            }
            return empty

        total = len(timeline)
        sampled_timeline: List[Dict[str, Any]] = []

        if total <= max_chart_points:
            sampled_timeline = timeline
        else:
            step = total / max_chart_points
            for i in range(max_chart_points):
                idx = min(int(i * step), total - 1)
                sampled_timeline.append(timeline[idx])
            sampled_timeline.sort(key=lambda x: x["ts"])

        labels = [item["datetime_str"] for item in sampled_timeline]
        timestamps = [item["ts"] for item in sampled_timeline]

        def _get_float_val(item: Dict[str, Any], key: str) -> Optional[float]:
            v = item["values"].get(key)
            if v is None or str(v).strip() == "":
                return None
            try:
                f = float(v)
                return round(f, 2) if not math.isnan(f) and not math.isinf(f) else None
            except (ValueError, TypeError):
                return None

        # Séries Gráfico 1: Vapor
        p_caldeira = [_get_float_val(item, "pressao_caldeira") for item in sampled_timeline]
        p_setpoint = [_get_float_val(item, "setpoint_pressao_alta") for item in sampled_timeline]
        p_alta_1_7 = [_get_float_val(item, "pressao_alta_prensas_1_7") for item in sampled_timeline]
        p_alta_8_12 = [_get_float_val(item, "pressao_alta_prensas_8_12") for item in sampled_timeline]

        # Séries Gráfico 2: Desvios em relação ao Setpoint
        desv_1_7 = [item.get("desvio_1_7") for item in sampled_timeline]
        desv_8_12 = [item.get("desvio_8_12") for item in sampled_timeline]

        # Séries Gráfico 3: Linhas de Baixa
        p_baixa_1_7 = [_get_float_val(item, "pressao_baixa_prensas_1_7") for item in sampled_timeline]
        p_baixa_8_12 = [_get_float_val(item, "pressao_baixa_prensas_8_12") for item in sampled_timeline]

        # Séries Gráfico 4: Condensado por Hora
        hourly_cond = condensate_stats.get("hourly_series", [])
        cond_labels = [h["hour"] for h in hourly_cond]
        cond_vols = [h["volume"] for h in hourly_cond]

        # Séries Gráfico 5: Ar Comprimido
        p_ar = [_get_float_val(item, "pressao_ar_comprimido") for item in sampled_timeline]

        # Séries Gráfico 6: Vácuo
        p_vacuo = [_get_float_val(item, "pressao_vacuo") for item in sampled_timeline]

        return {
            "chart_1_vapor": {
                "labels": labels,
                "timestamps": timestamps,
                "pressao_caldeira": p_caldeira,
                "setpoint": p_setpoint,
                "alta_1_7": p_alta_1_7,
                "alta_8_12": p_alta_8_12,
            },
            "chart_2_setpoint_desvio": {
                "labels": labels,
                "timestamps": timestamps,
                "desvio_1_7": desv_1_7,
                "desvio_8_12": desv_8_12,
            },
            "chart_3_baixa": {
                "labels": labels,
                "timestamps": timestamps,
                "baixa_1_7": p_baixa_1_7,
                "baixa_8_12": p_baixa_8_12,
            },
            "chart_4_condensado_hora": {
                "labels": cond_labels,
                "volumes": cond_vols,
            },
            "chart_5_ar_comprimido": {
                "labels": labels,
                "timestamps": timestamps,
                "ar_comprimido": p_ar,
            },
            "chart_6_vacuo": {
                "labels": labels,
                "timestamps": timestamps,
                "vacuo": p_vacuo,
            },
        }

    @classmethod
    def generate_excel_report(
        cls,
        start_dt: datetime,
        end_dt: datetime,
        generated_by: str = "Sistema",
    ) -> bytes:
        """
        Gera arquivo Excel (.xlsx) profissional contendo 3 abas:
        - Aba 1: Resumo Gerencial (indicadores, logo Freedom, notas, pronto para impressão A4 paisagem)
        - Aba 2: Gráficos (gráficos nativos do Excel para impressão)
        - Aba 3: Dados Históricos (tabela completa com auto-filtro e congelamento de painéis)
        """
        history = cls.get_synchronized_history(start_dt, end_dt)
        timeline = history["timeline"]
        stats = history["stats"]
        condensate_stats = history["condensate_stats"]
        var_stats = stats.get("variables", {})
        deviations = stats.get("deviations", {})
        comparisons = stats.get("comparisons", {})

        wb = openpyxl.Workbook()

        # Estilos Globais
        font_main_title = Font(name="Segoe UI", size=15, bold=True, color="0F172A")
        font_sub_title = Font(name="Segoe UI", size=10, italic=True, color="475569")
        font_section = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_table_hdr = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
        font_data_bold = Font(name="Segoe UI", size=9, bold=True, color="0F172A")
        font_data = Font(name="Segoe UI", size=9, color="1E293B")
        font_data_secondary = Font(name="Segoe UI", size=8, color="64748B")
        font_note = Font(name="Segoe UI", size=8, italic=True, color="475569")

        fill_navy = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        fill_gold = PatternFill(start_color="B8842E", end_color="B8842E", fill_type="solid")
        fill_header_gray = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        fill_kpi_label = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        border_thin = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )
        border_top_thick = Border(top=Side(style="medium", color="0F172A"))

        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Localizar Logo Oficial Freedom
        logo_path = os.path.join(settings.BASE_DIR, "maintenance", "static", "maintenance", "img", "logo_pneus_freedom_black.png")
        if not os.path.exists(logo_path):
            logo_path = os.path.join(settings.BASE_DIR, "staticfiles", "maintenance", "img", "logo_pneus_freedom_black.png")

        # ══════════════════════════════════════════════════════════════════════
        # ABA 1: RESUMO GERENCIAL
        # ══════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Resumo Gerencial"
        ws1.views.sheetView[0].showGridLines = True

        # Inserir Logo ou Fallback
        row_cursor = 1
        has_logo_inserted = False
        if os.path.exists(logo_path):
            try:
                from openpyxl.drawing.image import Image as OpenPyXLImage
                img = OpenPyXLImage(logo_path)
                img.width = 150
                img.height = 42
                ws1.add_image(img, "A1")
                has_logo_inserted = True
            except Exception:
                has_logo_inserted = False

        if not has_logo_inserted:
            ws1.merge_cells("A1:B2")
            logo_cell = ws1.cell(row=1, column=1, value="PNEUS FREEDOM")
            logo_cell.font = Font(name="Segoe UI", size=14, bold=True, color="B8842E")
            logo_cell.alignment = align_center

        ws1.merge_cells("C1:H1")
        title_cell = ws1.cell(row=1, column=3, value="RELATÓRIO DA CALDEIRA 2 & DISTRIBUIÇÃO DE VAPOR")
        title_cell.font = font_main_title
        title_cell.alignment = align_left

        ws1.merge_cells("C2:H2")
        sub_cell = ws1.cell(row=2, column=3, value="Auditoria de Pressões de Processo, Setpoint, Linhas de Vulcanização, Condensado e Utilidades")
        sub_cell.font = font_sub_title
        sub_cell.alignment = align_left

        row_cursor = 4

        # Bloco de Metadados da Emissão
        ws1.merge_cells(f"A{row_cursor}:H{row_cursor}")
        meta_hdr = ws1.cell(row=row_cursor, column=1, value="1. DADOS DA CONSULTA E METADADOS DO RELATÓRIO")
        meta_hdr.font = font_section
        meta_hdr.fill = fill_navy
        meta_hdr.alignment = align_left

        row_cursor += 1
        meta_rows = [
            ("Período Analisado:", f"{start_dt.strftime('%d/%m/%Y %H:%M')} até {end_dt.strftime('%d/%m/%Y %H:%M')} (Horário de Brasília)", "Usuário Responsável:", generated_by),
            ("Data e Hora da Geração:", timezone.localtime().strftime("%d/%m/%Y %H:%M:%S"), "Total de Amostras Sincronizadas:", f"{len(timeline)} registros"),
            ("Status dos Dados:", "Leituras válidas registradas" if history["has_data"] else "Não existem leituras para o período", "Cobertura de Condensado:", f"{condensate_stats.get('hours_covered', 0)} horas cobertas"),
        ]

        for label1, val1, label2, val2 in meta_rows:
            ws1.cell(row=row_cursor, column=1, value=label1).font = font_data_bold
            ws1.cell(row=row_cursor, column=1).fill = fill_kpi_label
            ws1.cell(row=row_cursor, column=1).border = border_thin

            ws1.merge_cells(f"B{row_cursor}:D{row_cursor}")
            c_val1 = ws1.cell(row=row_cursor, column=2, value=val1)
            c_val1.font = font_data
            c_val1.alignment = align_left
            for col in range(2, 5):
                ws1.cell(row=row_cursor, column=col).border = border_thin

            ws1.cell(row=row_cursor, column=5, value=label2).font = font_data_bold
            ws1.cell(row=row_cursor, column=5).fill = fill_kpi_label
            ws1.cell(row=row_cursor, column=5).border = border_thin

            ws1.merge_cells(f"F{row_cursor}:H{row_cursor}")
            c_val2 = ws1.cell(row=row_cursor, column=6, value=val2)
            c_val2.font = font_data
            c_val2.alignment = align_left
            for col in range(6, 9):
                ws1.cell(row=row_cursor, column=col).border = border_thin

            row_cursor += 1

        row_cursor += 1

        # Bloco de Indicadores Principais de Processo
        ws1.merge_cells(f"A{row_cursor}:H{row_cursor}")
        kpi_hdr = ws1.cell(row=row_cursor, column=1, value="2. RESUMO DOS INDICADORES OPERACIONAIS E ESTATÍSTICAS")
        kpi_hdr.font = font_section
        kpi_hdr.fill = fill_gold
        kpi_hdr.alignment = align_left

        row_cursor += 1
        headers_kpi = ["Variável / Sistema", "Unidade", "Média", "Mínimo", "Máximo", "Amplitude", "Desvio / Dif. Média", "Observações"]
        for c_idx, h_text in enumerate(headers_kpi, start=1):
            cell = ws1.cell(row=row_cursor, column=c_idx, value=h_text)
            cell.font = font_table_hdr
            cell.fill = fill_header_gray
            cell.alignment = align_center
            cell.border = border_thin

        row_cursor += 1

        def _fmt_f(val: Optional[float], decimals: int = 2) -> str:
            if val is None:
                return "-"
            return f"{val:.{decimals}f}".replace(".", ",")

        kpi_table_data = [
            # Geração de Vapor
            ("Pressão da Caldeira 2", "bar", var_stats.get("pressao_caldeira", {}).get("avg"), var_stats.get("pressao_caldeira", {}).get("min"), var_stats.get("pressao_caldeira", {}).get("max"), var_stats.get("pressao_caldeira", {}).get("amplitude"), "-", "Pressão medida na caldeira"),
            ("Pressão de alta configurada (Setpoint)", "bar", var_stats.get("setpoint_pressao_alta", {}).get("avg"), var_stats.get("setpoint_pressao_alta", {}).get("min"), var_stats.get("setpoint_pressao_alta", {}).get("max"), var_stats.get("setpoint_pressao_alta", {}).get("amplitude"), "-", "Meta regulada na válvula"),
            # Linhas de Alta
            ("Linha Alta — Prensas 1 a 7 (até 14 cavidades)", "bar", var_stats.get("pressao_alta_prensas_1_7", {}).get("avg"), var_stats.get("pressao_alta_prensas_1_7", {}).get("min"), var_stats.get("pressao_alta_prensas_1_7", {}).get("max"), var_stats.get("pressao_alta_prensas_1_7", {}).get("amplitude"), f"MAD: {_fmt_f(deviations.get('linha_1_7', {}).get('avg_abs'))} bar", "Desvio do Setpoint"),
            ("Linha Alta — Prensas 8 a 12 (até 18 cavidades)", "bar", var_stats.get("pressao_alta_prensas_8_12", {}).get("avg"), var_stats.get("pressao_alta_prensas_8_12", {}).get("min"), var_stats.get("pressao_alta_prensas_8_12", {}).get("max"), var_stats.get("pressao_alta_prensas_8_12", {}).get("amplitude"), f"MAD: {_fmt_f(deviations.get('linha_8_12', {}).get('avg_abs'))} bar", "Desvio do Setpoint"),
            # Comparação Alta
            ("Diferença entre Linhas de Alta (1-7 × 8-12)", "bar", comparisons.get("alta", {}).get("avg_signed"), comparisons.get("alta", {}).get("min"), comparisons.get("alta", {}).get("max"), "-", f"Abs: {_fmt_f(comparisons.get('alta', {}).get('avg_abs'))} bar", "Linhas com capacidades diferentes"),
            # Linhas de Baixa
            ("Linha Baixa — Prensas 1 a 7", "bar", var_stats.get("pressao_baixa_prensas_1_7", {}).get("avg"), var_stats.get("pressao_baixa_prensas_1_7", {}).get("min"), var_stats.get("pressao_baixa_prensas_1_7", {}).get("max"), var_stats.get("pressao_baixa_prensas_1_7", {}).get("amplitude"), "-", "Pressão de retorno/baixa"),
            ("Linha Baixa — Prensas 8 a 12", "bar", var_stats.get("pressao_baixa_prensas_8_12", {}).get("avg"), var_stats.get("pressao_baixa_prensas_8_12", {}).get("min"), var_stats.get("pressao_baixa_prensas_8_12", {}).get("max"), var_stats.get("pressao_baixa_prensas_8_12", {}).get("amplitude"), "-", "Pressão de retorno/baixa"),
            ("Diferença entre Linhas de Baixa", "bar", comparisons.get("baixa", {}).get("avg_signed"), comparisons.get("baixa", {}).get("min"), comparisons.get("baixa", {}).get("max"), "-", f"Abs: {_fmt_f(comparisons.get('baixa', {}).get('avg_abs'))} bar", "Diferença 1-7 × 8-12"),
            # Utilidades
            ("Pressão do Ar Comprimido (Vulcanização)", "bar", var_stats.get("pressao_ar_comprimido", {}).get("avg"), var_stats.get("pressao_ar_comprimido", {}).get("min"), var_stats.get("pressao_ar_comprimido", {}).get("max"), var_stats.get("pressao_ar_comprimido", {}).get("amplitude"), "-", "Conformação e descarga"),
            ("Pressão da Unidade de Vácuo", "bar", var_stats.get("pressao_vacuo", {}).get("avg"), var_stats.get("pressao_vacuo", {}).get("min"), var_stats.get("pressao_vacuo", {}).get("max"), var_stats.get("pressao_vacuo", {}).get("amplitude"), "-", "Drenagem e abertura do bladder"),
        ]

        for r_idx, (v_name, unit, avg_v, min_v, max_v, amp_v, dev_str, obs) in enumerate(kpi_table_data):
            fill_row = fill_zebra if r_idx % 2 == 1 else None

            ws1.cell(row=row_cursor, column=1, value=v_name).alignment = align_left
            ws1.cell(row=row_cursor, column=2, value=unit).alignment = align_center
            ws1.cell(row=row_cursor, column=3, value=avg_v if isinstance(avg_v, (int, float)) else "-").alignment = align_right
            ws1.cell(row=row_cursor, column=4, value=min_v if isinstance(min_v, (int, float)) else "-").alignment = align_right
            ws1.cell(row=row_cursor, column=5, value=max_v if isinstance(max_v, (int, float)) else "-").alignment = align_right
            ws1.cell(row=row_cursor, column=6, value=amp_v if isinstance(amp_v, (int, float)) else "-").alignment = align_right
            ws1.cell(row=row_cursor, column=7, value=dev_str).alignment = align_center
            ws1.cell(row=row_cursor, column=8, value=obs).alignment = align_left

            for col in range(1, 9):
                c = ws1.cell(row=row_cursor, column=col)
                c.font = font_data_bold if col == 1 else font_data
                c.border = border_thin
                if fill_row:
                    c.fill = fill_row
                if col in (3, 4, 5, 6) and isinstance(c.value, (int, float)):
                    c.number_format = "#,##0.00"

            row_cursor += 1

        row_cursor += 1

        # Bloco de Condensado
        ws1.merge_cells(f"A{row_cursor}:H{row_cursor}")
        cond_hdr = ws1.cell(row=row_cursor, column=1, value="3. VOLUME DE CONDENSADO MEDIDO NO PERÍODO")
        cond_hdr.font = font_section
        cond_hdr.fill = fill_navy
        cond_hdr.alignment = align_left

        row_cursor += 1
        cond_rows = [
            ("Volume Total do Período:", condensate_stats.get("formatted_liters", "Sem dados"), "Vazão Média Horária Derivada:", condensate_stats.get("formatted_avg_lh", "Sem dados")),
            ("Primeira Leitura Acumulada:", f"{condensate_stats.get('first_cumulative', '-')} L" if condensate_stats.get('first_cumulative') is not None else "-", "Última Leitura Acumulada:", f"{condensate_stats.get('last_cumulative', '-')} L" if condensate_stats.get('last_cumulative') is not None else "-"),
            ("Resets / Quebras Detectadas:", f"{condensate_stats.get('resets_count', 0)} ocorrência(s)", "Status do Totalizador:", condensate_stats.get("status_message", "-")),
        ]

        for label1, val1, label2, val2 in cond_rows:
            ws1.cell(row=row_cursor, column=1, value=label1).font = font_data_bold
            ws1.cell(row=row_cursor, column=1).fill = fill_kpi_label
            ws1.cell(row=row_cursor, column=1).border = border_thin

            ws1.merge_cells(f"B{row_cursor}:D{row_cursor}")
            c_val1 = ws1.cell(row=row_cursor, column=2, value=val1)
            c_val1.font = font_data_bold if "Total" in label1 else font_data
            c_val1.alignment = align_left
            for col in range(2, 5):
                ws1.cell(row=row_cursor, column=col).border = border_thin

            ws1.cell(row=row_cursor, column=5, value=label2).font = font_data_bold
            ws1.cell(row=row_cursor, column=5).fill = fill_kpi_label
            ws1.cell(row=row_cursor, column=5).border = border_thin

            ws1.merge_cells(f"F{row_cursor}:H{row_cursor}")
            c_val2 = ws1.cell(row=row_cursor, column=6, value=val2)
            c_val2.font = font_data_bold if "Vazão" in label2 else font_data
            c_val2.alignment = align_left
            for col in range(6, 9):
                ws1.cell(row=row_cursor, column=col).border = border_thin

            row_cursor += 1

        row_cursor += 1

        # Bloco de Notas e Contexto Industrial
        ws1.merge_cells(f"A{row_cursor}:H{row_cursor}")
        note_hdr = ws1.cell(row=row_cursor, column=1, value="4. NOTAS DE ENGENHARIA E CONTEXTO INDUSTRIAL")
        note_hdr.font = font_section
        note_hdr.fill = fill_header_gray
        note_hdr.alignment = align_left

        row_cursor += 1
        notes = [
            "• Oscilações de pressão podem ocorrer naturalmente quando várias prensas ou cavidades solicitam vapor simultaneamente.",
            "• As linhas possuem capacidades instaladas diferentes: até 14 cavidades nas prensas 1 a 7 e até 18 cavidades nas prensas 8 a 12.",
            "• Este relatório é descritivo e comparativo, não classificando automaticamente uma oscilação como falha ou defeito.",
            "• O volume de condensado é calculado exclusivamente pelas diferenças positivas consecutivas do totalizador acumulado VOLUME_CAL.",
        ]

        for n_text in notes:
            ws1.merge_cells(f"A{row_cursor}:H{row_cursor}")
            c_n = ws1.cell(row=row_cursor, column=1, value=n_text)
            c_n.font = font_note
            c_n.alignment = align_left
            for col in range(1, 9):
                ws1.cell(row=row_cursor, column=col).border = border_thin
            row_cursor += 1

        # Configurações de Impressão da Aba 1 (A4 Paisagem, 1 Página de Largura)
        ws1.page_setup.orientation = ws1.ORIENTATION_LANDSCAPE
        ws1.page_setup.paperSize = ws1.PAPERSIZE_A4
        ws1.page_setup.fitToWidth = 1
        ws1.page_setup.fitToHeight = 0
        ws1.sheet_properties.pageSetUpPr.fitToPage = True
        ws1.page_margins.left = 0.4
        ws1.page_margins.right = 0.4
        ws1.page_margins.top = 0.5
        ws1.page_margins.bottom = 0.5
        ws1.oddFooter.center.text = "Página &P de &N — Pneus Freedom • Relatório da Caldeira 2"

        # Ajuste de largura de colunas da Aba 1
        col_widths_ws1 = [32, 10, 12, 12, 12, 12, 20, 30]
        for idx, w in enumerate(col_widths_ws1, start=1):
            ws1.column_dimensions[get_column_letter(idx)].width = w

        # ══════════════════════════════════════════════════════════════════════
        # ABA 3: DADOS HISTÓRICOS (Criar antes dos gráficos para referências)
        # ══════════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet(title="Dados Históricos")
        ws3.views.sheetView[0].showGridLines = True

        raw_headers = [
            "Data e Hora",
            "Pressão Caldeira (bar)",
            "Setpoint Alta (bar)",
            "Alta Prensas 1-7 (bar)",
            "Alta Prensas 8-12 (bar)",
            "Desvio 1-7 × SP (bar)",
            "Desvio 8-12 × SP (bar)",
            "Dif. Linhas Alta (bar)",
            "Baixa Prensas 1-7 (bar)",
            "Baixa Prensas 8-12 (bar)",
            "Dif. Linhas Baixa (bar)",
            "Dif. Caldeira × 1-7 (bar)",
            "Dif. Caldeira × 8-12 (bar)",
            "Totalizador Condensado (L)",
            "Incremento Condensado (L)",
            "Ar Comprimido (bar)",
            "Unidade Vácuo (bar)",
            "Observações",
        ]

        ws3.row_dimensions[1].height = 28
        for col_idx, h_text in enumerate(raw_headers, start=1):
            c = ws3.cell(row=1, column=col_idx, value=h_text)
            c.font = font_table_hdr
            c.fill = fill_gold if col_idx in (1, 2, 3) else fill_navy
            c.alignment = align_center
            c.border = border_thin

        # Preenchimento das Linhas
        cond_deltas_map = condensate_stats.get("deltas_by_ts", {})

        for row_idx, item in enumerate(timeline, start=2):
            ws3.row_dimensions[row_idx].height = 18
            vals = item["values"]
            row_dt = item["datetime"]
            ts = item["ts"]

            # Coluna 1: Data e Hora
            c1 = ws3.cell(row=row_idx, column=1, value=row_dt.strftime("%d/%m/%Y %H:%M:%S"))
            c1.font = font_data
            c1.alignment = align_center
            c1.border = border_thin

            row_data_items = [
                (vals.get("pressao_caldeira"), True),
                (vals.get("setpoint_pressao_alta"), True),
                (vals.get("pressao_alta_prensas_1_7"), True),
                (vals.get("pressao_alta_prensas_8_12"), True),
                (item.get("desvio_1_7"), True),
                (item.get("desvio_8_12"), True),
                (item.get("diferenca_alta"), True),
                (vals.get("pressao_baixa_prensas_1_7"), True),
                (vals.get("pressao_baixa_prensas_8_12"), True),
                (item.get("diferenca_baixa"), True),
                (item.get("diferenca_caldeira_1"), True),
                (item.get("diferenca_caldeira_2"), True),
                (vals.get("volume_condensado"), True),
                (cond_deltas_map.get(ts, None), True),
                (vals.get("pressao_ar_comprimido"), True),
                (vals.get("pressao_vacuo"), True),
                ("OK" if item.get("desvio_1_7") is not None else "Parcial", False),
            ]

            for c_offset, (val_raw, is_num) in enumerate(row_data_items, start=2):
                cell = ws3.cell(row=row_idx, column=c_offset)
                cell.font = font_data
                cell.border = border_thin

                if val_raw is None or val_raw == "":
                    cell.value = "-"
                    cell.alignment = align_center
                elif is_num:
                    try:
                        num_f = round(float(val_raw), 2)
                        cell.value = num_f
                        cell.alignment = align_right
                        cell.number_format = "#,##0.00"
                    except (ValueError, TypeError):
                        cell.value = str(val_raw)
                        cell.alignment = align_left
                else:
                    cell.value = str(val_raw)
                    cell.alignment = align_left

        # Configurações de Usabilidade da Aba 3
        ws3.freeze_panes = "A2"
        if timeline:
            ws3.auto_filter.ref = ws3.dimensions
        ws3.print_title_rows = "1:1"
        ws3.page_setup.orientation = ws3.ORIENTATION_LANDSCAPE
        ws3.page_setup.paperSize = ws3.PAPERSIZE_A4
        ws3.oddFooter.center.text = "Página &P de &N — Pneus Freedom • Histórico da Caldeira 2"

        # Ajuste de largura da Aba 3
        for col in ws3.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws3.column_dimensions[col_letter].width = max(max_len + 3, 13)

        # ══════════════════════════════════════════════════════════════════════
        # ABA 2: GRÁFICOS (Gráficos Nativos do Excel)
        # ══════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet(title="Gráficos")
        ws2.views.sheetView[0].showGridLines = True

        ws2.cell(row=1, column=1, value="PAINEL DE GRÁFICOS OPERACIONAIS DA CALDEIRA 2").font = font_main_title
        ws2.cell(row=2, column=1, value=f"Período: {start_dt.strftime('%d/%m/%Y %H:%M')} até {end_dt.strftime('%d/%m/%Y %H:%M')}").font = font_sub_title

        if len(timeline) >= 2:
            num_rows = len(timeline) + 1

            # Gráfico 1: Pressão Caldeira, Setpoint e Linhas de Alta
            chart1 = LineChart()
            chart1.title = "1. Pressão da Caldeira, Setpoint e Linhas de Alta (bar)"
            chart1.style = 13
            chart1.y_axis.title = "Pressão (bar)"
            chart1.x_axis.title = "Data e Hora"
            chart1.width = 22
            chart1.height = 12

            data_ref1 = Reference(ws3, min_col=2, min_row=1, max_col=5, max_row=min(num_rows, 500))
            cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=min(num_rows, 500))
            chart1.add_data(data_ref1, titles_from_data=True)
            chart1.set_categories(cats_ref)
            ws2.add_chart(chart1, "A4")

            # Gráfico 2: Linhas de Baixa
            chart2 = LineChart()
            chart2.title = "2. Pressão das Linhas de Baixa (bar)"
            chart2.style = 10
            chart2.y_axis.title = "Pressão (bar)"
            chart2.x_axis.title = "Data e Hora"
            chart2.width = 22
            chart2.height = 12

            data_ref2 = Reference(ws3, min_col=9, min_row=1, max_col=10, max_row=min(num_rows, 500))
            chart2.add_data(data_ref2, titles_from_data=True)
            chart2.set_categories(cats_ref)
            ws2.add_chart(chart2, "A29")

            # Gráfico 3: Ar Comprimido e Vácuo
            chart3 = LineChart()
            chart3.title = "3. Utilidades Auxiliares: Ar Comprimido e Vácuo (bar)"
            chart3.style = 24
            chart3.y_axis.title = "Pressão (bar)"
            chart3.x_axis.title = "Data e Hora"
            chart3.width = 22
            chart3.height = 12

            data_ref3 = Reference(ws3, min_col=16, min_row=1, max_col=17, max_row=min(num_rows, 500))
            chart3.add_data(data_ref3, titles_from_data=True)
            chart3.set_categories(cats_ref)
            ws2.add_chart(chart3, "A54")

        else:
            ws2.cell(row=5, column=1, value="Não existem dados suficientes no período selecionado para gerar os gráficos.").font = font_note

        ws2.page_setup.orientation = ws2.ORIENTATION_LANDSCAPE
        ws2.page_setup.paperSize = ws2.PAPERSIZE_A4

        # Definir Aba 1 como selecionada ao abrir
        wb.active = ws1

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
